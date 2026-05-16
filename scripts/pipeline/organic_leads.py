"""
GenosAI Organic Leads Tracker
Creates/updates output/workbooks/GenosAI_OrganicLeads.xlsx
Run this script each time a new manually-sourced lead is added.
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r"c:/Users/deevansho/Desktop/GenosApollp clients")
OUT  = ROOT / "output" / "workbooks" / "GenosAI_OrganicLeads.xlsx"

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill(fgColor="1F3864", patternType="solid")
EB5_FILL      = PatternFill(fgColor="FFF2CC", patternType="solid")   # gold — investment/legal
NEW_FILL      = PatternFill(fgColor="D7FFD7", patternType="solid")   # green — new / to contact
CONTACTED_F   = PatternFill(fgColor="DDEEFF", patternType="solid")   # blue — contacted
WON_FILL      = PatternFill(fgColor="E8D5FF", patternType="solid")   # purple — won
HIGH_FILL     = PatternFill(fgColor="FFD7D7", patternType="solid")   # red — high priority

HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
NORMAL_FONT  = Font(size=9)
WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
TOP_ALIGN    = Alignment(vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def h_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = HEADER_FILL
    c.font      = HEADER_FONT
    c.alignment = CENTER_ALIGN
    c.border    = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def d_cell(ws, row, col, value, fill=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = NORMAL_FONT
    c.border    = BORDER
    c.alignment = WRAP_ALIGN if wrap else TOP_ALIGN
    if fill:
        c.fill = fill
    return c


# ── Lead data ─────────────────────────────────────────────────────────────────
# Each dict = one organic lead row. Add new leads here and re-run the script.
LEADS = [
    {
        "num":          1,
        "company":      "FirstPathway Partners",
        "contact":      "Asif Chhipa",
        "title":        "Investment Specialist",
        "email":        "info@firstpathway.com",
        "phone":        "+1 (414) 431-0742",
        "website":      "https://www.firstpathway.com",
        "linkedin":     "https://www.linkedin.com/company/firstpathway-partners/",
        "city":         "Milwaukee",
        "state_country":"WI, USA",
        "industry":     "EB-5 Immigration Investment",
        "source":       "Manual Research — LinkedIn + Website",
        "status":       "To Contact",
        "audit_score":  "4/10",
        "deck_path":    "output/decks/FirstPathway_Audit_GenosAI.pptx",
        "automation_opportunities": (
            "1. Multilingual investor chatbot (Hindi/Mandarin/Arabic) on website + WhatsApp + WeChat\n"
            "2. Automated client journey emails: I-526E → DS-260 → I-829 milestones (3-5 yr lifecycle)\n"
            "3. HNI lead qualification engine — score inbound by net worth & timeline, route to Asif\n"
            "4. EB-5 ROI/residency benefit calculator with lead capture gate\n"
            "5. SEO content engine: 12 articles targeting EB-5 India/UAE/China search terms\n"
            "6. LinkedIn thought leadership automation (3 posts/week for HNI audience)\n"
            "7. Hindi/Mandarin monthly webinar series with auto registration + replay email\n"
            "8. WeChat Official Account AI bot for Chinese investor qualification"
        ),
        "notes": (
            "High-value prospect: $800K+ per investor. Small team (Asif Chhipa + Wilka Toppins, 34yr legal exp). "
            "Serves India, UAE, China, Europe. Active projects: Ameyalli (Utah thermal resort) + Kaia (Kanab, UT). "
            "Has WeChat + Weibo accounts but no AI behind them. "
            "India is now #1 EB-5 applicant country — huge multilingual AI opportunity. "
            "3-5 year client lifecycle = massive automation value per client."
        ),
        "email_subject": (
            "Your India investors are WhatsApping competitors at 10pm IST"
        ),
        "email_body": (
            "Hi Asif,\n\n"
            "I looked through firstpathway.com this week. The track record is strong — I-829 approvals, Ameyalli, Kaia. "
            "The digital infrastructure doesn't match it.\n\n"
            "Three things stood out:\n\n"
            "1. India is now the #1 EB-5 applicant country. Your website is English-only with a contact form. "
            "Indian HNI families browsing at 10pm IST are simultaneously WhatsApping 3-4 EB-5 firms. "
            "The one that responds in their language in under 5 minutes books the consultation.\n\n"
            "2. Your WeChat and Weibo accounts exist but have no AI agent behind them. "
            "Chinese investors — your historical top market — aren't getting qualified there.\n\n"
            "3. The EB-5 process runs 3-5 years. There are no automated touchpoints between filing milestones. "
            "Investors go quiet. So do you.\n\n"
            "We build AI systems that fix exactly this: multilingual investor agents (Hindi, Mandarin, Arabic), "
            "automated client journey comms from I-526E to I-829, and a lead qualification engine that routes "
            "$800K+ investors to you and handles the rest automatically.\n\n"
            "I put together a short audit deck — 10 slides, specific to FirstPathway. Happy to send it over or walk you through it on a 15-minute call.\n\n"
            "Worth a conversation?\n\n"
            "Rohan Malik\n"
            "CEO, Genos AI\n"
            "hello@genosai.tech | +91 638-714-2699\n"
            "www.genosai.tech"
        ),
        "followup_1": (
            "Hi Asif,\n\n"
            "Following up on my note from last week.\n\n"
            "One number that might be relevant: a single additional qualified EB-5 investor "
            "represents $800K in investment plus a 3-5 year advisory relationship. "
            "Our multilingual investor agent typically pays for itself within the first converted inquiry.\n\n"
            "The audit deck I mentioned is ready — covers exactly what's missing on the digital side "
            "and what we'd build, phase by phase.\n\n"
            "Want me to send it across?\n\n"
            "Rohan\n"
            "Genos AI | hello@genosai.tech"
        ),
        "followup_2": (
            "Hi Asif,\n\n"
            "Last note from me on this.\n\n"
            "If now isn't the right time — no problem. If you'd like the audit deck for a future conversation, "
            "just reply and I'll send it.\n\n"
            "Either way, good luck with Ameyalli and Kaia. The Utah projects look strong.\n\n"
            "Rohan\n"
            "Genos AI | hello@genosai.tech"
        ),
    },
    {
        "num":          2,
        "company":      "Herboths Property",
        "contact":      "Lizette da Fonseca",
        "title":        "Sales Contact",
        "email":        "",
        "phone":        "+264 81 142 8785",
        "website":      "http://herboths.com",
        "linkedin":     "https://www.linkedin.com/company/herboths-property/",
        "city":         "Windhoek",
        "state_country":"Namibia",
        "industry":     "Real Estate — Lifestyle Estate Development",
        "source":       "Manual Research — LinkedIn + Website",
        "status":       "To Contact",
        "audit_score":  "4/10",
        "deck_path":    "output/decks/Herboths_Audit_GenosAI.pptx",
        "automation_opportunities": (
            "1. WhatsApp AI estate agent — 24/7 inquiry response + site visit booking (primary Namibia channel)\n"
            "2. 6-touch lead nurture sequence over 120 days for warm buyers on slow decision cycle\n"
            "3. Buyer journey automation: reservation → docs → transfer milestones → HOA welcome\n"
            "4. Social content engine: LinkedIn + Facebook 3x/week (currently only 29 LinkedIn followers)\n"
            "5. SEO content: 8 articles targeting 'lifestyle estate Windhoek', 'freehold plots Namibia'\n"
            "6. Plot availability bot integrated with WhatsApp agent\n"
            "7. Post-sale referral automation via WhatsApp at 30 and 90 days\n"
            "8. SSL certificate fix (site currently shows security warning on HTTPS)"
        ),
        "notes": (
            "Small team (2-10 people). Flagship: Woodlands @ Herboths — 31 freehold plots, 3-8ha, 160ha total, "
            "20 min from Windhoek. Previous project Herboths Blick sold out (strong validation). "
            "Features: solar, independent water/sewerage, wildlife corridors, HOA. "
            "Contact via WhatsApp (+264 81 142 8785) — no public email found. "
            "WhatsApp is primary business channel in Namibia. 29 LinkedIn followers — very low for premium estate. "
            "SSL cert invalid on HTTPS — site only works on HTTP."
        ),
        "email_subject": (
            "Herboths Blick sold out. Woodlands buyers are WhatsApping competitors right now."
        ),
        "email_body": (
            "Hi Lizette,\n\n"
            "I looked at Woodlands @ Herboths this week — 31 plots, 160 hectares, solar-independent, "
            "wildlife corridors, 20 minutes from Windhoek. The product is genuinely differentiated.\n\n"
            "The digital infrastructure isn't keeping up with it.\n\n"
            "Three specific things:\n\n"
            "1. WhatsApp is the primary property channel in Namibia. A buyer browsing on Sunday afternoon "
            "will WhatsApp 3-4 options simultaneously. The one that responds in under 5 minutes with "
            "full information books the site visit. With a 2-10 person team, Herboths can't win that race manually.\n\n"
            "2. Plot purchases take 3-6 months to decide. Buyers who inquire and don't get nurtured go cold — "
            "and then convert for a competitor who showed up consistently in their feed. "
            "Herboths Blick sold out. The pipeline just needs to stay warm.\n\n"
            "3. Woodlands has 29 LinkedIn followers. The Windhoek professional who would buy a 5-hectare "
            "solar-independent plot is on LinkedIn every day — and has never seen a Herboths post.\n\n"
            "We build AI systems that fix this: a WhatsApp estate agent that responds 24/7, a 6-touch "
            "nurture sequence for slow deciders, and a social content engine that builds the audience "
            "Woodlands deserves.\n\n"
            "I've put together a short audit deck — 10 slides, specific to Herboths. "
            "Happy to send it over or take 15 minutes to walk through it.\n\n"
            "Worth a conversation?\n\n"
            "Rohan Malik\n"
            "CEO, Genos AI\n"
            "hello@genosai.tech | +91 638-714-2699\n"
            "www.genosai.tech"
        ),
        "followup_1": (
            "Hi Lizette,\n\n"
            "Following up on my note from last week.\n\n"
            "One thing worth flagging: herboths.com currently shows a security warning on "
            "HTTPS-enabled browsers — every buyer who types the URL sees it. "
            "That's a quick fix and something we'd address in Week 1 alongside the WhatsApp agent.\n\n"
            "With only 31 plots available, each site visit is a significant opportunity. "
            "A WhatsApp AI agent costs the same whether it handles 5 inquiries or 50 — "
            "and it works at 3am on a Saturday when buyers are scrolling.\n\n"
            "Audit deck is ready if you'd like to see it.\n\n"
            "Rohan\n"
            "Genos AI | hello@genosai.tech"
        ),
        "followup_2": (
            "Hi Lizette,\n\n"
            "Last message from me.\n\n"
            "If the timing isn't right, no problem — happy to reconnect when it is. "
            "If you'd like the audit deck for reference, just reply and I'll send it across.\n\n"
            "Good luck with Woodlands. The estate looks like a genuinely special product.\n\n"
            "Rohan\n"
            "Genos AI | hello@genosai.tech"
        ),
    },
]


def build_leads_sheet(wb):
    ws = wb.create_sheet("Organic Leads")
    ws.freeze_panes = "A2"

    headers = [
        "#", "Company", "Contact Person", "Title", "Email", "Phone",
        "Website", "LinkedIn", "City", "State/Country",
        "Industry", "Source", "Status", "Audit Score", "Deck Path",
        "Automation Opportunities", "Notes",
    ]
    widths = [4, 28, 22, 20, 30, 18, 35, 40, 14, 14, 22, 30, 14, 10, 40, 70, 60]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        h_cell(ws, 1, col, h, w)
    ws.row_dimensions[1].height = 22

    for lead in LEADS:
        r = lead["num"] + 1
        fill = EB5_FILL  # default; adjust per industry/status as needed

        d_cell(ws, r,  1, lead["num"],                    fill)
        d_cell(ws, r,  2, lead["company"],                fill)
        d_cell(ws, r,  3, lead["contact"],                fill)
        d_cell(ws, r,  4, lead["title"],                  fill)
        d_cell(ws, r,  5, lead["email"],                  fill)
        d_cell(ws, r,  6, lead["phone"],                  fill)
        d_cell(ws, r,  7, lead["website"],                fill)
        d_cell(ws, r,  8, lead["linkedin"],               fill)
        d_cell(ws, r,  9, lead["city"],                   fill)
        d_cell(ws, r, 10, lead["state_country"],          fill)
        d_cell(ws, r, 11, lead["industry"],               fill)
        d_cell(ws, r, 12, lead["source"],                 fill)
        d_cell(ws, r, 13, lead["status"],                 fill)
        d_cell(ws, r, 14, lead["audit_score"],            fill)
        d_cell(ws, r, 15, lead["deck_path"],              fill)
        d_cell(ws, r, 16, lead["automation_opportunities"], fill, wrap=True)
        d_cell(ws, r, 17, lead["notes"],                  fill, wrap=True)
        ws.row_dimensions[r].height = 100

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def build_email_sheet(wb):
    ws = wb.create_sheet("Cold Emails")
    ws.freeze_panes = "A2"

    headers = [
        "#", "Company", "Contact", "Email",
        "Subject", "Body", "Follow-Up 1", "Follow-Up 2",
    ]
    widths = [4, 28, 22, 30, 55, 90, 70, 70]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        h_cell(ws, 1, col, h, w)
    ws.row_dimensions[1].height = 22

    for lead in LEADS:
        if not lead.get("email_subject"):
            continue
        r = lead["num"] + 1
        d_cell(ws, r, 1, lead["num"],             EB5_FILL)
        d_cell(ws, r, 2, lead["company"],          EB5_FILL)
        d_cell(ws, r, 3, lead["contact"],          EB5_FILL)
        d_cell(ws, r, 4, lead["email"],            EB5_FILL)
        d_cell(ws, r, 5, lead["email_subject"],    EB5_FILL)
        d_cell(ws, r, 6, lead["email_body"],       EB5_FILL, wrap=True)
        d_cell(ws, r, 7, lead["followup_1"],       EB5_FILL, wrap=True)
        d_cell(ws, r, 8, lead["followup_2"],       EB5_FILL, wrap=True)
        ws.row_dimensions[r].height = 160

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def build_pipeline_sheet(wb):
    """Summary pipeline view with stage tracking."""
    ws = wb.create_sheet("Pipeline")
    ws.freeze_panes = "A2"

    headers = [
        "#", "Company", "Contact", "Email", "Industry",
        "Audit Score", "Status", "Last Touchpoint", "Next Action", "Est. Deal Value",
    ]
    widths = [4, 28, 22, 30, 22, 10, 14, 20, 40, 16]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        h_cell(ws, 1, col, h, w)
    ws.row_dimensions[1].height = 22

    for lead in LEADS:
        r = lead["num"] + 1
        d_cell(ws, r,  1, lead["num"],         EB5_FILL)
        d_cell(ws, r,  2, lead["company"],     EB5_FILL)
        d_cell(ws, r,  3, lead["contact"],     EB5_FILL)
        d_cell(ws, r,  4, lead["email"],       EB5_FILL)
        d_cell(ws, r,  5, lead["industry"],    EB5_FILL)
        d_cell(ws, r,  6, lead["audit_score"], EB5_FILL)
        d_cell(ws, r,  7, lead["status"],      EB5_FILL)
        next_actions = {
            1: "Send audit deck + intro email to info@firstpathway.com",
            2: "WhatsApp Lizette at +264 81 142 8785 — no public email; attach deck PDF",
        }
        deal_values = {
            1: "$15,000–$30,000 (AI automation retainer est.)",
            2: "$8,000–$18,000 (AI automation retainer est.)",
        }
        d_cell(ws, r,  8, "—",                              EB5_FILL)
        d_cell(ws, r,  9, next_actions.get(lead["num"], "—"), EB5_FILL, wrap=True)
        d_cell(ws, r, 10, deal_values.get(lead["num"], "—"),  EB5_FILL)
        ws.row_dimensions[r].height = 40

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_leads_sheet(wb)
    build_email_sheet(wb)
    build_pipeline_sheet(wb)

    wb.save(OUT)
    print(f"Saved -> {OUT}")
    print(f"  {len(LEADS)} organic lead(s) written")
    print(f"  Sheets: Organic Leads | Cold Emails | Pipeline")


if __name__ == "__main__":
    main()
