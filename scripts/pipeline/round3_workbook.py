"""
Round 3 workbook builder — creates GenosAI_RestaurantDentist_Leads.xlsx
with 4 sheets: Leads, Website Audits, Cold Emails, Audit PPT Content.

Rules:
- PPT content only for leads that HAVE a website.
- Email content only for leads that HAVE an email.
- All 100 leads appear in Leads + Website Audits.
"""
import json
from pathlib import Path
from copy import copy
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

ROOT      = Path(r"c:/Users/91638/Desktop/GenosApollp clients")
LEADS_F   = ROOT / "data" / "pipeline" / "_round3_leads.json"
AUDITS_F  = ROOT / "data" / "pipeline" / "_round3_audits.json"
CONTENT_F = ROOT / "data" / "pipeline" / "_round3_content.json"
OUT       = ROOT / "output" / "workbooks" / "GenosAI_RestaurantDentist_Leads.xlsx"

HEADER_FILL    = PatternFill(fgColor="1F3864", patternType="solid")
REST_FILL      = PatternFill(fgColor="FFF2CC", patternType="solid")   # yellow for restaurants
DENT_FILL      = PatternFill(fgColor="DDEEFF", patternType="solid")   # blue for dentists
NO_WEB_FILL    = PatternFill(fgColor="F2F2F2", patternType="solid")   # grey for no-website leads
HIGH_FILL      = PatternFill(fgColor="FFD7D7", patternType="solid")
MED_FILL       = PatternFill(fgColor="FFFACD", patternType="solid")
LOW_FILL       = PatternFill(fgColor="D7FFD7", patternType="solid")

HEADER_FONT    = Font(bold=True, color="FFFFFF", size=10)
NORMAL_FONT    = Font(size=9)
WRAP_ALIGN     = Alignment(wrap_text=True, vertical="top")
TOP_ALIGN      = Alignment(vertical="top")

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def h_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill   = HEADER_FILL
    c.font   = HEADER_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def d_cell(ws, row, col, value, fill=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font   = NORMAL_FONT
    c.border = BORDER
    c.alignment = WRAP_ALIGN if wrap else TOP_ALIGN
    if fill:
        c.fill = fill
    return c


def short(s, n=50):
    return (s or "")[:n]


def row_fill(lead):
    return REST_FILL if lead["category"] == "restaurant" else DENT_FILL


def build_leads_sheet(wb, leads, audits):
    ws = wb.create_sheet("Leads")
    ws.freeze_panes = "A2"

    headers = ["#","Category","Business Name","Email","Email Status","Website","Phone",
               "Address","City","State","Rating","Industry","Source"]
    widths  = [4,12,35,30,12,30,15,40,15,8,7,12,25]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        h_cell(ws, 1, col, h, w)

    ws.row_dimensions[1].height = 20

    for lead in leads:
        r    = lead["num"] - leads[0]["num"] + 2
        fill = row_fill(lead)
        a_entry = audits.get(str(lead["num"]), audits.get(lead["num"], {}))
        audit   = a_entry.get("audit")

        d_cell(ws, r, 1,  lead["num"],            fill)
        d_cell(ws, r, 2,  lead["category"].title(),fill)
        d_cell(ws, r, 3,  lead["company"],         fill)
        d_cell(ws, r, 4,  lead["email"],            fill)
        d_cell(ws, r, 5,  lead["email_status"],     fill)
        d_cell(ws, r, 6,  lead["website"],          fill)
        d_cell(ws, r, 7,  lead["phone"],            fill)
        d_cell(ws, r, 8,  lead["address"],          fill)
        d_cell(ws, r, 9,  lead["city"],             fill)
        d_cell(ws, r, 10, lead["state"],            fill)
        d_cell(ws, r, 11, lead["rating"],           fill)
        d_cell(ws, r, 12, lead["industry"],         fill)
        d_cell(ws, r, 13, lead["query_source"],     fill)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def build_audit_sheet(wb, leads, audits):
    ws = wb.create_sheet("Website Audits")
    ws.freeze_panes = "A2"

    headers = ["#","Category","Business","Website","Platform","Mobile","Analytics",
               "Booking","Online Order / Chat","Reviews","Email Mktg",
               "Score","Priority","Top Issues","Gaps / Opportunities"]
    widths  = [4,12,30,28,12,8,10,9,18,9,11,7,8,50,50]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        h_cell(ws, 1, col, h, w)
    ws.row_dimensions[1].height = 20

    for lead in leads:
        r       = lead["num"] - leads[0]["num"] + 2
        a_entry = audits.get(str(lead["num"]), audits.get(lead["num"], {}))
        audit   = a_entry.get("audit")
        cat     = lead["category"]

        if not lead["website"]:
            p_fill = NO_WEB_FILL
        elif audit:
            p_fill = HIGH_FILL if audit["priority"]=="HIGH" else MED_FILL if audit["priority"]=="MED" else LOW_FILL
        else:
            p_fill = MED_FILL

        d_cell(ws, r, 1,  lead["num"],              p_fill)
        d_cell(ws, r, 2,  cat.title(),              p_fill)
        d_cell(ws, r, 3,  short(lead["company"],35),p_fill)
        d_cell(ws, r, 4,  lead["website"],          p_fill)

        if audit:
            extra1 = audit.get("online_order","?") if cat=="restaurant" else audit.get("chat","?")
            d_cell(ws, r, 5,  audit["platform"],     p_fill)
            d_cell(ws, r, 6,  audit["mobile"],       p_fill)
            d_cell(ws, r, 7,  audit["analytics"],    p_fill)
            d_cell(ws, r, 8,  audit["booking"],      p_fill)
            d_cell(ws, r, 9,  extra1,                p_fill)
            d_cell(ws, r, 10, audit["reviews"],      p_fill)
            d_cell(ws, r, 11, audit["email_mkt"],    p_fill)
            d_cell(ws, r, 12, f"{audit['score']}/10",p_fill)
            d_cell(ws, r, 13, audit["priority"],     p_fill)
            d_cell(ws, r, 14, "\n".join(audit["issues"]), p_fill, wrap=True)
            d_cell(ws, r, 15, "\n".join(audit["gaps"]),   p_fill, wrap=True)
        elif not lead["website"]:
            for col in range(5, 16):
                d_cell(ws, r, col, "No website", p_fill)
        else:
            for col in range(5, 14):
                d_cell(ws, r, col, "Fetch failed", p_fill)
            d_cell(ws, r, 14, "Manual review needed", p_fill)
            d_cell(ws, r, 15, "Add AI booking + chatbot + review system", p_fill)

        ws.row_dimensions[r].height = 45

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def build_email_sheet(wb, leads, contents):
    ws = wb.create_sheet("Cold Emails")
    ws.freeze_panes = "A2"

    headers = ["#","Category","Business","Email","Subject","Body","Follow-Up 1","Follow-Up 2"]
    widths  = [4,12,30,30,45,80,60,60]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        h_cell(ws, 1, col, h, w)
    ws.row_dimensions[1].height = 20

    email_leads = [l for l in leads if l["email"]]
    no_email_leads = [l for l in leads if not l["email"]]

    for lead in email_leads:
        r    = lead["num"] - leads[0]["num"] + 2
        fill = row_fill(lead)
        c    = contents.get(str(lead["num"]), contents.get(lead["num"]))
        if not c:
            continue
        d_cell(ws, r, 1, lead["num"],              fill)
        d_cell(ws, r, 2, lead["category"].title(), fill)
        d_cell(ws, r, 3, short(lead["company"],35),fill)
        d_cell(ws, r, 4, lead["email"],            fill)
        d_cell(ws, r, 5, c["subject"],             fill)
        d_cell(ws, r, 6, c["body"],                fill, wrap=True)
        d_cell(ws, r, 7, c["fu1"],                 fill, wrap=True)
        d_cell(ws, r, 8, c["fu2"],                 fill, wrap=True)
        ws.row_dimensions[r].height = 100

    # grey rows for leads without email
    offset = len(email_leads)
    for j, lead in enumerate(no_email_leads):
        r = offset + j + 2
        d_cell(ws, r, 1, lead["num"],              NO_WEB_FILL)
        d_cell(ws, r, 2, lead["category"].title(), NO_WEB_FILL)
        d_cell(ws, r, 3, short(lead["company"],35),NO_WEB_FILL)
        d_cell(ws, r, 4, "Email not found",        NO_WEB_FILL)
        for col in range(5, 9):
            d_cell(ws, r, col, "-",                NO_WEB_FILL)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def build_ppt_sheet(wb, leads, contents, audits):
    ws = wb.create_sheet("Audit PPT Content")
    ws.freeze_panes = "A2"

    headers = ["#","Category","Business","Email","Website","Score","Priority",
               "Slide 1","Slide 2","Slide 3","Slide 4","Slide 5",
               "Slide 6","Slide 7","Slide 8","Slide 9"]
    widths  = [4,12,30,30,28,7,8] + [50]*9
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        h_cell(ws, 1, col, h, w)
    ws.row_dimensions[1].height = 20

    website_leads = [l for l in leads if l["website"]]

    for idx, lead in enumerate(website_leads):
        r       = idx + 2
        fill    = row_fill(lead)
        a_entry = audits.get(str(lead["num"]), audits.get(lead["num"], {}))
        audit   = a_entry.get("audit")
        c       = contents.get(str(lead["num"]), contents.get(lead["num"]))
        if not c:
            continue

        score_disp = f"{audit['score']}/10" if audit else "-"
        priority   = audit["priority"] if audit else "MED"

        d_cell(ws, r, 1, lead["num"],              fill)
        d_cell(ws, r, 2, lead["category"].title(), fill)
        d_cell(ws, r, 3, short(lead["company"],35),fill)
        d_cell(ws, r, 4, lead["email"] or "-",     fill)
        d_cell(ws, r, 5, lead["website"],          fill)
        d_cell(ws, r, 6, score_disp,               fill)
        d_cell(ws, r, 7, priority,                 fill)
        for i, slide_text in enumerate(c["slides"], 0):
            d_cell(ws, r, 8+i, slide_text, fill, wrap=True)
        ws.row_dimensions[r].height = 120

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def main():
    leads    = json.loads(LEADS_F.read_text(encoding="utf-8"))
    audits   = json.loads(AUDITS_F.read_text(encoding="utf-8"))
    contents = json.loads(CONTENT_F.read_text(encoding="utf-8"))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_leads_sheet(wb, leads, audits)
    build_audit_sheet(wb, leads, audits)
    build_email_sheet(wb, leads, contents)
    build_ppt_sheet(wb, leads, contents, audits)

    wb.save(OUT)

    restaurants = sum(1 for l in leads if l["category"]=="restaurant")
    dentists    = sum(1 for l in leads if l["category"]=="dentist")
    with_web    = sum(1 for l in leads if l["website"])
    with_email  = sum(1 for l in leads if l["email"])

    print(f"Saved -> {OUT}")
    print(f"  {len(leads)} leads total  |  {restaurants} restaurants  |  {dentists} dentists")
    print(f"  {with_web} with website (PPT rows)  |  {with_email} with email (Cold Email rows)")


if __name__ == "__main__":
    main()
