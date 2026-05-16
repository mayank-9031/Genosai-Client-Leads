"""
Round 2 workbook update:
  - Append 46 new lead rows to Leads, Website Audits, Cold Emails, Audit PPT Content.
  - Apply green fill to rows 2..44 (the round-1 block) on those 4 data sheets.
Saves in place to GenosAI_RealEstate_Leads (2).xlsx.
"""
import json, sys
from copy import copy
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent.parent
XLSX = ROOT / "output" / "workbooks" / "GenosAI_RealEstate_Leads (2).xlsx"

LEADS = ROOT / "data" / "pipeline" / "_round2_leads.json"
AUDITS = ROOT / "data" / "pipeline" / "_round2_audits.json"
CONTENT = ROOT / "data" / "pipeline" / "_round2_content.json"
DECK_PATHS = ROOT / "data" / "pipeline" / "_round2_deck_paths.json"

GREEN = PatternFill(fgColor="C6EFCE", patternType="solid")

def short_company(c):
    return c.split(",")[0].split("|")[0].strip()[:80]

def copy_row_style(ws, src_row, dst_row, max_col):
    """Copy styling (font, fill, alignment, border, number_format) from src_row to dst_row for cols 1..max_col."""
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.alignment = copy(src.alignment)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)

def main():
    if not XLSX.exists():
        print("xlsx missing", file=sys.stderr); sys.exit(1)
    try:
        wb = openpyxl.load_workbook(XLSX)
    except PermissionError:
        print("ERROR: workbook is locked. Close it in Excel and re-run.", file=sys.stderr)
        sys.exit(2)

    leads = json.loads(LEADS.read_text())
    audits = json.loads(AUDITS.read_text())
    contents = json.loads(CONTENT.read_text())
    deck_paths = json.loads(DECK_PATHS.read_text())

    # ---------- Sheet: Leads ----------
    ws = wb["Leads"]
    SRC_ROW = 44  # last data row of round 1 (header is 1, data 2..44)
    MAX_COL = 17  # A-Q
    for lead in leads:
        r = lead["num"] + 1  # data #44 -> row 45
        a = audits.get(lead["email"], {}).get("audit")
        ws.cell(row=r, column=1, value=lead["num"])
        ws.cell(row=r, column=2, value=lead["first_name"])
        ws.cell(row=r, column=3, value=lead["last_name"])
        ws.cell(row=r, column=4, value=lead["full_name"])
        ws.cell(row=r, column=5, value=lead["title"])
        ws.cell(row=r, column=6, value=short_company(lead["company"]))
        ws.cell(row=r, column=7, value=lead["email"])
        ws.cell(row=r, column=8, value=lead["email_status"])
        ws.cell(row=r, column=9, value=lead["website"])
        ws.cell(row=r, column=10, value=lead["phone"])
        ws.cell(row=r, column=11, value=lead["city"])
        ws.cell(row=r, column=12, value=lead["state"])
        ws.cell(row=r, column=13, value=lead["country"])
        ws.cell(row=r, column=14, value=lead["employees"])
        ws.cell(row=r, column=15, value=lead["industry"])
        ws.cell(row=r, column=16, value=lead["linkedin"])
        ws.cell(row=r, column=17, value=deck_paths.get(lead["email"], ""))
        copy_row_style(ws, SRC_ROW, r, MAX_COL)
    # Apply green fill to round-1 block (rows 2..44)
    for r in range(2, 45):
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r, column=c).fill = GREEN

    # ---------- Sheet: Website Audits ----------
    ws = wb["Website Audits"]
    SRC_ROW = 44
    MAX_COL = 16  # A-P
    for lead in leads:
        r = lead["num"] + 1
        a = audits.get(lead["email"], {}).get("audit")
        if a:
            issues_str = "\n".join(a["issues"])
            gaps_str = "\n".join(a["gaps"])
            row_vals = [
                lead["num"], short_company(lead["company"]), lead["website"],
                a["platform"], a["mobile"], a["analytics"], a["crm"],
                a["chat"], a["ai_tools"], a["cdn"], a["remarketing"], a["email_mkt"],
                a["score"], a["priority"], issues_str, gaps_str,
            ]
        else:
            row_vals = [
                lead["num"], short_company(lead["company"]), lead["website"],
                "Custom", "?", "?", "?", "?", "?", "?", "?", "?",
                "-", "MED",
                "Site fetch failed during audit - manual review needed",
                "AI chatbot for 24/7 lead qualification\nAI lead capture & CRM automation\nAutomated follow-up & nurture",
            ]
        for col, v in enumerate(row_vals, 1):
            ws.cell(row=r, column=col, value=v)
        copy_row_style(ws, SRC_ROW, r, MAX_COL)
    for r in range(2, 45):
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r, column=c).fill = GREEN

    # ---------- Sheet: Cold Emails ----------
    ws = wb["Cold Emails"]
    SRC_ROW = 44
    MAX_COL = 8  # A-H
    for lead in leads:
        r = lead["num"] + 1
        c = contents[lead["email"]]
        ws.cell(row=r, column=1, value=lead["num"])
        ws.cell(row=r, column=2, value=lead["first_name"])
        ws.cell(row=r, column=3, value=short_company(lead["company"]))
        ws.cell(row=r, column=4, value=lead["email"])
        ws.cell(row=r, column=5, value=c["subject"])
        ws.cell(row=r, column=6, value=c["body"])
        ws.cell(row=r, column=7, value=c["fu1"])
        ws.cell(row=r, column=8, value=c["fu2"])
        copy_row_style(ws, SRC_ROW, r, MAX_COL)
    for r in range(2, 45):
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r, column=c).fill = GREEN

    # ---------- Sheet: Audit PPT Content ----------
    ws = wb["Audit PPT Content"]
    SRC_ROW = 44
    MAX_COL = 16  # A-P (#, Company, Contact, Email, Website, Score, Priority, S1..S9)
    for lead in leads:
        r = lead["num"] + 1
        a = audits.get(lead["email"], {}).get("audit")
        c = contents[lead["email"]]
        score_disp = f"{a['score']}" if a else "-"
        priority = a["priority"] if a else "MED"
        ws.cell(row=r, column=1, value=lead["num"])
        ws.cell(row=r, column=2, value=short_company(lead["company"]))
        ws.cell(row=r, column=3, value=f"{lead['full_name']} ({lead['title']})")
        ws.cell(row=r, column=4, value=lead["email"])
        ws.cell(row=r, column=5, value=lead["website"])
        ws.cell(row=r, column=6, value=score_disp)
        ws.cell(row=r, column=7, value=priority)
        for i, slide_text in enumerate(c["slides"], 0):
            ws.cell(row=r, column=8 + i, value=slide_text)
        copy_row_style(ws, SRC_ROW, r, MAX_COL)
    for r in range(2, 45):
        for c in range(1, MAX_COL + 1):
            ws.cell(row=r, column=c).fill = GREEN

    try:
        wb.save(XLSX)
    except PermissionError:
        print("ERROR on save: workbook is locked. Close in Excel and re-run.", file=sys.stderr)
        sys.exit(2)
    print(f"Saved {XLSX} - 46 new rows appended to 4 sheets, green fill on rows 2..44.")

if __name__ == "__main__":
    main()
