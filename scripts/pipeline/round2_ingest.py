"""
Round 2 ingest: dedupe 46 new leads from 3 Apollo CSVs vs existing xlsx.
Writes scripts/_round2_leads.json with the cleaned + ordered lead list (#44-89).
"""
import csv, json, os, re, sys
from pathlib import Path
import openpyxl

ROOT = Path(__file__).parent.parent.parent
XLSX = ROOT / "output" / "workbooks" / "GenosAI_RealEstate_Leads (2).xlsx"
CSVS = [ROOT / "data" / "exports" / f"apollo-contacts-export ({n}).csv" for n in (2, 3, 4)]
OUT = ROOT / "data" / "pipeline" / "_round2_leads.json"

def norm(s):
    return (s or "").strip()

def strip_apollo_phone(p):
    p = norm(p)
    if p.startswith("'"):
        p = p[1:]
    return p.strip()

def slugify(s):
    s = re.sub(r"[^A-Za-z0-9]+", "_", s).strip("_")
    return s[:40]

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Leads"]
    existing = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[6]:
            existing.add(str(row[6]).strip().lower())

    seen = set()
    leads = []
    next_num = ws.max_row  # last data row number = 44 (header is 1, data 1..43 -> row 2..44)
    # existing data rows: row 2..44 -> #1..43; new start at #44 row 45
    next_idx = 44
    for csv_path in CSVS:
        with open(csv_path, encoding="utf-8", errors="replace") as fh:
            for r in csv.DictReader(fh):
                email = norm(r.get("Email", "")).lower()
                if not email or email in seen or email in existing:
                    continue
                if norm(r.get("Email Status", "")) != "Verified":
                    continue
                seen.add(email)
                phone = (
                    strip_apollo_phone(r.get("Mobile Phone", ""))
                    or strip_apollo_phone(r.get("Corporate Phone", ""))
                    or strip_apollo_phone(r.get("Work Direct Phone", ""))
                )
                first = norm(r.get("First Name", ""))
                last = norm(r.get("Last Name", ""))
                company = norm(r.get("Company Name", ""))
                lead = {
                    "num": next_idx,
                    "first_name": first,
                    "last_name": last,
                    "full_name": f"{first} {last}".strip(),
                    "title": norm(r.get("Title", "")),
                    "company": company,
                    "email": email,
                    "email_status": "Verified",
                    "website_raw": norm(r.get("Website", "")),
                    "phone": phone,
                    "city": norm(r.get("City", "")),
                    "state": norm(r.get("State", "")),
                    "country": norm(r.get("Country", "")),
                    "employees": norm(r.get("# Employees", "")),
                    "industry": norm(r.get("Industry", "")) or norm(r.get("Keywords", "")).split(",")[0].strip(),
                    "linkedin": norm(r.get("Person Linkedin Url", "")),
                    "company_linkedin": norm(r.get("Company Linkedin Url", "")),
                    "technologies": norm(r.get("Technologies", "")),
                    "keywords": norm(r.get("Keywords", "")),
                    "annual_revenue": norm(r.get("Annual Revenue", "")),
                    "slug": slugify(company) or slugify(first + "_" + last),
                }
                # normalize website (strip protocol for sheet display, keep full for fetching)
                w = lead["website_raw"]
                w = re.sub(r"^https?://", "", w, flags=re.I).rstrip("/")
                lead["website"] = w
                lead["website_url"] = lead["website_raw"] if lead["website_raw"].startswith("http") else f"https://{w}"
                leads.append(lead)
                next_idx += 1

    OUT.parent.mkdir(exist_ok=True)
    OUT.write_text(json.dumps(leads, indent=2))
    print(f"Wrote {len(leads)} leads to {OUT}")
    for l in leads[:3]:
        print(" ", l["num"], l["full_name"], "|", l["company"], "|", l["website"])

if __name__ == "__main__":
    main()
