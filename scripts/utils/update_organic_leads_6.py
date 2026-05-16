"""
Append 6 newly-audited leads (Eris, Befitting, AG Luxury, SIOR, BT Industries,
Tokuti) to GenosAI_OrganicLeads.xlsx — Organic Leads, Cold Emails and Pipeline
sheets. Mirrors the row layout used for FirstPathway and Herboths.
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from copy import copy

WB_PATH = Path(r"c:\Users\deevansho\Desktop\GenosApollp clients\output\workbooks\GenosAI_OrganicLeads.xlsx")

LEADS = [
    {
        # 3
        "company": "Eris Property Group",
        "contact": "General Enquiries",
        "title": "Marketing / Digital Lead (decision maker TBC)",
        "email": "info@eris.co.za",
        "phone": "+27 11 666 1300",
        "website": "https://www.eris.co.za",
        "linkedin": "https://www.linkedin.com/company/eris-property-group/",
        "city": "Johannesburg",
        "state": "South Africa (also GH, MU, MW)",
        "industry": "Commercial Real Estate — Development, Management & Asset Services",
        "source": "Manual Research — LinkedIn + Website",
        "status": "To Contact",
        "audit_score": "5/10",
        "deck_path": "Eris_WebDev_AI_Automation_Audit.pptx",
        "opps": (
            "1. AI chatbot on eris.co.za — 24/7 enquiry handling for tenants, investors and brokers\n"
            "2. Automated property listing pipeline: spec sheet → AI-written description → multi-portal publish\n"
            "3. POPIA + WCAG AA compliance sprint (cookie consent, data request forms, accessibility fixes)\n"
            "4. Next.js 15 rebuild + Sanity headless CMS — fixes mobile PageSpeed 48 → 90+\n"
            "5. Algolia faceted property search (type, location, size, price) + saved searches\n"
            "6. Investor/tenant portal: lease docs, payment history, maintenance tickets\n"
            "7. AI-generated weekly market intelligence reports across SA / Ghana / Mauritius / Malawi\n"
            "8. LinkedIn + social auto-posting for new listings & market updates\n"
            "9. Invoice processing & AP automation; AI contract review with clause flagging"
        ),
        "notes": (
            "Established Pan-African commercial real estate group (SA, Ghana, Mauritius, Malawi) "
            "spanning Development, Management, Asset, Facilities and Retail. Mobile PageSpeed 48/100, "
            "desktop 71/100, SEO 62/100 — partial WCAG AA, POPIA exposure on cookie consent. "
            "Listing search UX is dated vs. competitors. Strong digital revenue lever — Year-1 estimated "
            "savings/upside R 560 000–R 900 000 against R 530 000–R 865 000 spend. Property schema "
            "markup + AI listing engine is the highest-ROI Phase 1 move."
        ),
    },
    {
        # 4
        "company": "Befitting Group",
        "contact": "General Enquiries",
        "title": "Marketing / Sales Lead (decision maker TBC)",
        "email": "info@befittingproperties.com",
        "phone": "+234 (0) 1 700 7300",
        "website": "https://befittingproperties.com",
        "linkedin": "https://www.linkedin.com/company/befittinggroup/",
        "city": "Lagos",
        "state": "Nigeria (Pan-African)",
        "industry": "Real Estate — Development, Management & Investment",
        "source": "Manual Research — LinkedIn + Website + App Store",
        "status": "To Contact",
        "audit_score": "4/10",
        "deck_path": "Befitting_WebDev_AI_Automation_Audit.pptx",
        "opps": (
            "1. Multilingual WhatsApp + web AI agent — Nigeria's #1 property channel\n"
            "2. Diaspora investor journey: lead capture → KYC → off-plan progress emails (Nigeria → UK / US / UAE)\n"
            "3. Mobile PageSpeed urgent fix: 38 → 85+ (Next.js + image optimisation)\n"
            "4. SEO + schema markup sprint (currently ~55/100; thin content)\n"
            "5. AI listing engine — auto-write descriptions for residential/commercial/mixed-use stock\n"
            "6. Live social feed integration (Instagram @befittingproperties + Facebook) on site\n"
            "7. Mobile app (iOS) → AI-powered property recommendations + payment plan calculator\n"
            "8. Pan-African market dashboard: track Lagos, Abuja, Accra, Nairobi market signals weekly"
        ),
        "notes": (
            "12+ years in African real estate; Lagos HQ with Pan-African footprint. Has iOS app on App Store "
            "(strong asset). Digital health is weak: mobile PageSpeed 38/100 (urgent), SEO 55/100, no live "
            "social feed, minimal AI deployment beyond limited virtual tours. Diaspora investors (UK / US / "
            "UAE) are a high-value segment that needs WhatsApp + multilingual nurture — currently underserved. "
            "Large opportunity to be Africa's first AI-native real estate developer."
        ),
    },
    {
        # 5
        "company": "AG Luxury",
        "contact": "Sales Team",
        "title": "Sales / Founder (decision maker TBC)",
        "email": "Sales@agluxurydev.com",
        "phone": "+234 081 3446 5400",
        "website": "https://agluxurydev.com",
        "linkedin": "https://www.linkedin.com/company/ag-luxury-properties/",
        "city": "Lekki Phase 1, Lagos",
        "state": "Nigeria",
        "industry": "Luxury Residential Real Estate Development",
        "source": "Manual Research — Website + Social",
        "status": "To Contact",
        "audit_score": "5/10",
        "deck_path": "AGLuxury_WebDev_AI_Automation_Audit.pptx",
        "opps": (
            "1. AI concierge chatbot — luxury tone, 24/7, qualifies HNW buyers & diaspora\n"
            "2. 3D virtual tours + AI staging for AG Rosa units (currently absent — major luxury gap)\n"
            "3. Diaspora investor capture: USD/GBP pricing toggle, KYC, payment plan flows\n"
            "4. Premium brand consistency sweep across pages (Next.js stack already in place — great foundation)\n"
            "5. Live Instagram + TikTok feed embed (current zero integration; both channels active)\n"
            "6. SEO sprint — 'luxury homes Lekki', 'Lagos premium real estate' (currently ~58/100)\n"
            "7. AI-driven recommendation engine: match buyer profile to AG Rosa / Festus Court inventory\n"
            "8. Automated post-handover concierge journey for ultra-premium client retention"
        ),
        "notes": (
            "Premium luxury developer in Lekki Phase 1 with Next.js modern stack already in place — "
            "rare and a strong foundation. Flagship AG Rosa (current) + Festus Court (completed 2021). "
            "Targets affluent buyers, diaspora investors and HNW individuals. Digital gaps: no virtual "
            "tours / 3D (huge for luxury), no live social feed integration, mobile PageSpeed 52/100 "
            "(luxury needs 85+), SEO thin. Direct sales email and phone available — fast outreach path."
        ),
    },
    {
        # 6
        "company": "Society of Industrial and Office Realtors (SIOR)",
        "contact": "Marketing / Membership Team",
        "title": "Chief Marketing Officer / Membership Director (decision maker TBC)",
        "email": "info@sior.com",
        "phone": "+1 (202) 449-8200",
        "website": "https://www.sior.com",
        "linkedin": "https://www.linkedin.com/company/sior/",
        "city": "Washington",
        "state": "DC, USA (Global — 34 countries)",
        "industry": "Professional Association — Commercial Real Estate",
        "source": "Manual Research — Website + LinkedIn",
        "status": "To Contact",
        "audit_score": "6/10",
        "deck_path": "SIOR_WebDev_AI_Automation_Audit.pptx",
        "opps": (
            "1. MySIOR portal modernisation — replace dated Higher Logic UX with Next.js + SSO\n"
            "2. AI-powered member directory search across 4,000+ designees / 630+ cities\n"
            "3. Auto-generated weekly market intelligence reports from 80,000+ annual member transactions\n"
            "4. AI event concierge bot — Fall Event NYC (Oct 2026) + Lisbon International Conference\n"
            "5. SIOR Report magazine + Pulse Blog content engine: AI summaries, audio versions, segmentation\n"
            "6. Multilingual member onboarding journey (34 countries) with chapter-routed nurture\n"
            "7. Transaction submission automation — OCR + AI extraction from broker emails / PDFs\n"
            "8. Accessibility (WCAG AA) + schema markup sweep across sior.com\n"
            "9. Sponsor / partner ROI dashboard — auto-generated quarterly engagement reports"
        ),
        "notes": (
            "The world's foremost CRE professional association — 4,000+ designees, 630+ cities, 34 countries, "
            "~80,000 deals/year submitted by members. Strong content asset (SIOR Report magazine, Pulse Blog, "
            "weekly newsletter). Digital weak spots: MySIOR portal is dated (Higher Logic showing age), "
            "mobile PageSpeed 61/100, partial WCAG AA. Massive proprietary data set (80k annual transactions) "
            "= AI market-intelligence goldmine that no competitor association can replicate. Highest-prestige "
            "logo on our roster if won."
        ),
    },
    {
        # 7
        "company": "BT Industries",
        "contact": "Sales / Marketing Team",
        "title": "Sales / Marketing Director (decision maker TBC)",
        "email": "info@btindustries.co.za",
        "phone": "+27 11 552 8800",
        "website": "https://btindustries.co.za",
        "linkedin": "https://www.linkedin.com/company/bt-industries-africa/",
        "city": "Johannesburg",
        "state": "South Africa",
        "industry": "Industrial Distribution & Supply Chain — Automotive, Mining, Solar",
        "source": "Manual Research — LinkedIn + Website",
        "status": "To Contact",
        "audit_score": "5/10",
        "deck_path": "BTIndustries_WebDev_AI_Automation_Audit.pptx",
        "opps": (
            "1. B2B customer portal: OEM / mining customers see SKUs, stock, pricing tiers, order history\n"
            "2. AI fastener / part identifier — upload photo or paste spec → match BT SKU\n"
            "3. Solar-vertical lead engine for DoGo Power partnership (renewable energy entrants 2026)\n"
            "4. Quotation automation: RFQ email in → AI extracts parts → priced quote out in minutes\n"
            "5. ISO 9001 + B-BBEE Level 2 compliance content sprint (proof points underused on site)\n"
            "6. SAP / ERP integration with website inventory + customer credit limits\n"
            "7. AI catalogue search across fasteners, components & engineered parts (currently no search)\n"
            "8. LinkedIn thought-leadership engine: 65-year heritage, supply-chain resilience narrative\n"
            "9. Predictive stock-out alerts for top OEM customers (retention play)"
        ),
        "notes": (
            "65+ year heritage in South African industrial distribution — 'Confidence of Supply' positioning. "
            "Sectors: Automotive/OEM, Industrial, Aftermarket, Mining, plus new Renewable Energy / Solar via "
            "DoGo Power partnership (2026). ISO 9001 + Level 2 B-BBEE certified — strong B2B procurement "
            "credentials but website doesn't lead with them. Digital site is brochureware — no customer "
            "portal, no catalogue search, no quote automation. Big back-office automation upside given the "
            "long sales cycles and OEM-grade documentation requirements."
        ),
    },
    {
        # 8
        "company": "Tokuti (Tokuti.io)",
        "contact": "Mark Mariampillai",
        "title": "CEO & Founder",
        "email": "mark@tokuti.io",
        "phone": None,
        "website": "https://tokuti.io",
        "linkedin": "https://www.linkedin.com/in/markmariampillai/",
        "city": "Global (Web3)",
        "state": "UK / Global",
        "industry": "Web3 / PropTech — Tokenized Real Estate Auctions",
        "source": "Manual Research — Founder LinkedIn + Tokuti.io",
        "status": "To Contact",
        "audit_score": "5/10",
        "deck_path": "Tokuti_WebDev_Automation_Proposal.pptx",
        "opps": (
            "1. Move off Webflow → Next.js 15 + wagmi/viem for real Web3 wallet auth (MetaMask / WalletConnect)\n"
            "2. Real-time auction engine: WebSockets + smart-contract event listeners (Webflow can't do this)\n"
            "3. KYC / AML / accredited-investor flow via Sumsub or Onfido (institutional-grade compliance)\n"
            "4. Smart-contract auction logic (Solidity / EVM) with on-chain bid history\n"
            "5. Crypto + fiat payments: Stripe + USDC / ETH / BTC gateway integration\n"
            "6. AI deal-room: auto-generate property memos, legal summaries, jurisdictional access rules\n"
            "7. Investor CRM with on-chain wallet ID — track wallet → bid → KYC → deposit lifecycle\n"
            "8. GATENet ecosystem alignment (GATE token + staking) — unified investor identity across products\n"
            "9. Auction-launch traffic stack: Cloudflare + Vercel for 100× spike handling"
        ),
        "notes": (
            "Founder: Mark Mariampillai — 25 yrs finance background (Credit Suisse, HSBC). Building Tokuti "
            "Auction (tokenized real estate auctions) currently on Webflow MVP — perfect for validation, but "
            "Webflow cannot support Web3 wallet auth, smart-contract auction logic, real-time bidding, KYC/AML "
            "flows, or 100× traffic spikes. Linked GATENet token + staking ecosystem. Tokenized RWA market: "
            "$4T projected by 2030 (Deloitte), 50%+ CAGR, no dominant auction platform yet — strong first-mover "
            "window. Founder-led decision; relatively fast sales cycle once trust established."
        ),
    },
]

# ── Cold emails ───────────────────────────────────────────────────────────────
SIG = (
    "\n\nRohan Malik\n"
    "CEO, Genos AI\n"
    "hello@genosai.tech | +91 638-714-2699\n"
    "www.genosai.tech"
)
SIG_SHORT = "\n\nRohan\nGenos AI | hello@genosai.tech"

EMAILS = [
    # 3 — Eris
    {
        "subject": "eris.co.za loads in 4+ seconds on mobile — and POPIA is exposed",
        "body": (
            "Hi team,\n\n"
            "I spent time on eris.co.za this week. The portfolio is serious — development, asset, facilities, "
            "retail, four countries. The digital infrastructure isn't matching it.\n\n"
            "Three specific things:\n\n"
            "1. Mobile PageSpeed scores 48/100. Industry average is 2.5s page load; eris.co.za is over 4s. "
            "Brokers and tenants screening properties on a phone leave silently.\n\n"
            "2. POPIA cookie consent is missing. The Information Regulator has issued enforcement notices on "
            "exactly this in 2025. A 1-day fix that removes legal exposure.\n\n"
            "3. There is no AI on the site. Property enquiries get a contact form. Competitors are deploying "
            "chatbots that qualify in real time and route hot leads to the asset manager within seconds.\n\n"
            "We build AI systems that fix this: a chatbot trained on the Eris portfolio, an AI listing engine "
            "that writes descriptions from a spec sheet, and an investor/tenant portal that handles lease "
            "docs, maintenance and payments without staff time.\n\n"
            "I put together a 12-slide audit deck specific to Eris — PageSpeed targets, POPIA action items, "
            "Phase 1 roadmap with indicative ZAR pricing.\n\n"
            "Happy to send it across, or walk through it on a 15-minute call.\n\n"
            "Worth a conversation?"
            + SIG
        ),
        "fu1": (
            "Hi team,\n\n"
            "Following up on last week's note.\n\n"
            "One data point worth flagging: 73% of property enquirers convert with whoever responds first. "
            "On Eris's current setup, after-hours enquiries wait until the next business day. An AI agent "
            "trained on the Eris portfolio responds in under 2 seconds, 24/7, and books the viewing.\n\n"
            "The 12-slide audit deck is ready — covers exactly what's missing on the digital side, what "
            "we'd build, and Year-1 ROI vs spend.\n\n"
            "Want me to send it across?"
            + SIG_SHORT
        ),
        "fu2": (
            "Hi team,\n\n"
            "Last note from me.\n\n"
            "If now isn't the right time — no problem. If you'd like the audit deck for a future "
            "conversation, just reply and I'll send it over.\n\n"
            "Either way, good luck with the portfolio."
            + SIG_SHORT
        ),
    },
    # 4 — Befitting
    {
        "subject": "Mobile PageSpeed 38 in Lagos = diaspora buyers leaving befittingproperties.com",
        "body": (
            "Hi team,\n\n"
            "I went through befittingproperties.com this week. 12 years of African real estate, an iOS app, "
            "Pan-African operations — the foundation is genuinely strong.\n\n"
            "The website is leaking diaspora buyers.\n\n"
            "Three things:\n\n"
            "1. Mobile PageSpeed is 38/100. A diaspora investor in London, Dubai or Houston browsing on a "
            "3G connection waits 6+ seconds for the homepage. They don't wait — they tap back. The MTN and "
            "Airtel networks are not the bottleneck. The site is.\n\n"
            "2. There is no WhatsApp lead capture. WhatsApp is the #1 property channel in Nigeria. Diaspora "
            "buyers WhatsApp 3–4 developers simultaneously at 10pm local — the one that answers in their "
            "timezone with off-plan progress photos books the unit.\n\n"
            "3. The Instagram (@befittingproperties) and Facebook are active, but there's no live feed on "
            "the site. Buyers who land on the homepage from Google can't see the social proof Befitting is "
            "already creating.\n\n"
            "We build AI systems that fix this: a multilingual WhatsApp + web agent that responds 24/7 in "
            "English / Yoruba / Hausa / French, a diaspora investor journey from KYC to off-plan progress "
            "emails, and an AI listing engine that turns a spec sheet into a published unit in minutes.\n\n"
            "I put together a 12-slide audit deck specific to Befitting — PageSpeed targets, diaspora "
            "investor journey map, Phase 1 roadmap with indicative NGN/USD pricing.\n\n"
            "Happy to send it over or walk through it on a 15-minute call.\n\n"
            "Worth a conversation?"
            + SIG
        ),
        "fu1": (
            "Hi team,\n\n"
            "Following up on my note from last week.\n\n"
            "One thing worth flagging: a single diaspora off-plan unit can be 25–80M Naira. A WhatsApp AI "
            "agent that converts even one additional diaspora buyer per quarter pays for the entire Year-1 "
            "build several times over.\n\n"
            "The audit deck is ready if you'd like to see it."
            + SIG_SHORT
        ),
        "fu2": (
            "Hi team,\n\n"
            "Last message from me.\n\n"
            "If the timing isn't right, no problem. If you'd like the deck for reference, just reply.\n\n"
            "Good luck with the Pan-African expansion — the app is a great move."
            + SIG_SHORT
        ),
    },
    # 5 — AG Luxury
    {
        "subject": "AG Rosa has no virtual tour. Lekki HNW buyers expect one in 2026.",
        "body": (
            "Hi team,\n\n"
            "I looked at agluxurydev.com this week. The brand positioning is strong — Periwinkle, Lekki "
            "Phase 1, premium turnkey, Festus Court completed, AG Rosa current. Next.js is already in place, "
            "which is rare and shows real engineering judgement.\n\n"
            "Three things stood out:\n\n"
            "1. There are no 3D / virtual tours of AG Rosa units. For a luxury developer in 2026, this is "
            "the single biggest gap. Diaspora HNW buyers in London, Houston, Dubai will not fly in to view "
            "— they'll buy from the developer that lets them walk the unit on a phone first.\n\n"
            "2. Instagram and TikTok are active, but neither feed is embedded on the site. Affluent buyers "
            "researching AG Luxury on a laptop can't see the AG Rosa construction progress unless they "
            "leave the site. That's a buying-intent leak.\n\n"
            "3. The premium brand consistency drops on internal pages. The hero is luxury; some interior "
            "pages aren't. For a ₦150M+ unit, every page has to feel like the price tag.\n\n"
            "We build AI systems for exactly this segment: a luxury-toned AI concierge that qualifies HNW "
            "buyers 24/7, diaspora investor flows with USD/GBP toggle and KYC, an AI recommendation engine "
            "that matches buyer profile to AG Rosa / Festus Court inventory, and a premium brand consistency "
            "sweep across every page.\n\n"
            "I put together a 12-slide audit deck specific to AG Luxury — virtual tour cost/ROI, diaspora "
            "buyer journey, Phase 1 roadmap with indicative NGN/USD pricing.\n\n"
            "Happy to send it over or take 15 minutes on a call.\n\n"
            "Worth a conversation?"
            + SIG
        ),
        "fu1": (
            "Hi team,\n\n"
            "Following up on my note from last week.\n\n"
            "One quick note: a single ₦150M unit lost to a competitor with virtual tours is more than the "
            "entire AI concierge + virtual tour build for AG Rosa. The maths is one-sided — and the "
            "diaspora window for AG Rosa won't stay open forever.\n\n"
            "The audit deck is ready if you'd like to see it."
            + SIG_SHORT
        ),
        "fu2": (
            "Hi team,\n\n"
            "Last note from me.\n\n"
            "If the timing isn't right, no problem. If you'd like the deck for a future conversation, "
            "just reply.\n\n"
            "Good luck with AG Rosa. The product looks genuinely special."
            + SIG_SHORT
        ),
    },
    # 6 — SIOR
    {
        "subject": "80,000 SIOR transactions/year — and no AI market report on sior.com",
        "body": (
            "Hi team,\n\n"
            "I spent time on sior.com and the MySIOR portal this week. 4,000+ designees, 630+ cities, 34 "
            "countries, ~80,000 CRE transactions submitted by members every year — the data set is the "
            "single most valuable thing in industrial / office CRE that isn't being fully expressed.\n\n"
            "Three observations:\n\n"
            "1. No competitor association (NAIOP, CCIM, ULI) has anything close to the proprietary deal "
            "data SIOR sits on. But sior.com doesn't surface it. There is no auto-generated weekly market "
            "report, no segment dashboard by metro / property type, no member-facing intelligence layer. "
            "That data could be SIOR's most-cited brand asset in 12 months.\n\n"
            "2. MySIOR (Higher Logic) is showing its age. Member directory search is dated, transaction "
            "submission is form-heavy, the SSO experience feels like 2018. For an association whose value "
            "proposition is exclusivity and professionalism, the portal undersells the brand.\n\n"
            "3. Fall Event NYC (October 2026) and Lisbon International Conference are flagship moments. "
            "An AI event concierge — schedule personalisation, attendee matching, on-demand session "
            "summaries — turns a single conference into a year-long content engine.\n\n"
            "We build AI systems that fit exactly this: an AI market-intelligence engine on top of "
            "SIOR's 80k transactions, a modern Next.js + SSO portal to replace MySIOR's dated UX, an AI "
            "event concierge for Fall Event + Lisbon, and a content engine that turns SIOR Report magazine "
            "and Pulse Blog into segmented member nurture across 34 countries.\n\n"
            "I put together a 13-slide audit deck specific to SIOR — MySIOR gap analysis, market-report "
            "ROI model, event concierge spec, and a phased 12-month roadmap.\n\n"
            "Happy to send it across, or take 30 minutes on a call.\n\n"
            "Worth a conversation?"
            + SIG
        ),
        "fu1": (
            "Hi team,\n\n"
            "Following up on my note from last week.\n\n"
            "One framing that might be useful: NAIOP and CCIM cannot replicate SIOR's transaction "
            "data — it's the moat. An AI-generated quarterly Market Pulse report under the SIOR brand "
            "would be cited in WSJ, Bloomberg and Bisnow within two quarters. That's brand equity "
            "compounding off existing data.\n\n"
            "The audit deck is ready if you'd like to see it."
            + SIG_SHORT
        ),
        "fu2": (
            "Hi team,\n\n"
            "Last note from me on this thread.\n\n"
            "If the timing isn't right, no problem. If you'd like the deck for a future SIOR digital-"
            "strategy conversation, just reply.\n\n"
            "Either way, all the best with Fall Event NYC."
            + SIG_SHORT
        ),
    },
    # 7 — BT Industries
    {
        "subject": "65 years of supply-chain expertise, and btindustries.co.za is still a brochure",
        "body": (
            "Hi team,\n\n"
            "I looked at btindustries.co.za this week. 65+ years, 'Confidence of Supply', ISO 9001, "
            "Level 2 B-BBEE, automotive / mining / aftermarket, plus the new DoGo Power renewables play. "
            "There is real B2B substance here.\n\n"
            "But the website is doing none of the heavy lifting.\n\n"
            "Three things:\n\n"
            "1. There is no customer portal. OEM and mining buyers can't see their pricing tier, stock "
            "availability, open orders, or invoice history without phoning a rep. Competitors with even "
            "basic portals are winning supply contracts on procurement-team UX alone.\n\n"
            "2. There is no catalogue search. A procurement engineer with a fastener spec or part number "
            "has no way to check whether BT carries it without an email. Every one of those is a quote "
            "request lost to a competitor with a working search.\n\n"
            "3. RFQ-to-quote is fully manual. An AI quotation pipeline (RFQ email → AI extracts parts → "
            "priced quote out within minutes) would shrink the sales cycle from days to hours on standard "
            "requests, and free the team for complex engineered work.\n\n"
            "We build AI systems that fix exactly this for industrial B2B: a customer portal with SAP/ERP "
            "integration, AI catalogue search across fasteners and components, AI quotation automation, "
            "and a solar-vertical lead engine to capitalise on the DoGo Power partnership before "
            "competitors notice the renewables window.\n\n"
            "I put together a 12-slide audit deck specific to BT Industries — portal scope, quote-"
            "automation ROI, solar-vertical opportunity, and a phased roadmap with indicative ZAR pricing.\n\n"
            "Happy to send it over or take 15 minutes on a call.\n\n"
            "Worth a conversation?"
            + SIG
        ),
        "fu1": (
            "Hi team,\n\n"
            "Following up on last week's note.\n\n"
            "One number worth flagging: industrial distributors that deploy AI quote automation typically "
            "see RFQ-to-quote drop from 2–3 days to under 30 minutes — and win-rate improvements of 15–25% "
            "on standard requests purely because the customer hears back first. Mining and OEM procurement "
            "teams place the order with whoever responds, not whoever is cheapest.\n\n"
            "The audit deck is ready if you'd like to see it."
            + SIG_SHORT
        ),
        "fu2": (
            "Hi team,\n\n"
            "Last note from me.\n\n"
            "If the timing isn't right, no problem. If you'd like the deck for reference, just reply.\n\n"
            "Either way, all the best with the DoGo Power partnership — the solar entry is well-timed."
            + SIG_SHORT
        ),
    },
    # 8 — Tokuti (Mark)
    {
        "subject": "Webflow can't run a Web3 auction. Here's exactly where it breaks.",
        "body": (
            "Hi Mark,\n\n"
            "I went through tokuti.io and the auction concept this week. 25 years at Credit Suisse and "
            "HSBC, GATENet ecosystem, RWA tokenization — the thesis is clean and you're early enough that "
            "the first-mover window is genuinely open ($4T projected market by 2030, no dominant auction "
            "platform yet).\n\n"
            "Webflow was the right MVP choice. It is not the right Tokuti.\n\n"
            "Five hard ceilings Webflow can't get past:\n\n"
            "1. No wallet authentication. MetaMask, WalletConnect, Coinbase Wallet — none of them work "
            "natively on Webflow. Every wallet interaction needs custom code stitched on the side.\n\n"
            "2. No real-time auction engine. Live bidding, countdown timers synced to on-chain state, "
            "and dynamic price updates need WebSockets and server logic Webflow's static hosting can't run.\n\n"
            "3. No KYC / AML / accredited-investor flows. Sumsub, Onfido and jurisdictional gating need "
            "backend infrastructure that Webflow forms can't deliver to financial-grade standards.\n\n"
            "4. No scalability for launch events. A flagship tokenized property auction could spike "
            "traffic 100× overnight. Webflow's CDN is built for marketing sites, not financial platforms.\n\n"
            "5. No code ownership. You can't fully self-host, can't run custom deploy pipelines, can't "
            "let auditors review a deployment that lives on Webflow's infrastructure.\n\n"
            "We build production Web3 platforms exactly for this: Next.js 15 + wagmi/viem for wallet "
            "auth, Solidity smart contracts for on-chain auction logic, WebSockets for live bidding, "
            "Sumsub for KYC, Stripe + USDC/ETH/BTC for payments, Vercel + Cloudflare for 100× spikes — "
            "and full code ownership from day one.\n\n"
            "I put together a 12-slide proposal specific to Tokuti — Webflow ceiling analysis, target "
            "architecture, phased build plan, and indicative pricing with milestones.\n\n"
            "Happy to send it over or take 30 minutes on a call. Given the GATENet alignment, there's "
            "also an obvious unified-identity play worth discussing.\n\n"
            "Worth a conversation?"
            + SIG
        ),
        "fu1": (
            "Hi Mark,\n\n"
            "Following up on my note from last week.\n\n"
            "One framing worth flagging: every additional month on Webflow is a month where Tokuti "
            "cannot run a real auction, cannot accept accredited-investor KYC, and cannot demo wallet "
            "auth to institutional partners. The platform you'd want for the first major auction has to "
            "be live and battle-tested before that auction — not built during it.\n\n"
            "The proposal deck is ready if you'd like to see it."
            + SIG_SHORT
        ),
        "fu2": (
            "Hi Mark,\n\n"
            "Last note from me.\n\n"
            "If the timing isn't right, no problem. If you'd like the proposal for a future Tokuti "
            "platform conversation, just reply and I'll send it across.\n\n"
            "Either way, all the best with Tokuti and GATENet — the thesis is one of the cleanest "
            "RWA setups I've seen this year."
            + SIG_SHORT
        ),
    },
]

# ── Pipeline (Next Action + Est. Deal Value) ─────────────────────────────────
PIPELINE = [
    {"next_action": "Send audit deck + intro email to info@eris.co.za; LinkedIn outreach to Marketing / Digital lead",
     "deal":  "$20,000–$45,000 (Year-1 web rebuild + AI automation retainer est.)"},
    {"next_action": "Send audit deck + intro email to info@befittingproperties.com; LinkedIn outreach + WhatsApp test",
     "deal":  "$15,000–$32,000 (Year-1 web + WhatsApp AI agent + mobile UX est.)"},
    {"next_action": "Email Sales@agluxurydev.com with virtual-tour proposal; LinkedIn to AG Luxury founder",
     "deal":  "$12,000–$28,000 (AI concierge + virtual tours + diaspora flow est.)"},
    {"next_action": "Send audit deck to info@sior.com; warm intro via LinkedIn to CMO / Membership Director",
     "deal":  "$60,000–$150,000+ (MySIOR portal modernisation + AI market intelligence — flagship engagement)"},
    {"next_action": "Send audit deck + intro email to info@btindustries.co.za; LinkedIn to Sales / Marketing Director",
     "deal":  "$18,000–$40,000 (Customer portal + AI quotation automation + solar lead engine)"},
    {"next_action": "Send proposal deck to mark@tokuti.io; LinkedIn DM Mark Mariampillai; offer 30-min architecture call",
     "deal":  "$35,000–$90,000+ (Web3 auction platform build — Next.js + smart contracts + KYC)"},
]


def main() -> None:
    wb = openpyxl.load_workbook(WB_PATH)
    ws_leads = wb["Organic Leads"]
    ws_mail  = wb["Cold Emails"]
    ws_pipe  = wb["Pipeline"]

    # Find next free row + next index in each sheet (skip blank rows)
    def next_row(ws):
        r = ws.max_row
        while r > 1 and all(ws.cell(row=r, column=c).value in (None, "") for c in range(1, ws.max_column + 1)):
            r -= 1
        return r + 1

    def next_idx(ws):
        last_idx = 0
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, int):
                last_idx = max(last_idx, v)
        return last_idx + 1

    # Inherit style from the last existing data row so new rows blend in
    def style_template(ws, row):
        return [copy(ws.cell(row=row, column=c)._style) for c in range(1, ws.max_column + 1)]

    leads_style = style_template(ws_leads, ws_leads.max_row) if ws_leads.max_row > 1 else None
    mail_style  = style_template(ws_mail,  ws_mail.max_row)  if ws_mail.max_row  > 1 else None
    pipe_style  = style_template(ws_pipe,  ws_pipe.max_row)  if ws_pipe.max_row  > 1 else None

    leads_row_start = next_row(ws_leads)
    mail_row_start  = next_row(ws_mail)
    pipe_row_start  = next_row(ws_pipe)

    leads_idx = next_idx(ws_leads)
    mail_idx  = next_idx(ws_mail)
    pipe_idx  = next_idx(ws_pipe)

    for i, lead in enumerate(LEADS):
        # Organic Leads
        row = leads_row_start + i
        values = [
            leads_idx + i,
            lead["company"],
            lead["contact"],
            lead["title"],
            lead["email"],
            lead["phone"],
            lead["website"],
            lead["linkedin"],
            lead["city"],
            lead["state"],
            lead["industry"],
            lead["source"],
            lead["status"],
            lead["audit_score"],
            lead["deck_path"],
            lead["opps"],
            lead["notes"],
        ]
        for c, v in enumerate(values, 1):
            cell = ws_leads.cell(row=row, column=c, value=v)
            if leads_style:
                cell._style = copy(leads_style[c - 1])
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Cold Emails
        m = EMAILS[i]
        row = mail_row_start + i
        values = [
            mail_idx + i,
            lead["company"],
            lead["contact"],
            lead["email"],
            m["subject"],
            m["body"],
            m["fu1"],
            m["fu2"],
        ]
        for c, v in enumerate(values, 1):
            cell = ws_mail.cell(row=row, column=c, value=v)
            if mail_style:
                cell._style = copy(mail_style[c - 1])
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Pipeline
        p = PIPELINE[i]
        row = pipe_row_start + i
        values = [
            pipe_idx + i,
            lead["company"],
            lead["contact"],
            lead["email"],
            lead["industry"],
            lead["audit_score"],
            lead["status"],
            "—",
            p["next_action"],
            p["deal"],
        ]
        for c, v in enumerate(values, 1):
            cell = ws_pipe.cell(row=row, column=c, value=v)
            if pipe_style:
                cell._style = copy(pipe_style[c - 1])
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(WB_PATH)
    print(f"Saved → {WB_PATH}")
    print(f"  Added {len(LEADS)} rows to 'Organic Leads' (idx {leads_idx}..{leads_idx+len(LEADS)-1})")
    print(f"  Added {len(EMAILS)} rows to 'Cold Emails' (idx {mail_idx}..{mail_idx+len(EMAILS)-1})")
    print(f"  Added {len(PIPELINE)} rows to 'Pipeline' (idx {pipe_idx}..{pipe_idx+len(PIPELINE)-1})")


if __name__ == "__main__":
    main()
