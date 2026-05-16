from pathlib import Path
import openpyxl
from openpyxl import load_workbook
import copy as cp
import os

ROOT = Path(__file__).parent.parent.parent
WB_DIR = ROOT / "output" / "workbooks"
wb = load_workbook(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx")

subject = "Saw your VP Land Acquisition post - free SCI audit attached"

body = """\
Hi Thomas,

Came across your November post on the VP of Land Acquisition search. A Class A multifamily land play backed by strong capital in this environment takes a sharp team to move fast on.

I'm Rohan Malik, CEO of Genos AI. We help real estate executive search firms cut admin time so consultants spend more hours on mandates.

We ran a free audit of specialtyconsultants.com and attached the report. Two things stood out.

The website doesn't match the firm. SCI has Forbes recognition, 55 years, and a client roster (Prologis, Brookfield, Related) most firms spend decades building. But when a hiring committee lands on your site before a mandate call, the design works against you. CBRE, JLL, and Newmark have all gone premium and motion-led. We'd rebuild SCI's site to reflect what the firm actually is, and give SCI.ON its own landing page with a real funnel.

On the ops side, a few things worth a look:
- AI pre-screening cuts time-to-shortlist by 30-50%, so your consultants work the top candidates instead of all of them
- Search announcement automation publishes new mandates in minutes, not a half-day of back and forth
- Executive movement alerts so SCI gets to the relationship before someone else does
- AI chat captures candidate and client inquiries overnight, on weekends, whenever

Audit's attached, no strings. Happy to walk through it on a quick call if any of this lands.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

follow_up_1 = """\
Hi Thomas, just wanted to make sure the audit didn't get buried. The AI screening section is probably the most relevant given the pace of searches you're running. Happy to walk through it in 15 minutes. - Rohan\
"""

follow_up_2 = """\
Hi Thomas, last one from me. Dropping the audit link here in case the timing is better later. Either way, 55 years of real estate executive search is a serious body of work. If it ever makes sense to talk, you know where to find me. - Rohan\
"""

# Update Cold Emails sheet row 91
ws3 = wb["Cold Emails"]
ws3.cell(91, 5).value = subject
ws3.cell(91, 6).value = body
ws3.cell(91, 7).value = follow_up_1
ws3.cell(91, 8).value = follow_up_2

# Update Organic Leads sheet column U (Cold Email Body)
ws5 = wb["Organic Leads"]
# Find the Cold Email Body column
col_u = None
for col in range(1, ws5.max_column + 1):
    if ws5.cell(1, col).value == "Cold Email Body":
        col_u = col
        break

if col_u:
    ws5.cell(2, col_u).value = body
    print(f"Updated Organic Leads col {col_u} (Cold Email Body)")
else:
    print("Warning: Cold Email Body column not found in Organic Leads")

wb.save(WB_DIR / "GenosAI_RealEstate_Leads_UPDATED.xlsx")
print("Done - humanized email saved to both sheets.")
