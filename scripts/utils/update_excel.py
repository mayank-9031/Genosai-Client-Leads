from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy as cp
import os

ROOT = Path(__file__).parent.parent.parent
WB_DIR = ROOT / "output" / "workbooks"
wb = load_workbook(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx")

# ── Lead data ──
lead = {
    "num": 90,
    "first": "Thomas",
    "last": "Williams",
    "full": "Thomas Williams",
    "title": "COO & EVP",
    "company": "Specialty Consultants Inc.",
    "email": "twilliams@specialtyconsultants.com",
    "email_status": "Valid",
    "website": "https://www.specialtyconsultants.com/",
    "phone": "(412) 355-8200",
    "city": "Pittsburgh",
    "state": "PA",
    "country": "US",
    "employees": "11-50",
    "industry": "Real Estate Executive Search",
    "linkedin": "https://www.linkedin.com/in/thomasgwilliams1",
    "ppt": "SCI_Audit_GenosAI.pptx"
}

def copy_style(src, dst):
    if src.has_style:
        dst.font = cp.copy(src.font)
        dst.fill = cp.copy(src.fill)
        dst.border = cp.copy(src.border)
        dst.alignment = cp.copy(src.alignment)
        dst.number_format = src.number_format

# ── 1. Add to Leads sheet ──
ws = wb["Leads"]
row = 91
vals = [lead["num"], lead["first"], lead["last"], lead["full"], lead["title"],
        lead["company"], lead["email"], lead["email_status"], lead["website"],
        lead["phone"], lead["city"], lead["state"], lead["country"],
        lead["employees"], lead["industry"], lead["linkedin"], lead["ppt"]]
for col, v in enumerate(vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# ── 2. Add to Website Audits sheet ──
ws2 = wb["Website Audits"]
audit_vals = [
    lead["num"], lead["company"], lead["website"],
    "WordPress/Custom", "Yes", "Google Analytics (Likely)", "Unknown", "No",
    "None Detected", "Unknown", "No", "Newsletter Only", 5, "High",
    "No live chat; weak CTAs; only 1 blog post; no testimonials; SCI.ON value prop unclear; no structured data; limited content strategy",
    "AI candidate pre-screening; automated outreach sequences; AI chat for candidate/client capture; SCI.ON AI matching engine; automated client reporting; executive movement intelligence"
]
row2 = 91
for col, v in enumerate(audit_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# ── 3. Add to Cold Emails sheet ──
ws3 = wb["Cold Emails"]
subject = "Free Website Audit + AI Growth Opportunities for Specialty Consultants"

body = (
    "Hi Thomas,\n\n"
    "My name is Rohan Malik from Genos AI — we help real estate executive search firms automate "
    "their recruiting workflows and attract more clients through AI-powered tools.\n\n"
    "I came across Specialty Consultants and was genuinely impressed by your Forbes recognition "
    "as one of \"America's Best Executive Search Firms\" — five decades in business and 7,500+ "
    "completed searches is a track record very few firms can match.\n\n"
    "That said, we ran a complimentary audit of your website and identified a few key opportunities "
    "that could help SCI scale mandates and reduce time-to-placement:\n\n"
    "• Automate candidate sourcing and pre-screening with AI — cut time-to-shortlist by 30–50%\n"
    "• Deploy AI chat to capture and qualify candidate and client inquiries 24/7\n"
    "• Enhance SCI.ON with AI matching for faster, more accurate site-level staffing placements\n"
    "• Improve search visibility so real estate executives find SCI when they need search partners\n"
    "• Automate client reporting and search updates — freeing consultants to focus on relationships\n\n"
    "We've attached the free audit report — no strings attached. It outlines exactly what's working, "
    "what isn't, and the steps to improve it.\n\n"
    "If any of this resonates, I'd love a quick 15-minute call to walk you through our findings "
    "and show you what's possible for SCI.\n\n"
    "Best regards,\n"
    "Rohan Malik\n"
    "CEO, Genos AI\n"
    "📞 +91 638-714-2699\n"
    "🌐 www.genosai.tech\n"
    "hello@genosai.tech"
)

fu1 = (
    "Hi Thomas, just wanted to make sure my previous email didn't get buried. "
    "Did you get a chance to review the audit? Even one of the AI automation opportunities — "
    "like candidate pre-screening or 24/7 AI chat — could meaningfully reduce time-to-placement at SCI. "
    "Happy to walk you through it in 15 minutes. — Rohan"
)

fu2 = (
    "Hi Thomas, last follow-up from me. I'll leave this here in case the timing is better later — "
    "happy to send the full audit whenever it's useful. If it's not the right fit right now, no hard feelings. "
    "Either way, great work building SCI into what it is today. — Rohan"
)

email_vals = [lead["num"], lead["first"], lead["company"], lead["email"], subject, body, fu1, fu2]
row3 = 91
for col, v in enumerate(email_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

# ── 4. Add to Audit PPT Content sheet ──
ws4 = wb["Audit PPT Content"]
ppt_vals = [
    lead["num"], lead["company"], lead["full"], lead["email"], lead["website"], 5, "High",
    "Specialty Consultants Inc. — Website & AI Growth Audit | Prepared by Genos AI",
    "Forbes 'America's Best Executive Search Firms' | 5 decades in business | 7,500+ searches completed | Clients: Prologis, Brookfield, Related, Truist | SCI.ON platform in market",
    "Strong brand credibility and Forbes recognition; blue-chip client roster (Prologis, Brookfield, Related); clear service structure; SCI.ON product initiative shows forward-thinking",
    "No live chat or AI bot (losing after-hours inquiries); only 1 blog post limits SEO reach; zero client testimonials or success metrics; weak CTAs throughout; SCI.ON value prop unexplained; no structured data",
    "1. AI candidate pre-screening (30-50% faster shortlists) | 2. AI chat for 24/7 candidate + client capture | 3. SCI.ON AI matching engine | 4. Automated multi-touch candidate outreach sequences | 5. Executive movement intelligence (signal-based sourcing) | 6. Automated client search progress reports",
    "No AI screening = consultants spending 60%+ of time on manual resume review. No chat = losing warm candidate and client leads after hours. 1 blog post = invisible to thousands of monthly Google searches by real estate hiring managers.",
    "AI candidate screening pipeline (ATS-integrated) | AI chat widget trained on SCI services | SCI.ON AI matching layer | Automated outreach sequences for candidate pipelines | Content engine: 4-8 SEO articles/month targeting real estate hiring execs",
    "Month 1: AI chat live, capturing inquiries 24/7 | Month 2: AI screening live, 30% faster shortlists | Month 3: SCI.ON AI matching tested; 2x content velocity; automated client reporting active",
    "Reply to this email or book a 15-min call | Rohan Malik, CEO | hello@genosai.tech | www.genosai.tech"
]
row4 = 91
for col, v in enumerate(ppt_vals, 1):
    cell = ws4.cell(row4, col)
    cell.value = v
    copy_style(ws4.cell(row4 - 1, col), cell)

# ── 5. Create "Organic Leads" sheet ──
if "Organic Leads" not in wb.sheetnames:
    ws5 = wb.create_sheet("Organic Leads")
else:
    ws5 = wb["Organic Leads"]

headers = [
    "#", "First Name", "Last Name", "Full Name", "Title", "Company", "Email",
    "Email Status", "Website", "Phone", "City", "State", "Country",
    "Employees", "Industry", "LinkedIn", "Source", "Date Added", "PPT File", "Status"
]

header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
thin = Side(style="thin", color="444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
col_widths = [5, 12, 12, 20, 22, 28, 34, 14, 36, 16, 14, 8, 10, 10, 28, 38, 18, 14, 24, 22]

for col, h in enumerate(headers, 1):
    cell = ws5.cell(1, col)
    cell.value = h
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    ws5.column_dimensions[get_column_letter(col)].width = col_widths[col - 1]

ws5.row_dimensions[1].height = 30

# First data row
row_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
row_font = Font(name="Calibri", size=10, color="1A1A2E")

organic_vals = [
    1, "Thomas", "Williams", "Thomas Williams", "COO & EVP",
    "Specialty Consultants Inc.", "twilliams@specialtyconsultants.com",
    "Valid", "https://www.specialtyconsultants.com/", "(412) 355-8200",
    "Pittsburgh", "PA", "US", "11-50", "Real Estate Executive Search",
    "https://www.linkedin.com/in/thomasgwilliams1", "Organic/Manual",
    "2026-05-06", "SCI_Audit_GenosAI.pptx", "Email Sent - Pending Reply"
]

for col, v in enumerate(organic_vals, 1):
    cell = ws5.cell(2, col)
    cell.value = v
    cell.font = row_font
    cell.fill = row_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border

ws5.row_dimensions[2].height = 22
ws5.freeze_panes = "A2"

wb.save(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx")
print("Excel updated successfully — all 5 sheets done.")
