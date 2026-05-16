from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy as cp
import os

ROOT = Path(__file__).parent.parent.parent
WB_DIR = ROOT / "output" / "workbooks"
wb = load_workbook(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx")

# ────────────────────────────────────────────────────────────────
# RESEARCHED EMAIL — Thomas G. Williams, COO & EVP, SCI
# Research used:
#   - His Nov 2025 article: "VP of Land Acquisition – Multifamily"
#   - SCI founded 1970 (55 years, not just "5 decades")
#   - Education: Duquesne University + Regis University
#   - SCI is actively publishing compensation reports (data-driven)
#   - Recent placements: affordable housing, healthcare, multifamily
#   - Forbes recognition still current
#   - SCI.ON product under-marketed
# ────────────────────────────────────────────────────────────────

subject = "Saw your VP Land Acquisition post — free SCI audit attached"

body = """\
Hi Thomas,

Came across your November post announcing the VP of Land Acquisition search — a Class A multifamily land play backed by strong capital in this environment takes a sharp team to move fast on. Quality mandate.

I'm Rohan Malik, CEO of Genos AI. We work with real estate executive search firms to cut the time consultants spend on admin — so more hours go to closing mandates and building client relationships, not manual review.

We ran a complimentary audit of specialtyconsultants.com and attached the full report. A few things stood out:

─── WEBSITE ───────────────────────────────────────────────
SCI has Forbes recognition, 55 years of track record, and a client roster (Prologis, Brookfield, Related) that most firms spend decades building. But the website doesn't reflect that. CBRE, JLL, and Newmark have moved to dark, premium, motion-led designs. When a hiring committee lands on your site before a mandate conversation, the design gap quietly works against the credibility SCI has already earned.

We'd rebuild it to match the firm: fast, premium, conversion-focused — with SCI.ON finally getting its own landing page and funnel.

─── AI AUTOMATION ─────────────────────────────────────────
• AI candidate pre-screening — reduce time-to-shortlist 30–50%, so your consultants spend time on the top 10% instead of reviewing everything
• Search announcement automation — publish and syndicate new mandates in minutes, not hours (relevant given your pace of announcements in 2025)
• Executive movement intelligence — alerts when a placed exec changes roles or a target company signals a hire, so SCI gets to the relationship first
• AI chat on-site — capture candidate and client inquiries 24/7, not just business hours
• SCI.ON AI matching layer — score site-level candidates against open roles automatically, which is what SCI.ON needs to scale beyond a tagline

The audit is attached — no pitch, no strings. It maps out exactly what's working and what's costing SCI leads and time.

If any of this is relevant, I'm easy to reach for a quick 15 minutes.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

follow_up_1 = """\
Hi Thomas, wanted to make sure the audit didn't get buried — it's a quick read and the AI screening section in particular is relevant to the pace of searches you're running right now. Happy to walk through it in 15 minutes at your convenience. — Rohan\
"""

follow_up_2 = """\
Hi Thomas, last note from me. I'll leave the audit link here in case the timing is better down the road. Either way, 55 years of executive search in real estate is a serious body of work — I respect what SCI has built. If it ever makes sense to talk, you know where to find me. — Rohan\
"""

# ── Update Cold Emails sheet row 91 ──
ws3 = wb["Cold Emails"]
new_vals = {
    5: subject,
    6: body,
    7: follow_up_1,
    8: follow_up_2,
}
for col, val in new_vals.items():
    ws3.cell(91, col).value = val

# ── Add "Email Body" column to Organic Leads sheet ──
ws5 = wb["Organic Leads"]

# Find max column
max_col = ws5.max_column
new_col = max_col + 1
new_col_letter = get_column_letter(new_col)

# Header style
header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
thin = Side(style="thin", color="444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Write header
hcell = ws5.cell(1, new_col)
hcell.value = "Cold Email Body"
hcell.font = header_font
hcell.fill = header_fill
hcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
hcell.border = border
ws5.column_dimensions[new_col_letter].width = 60

# Write email body into row 2 (Thomas Williams)
dcell = ws5.cell(2, new_col)
dcell.value = body
dcell.font = Font(name="Calibri", size=10, color="1A1A2E")
dcell.fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
dcell.alignment = Alignment(vertical="top", wrap_text=True)
dcell.border = border
ws5.row_dimensions[2].height = 180  # tall enough to show email preview

wb.save(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx")
print("Done — Cold Emails row 91 updated + 'Cold Email Body' column added to Organic Leads.")
print(f"New column: {new_col_letter} (col {new_col})")
