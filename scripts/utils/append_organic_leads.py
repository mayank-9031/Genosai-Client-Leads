"""Append 6 new leads (Kaden, NewExGen, RealtorCoJo, IdiliAfrica, ShalomPark, Fortress)
to output/workbooks/GenosAI_OrganicLeads.xlsx using the existing row format/style."""
import openpyxl
from copy import copy
from openpyxl.utils import get_column_letter

XLSX = r"output/workbooks/GenosAI_OrganicLeads.xlsx"
SHEET = "Organic Leads"

leads = [
    {
        "Company": "Kaden Investment",
        "Contact Person": "General Enquiries",
        "Title": "Marketing / Digital Lead (decision maker TBC)",
        "Email": None,
        "Phone": None,
        "Website": "https://www.kaden-sa.com",
        "LinkedIn": "https://www.linkedin.com/company/kaden-investment/",
        "City": "Riyadh",
        "State/Country": "Saudi Arabia",
        "Industry": "Real Estate Development & Asset Management",
        "Source": "Manual Research — Website + LinkedIn",
        "Status": "To Contact",
        "Audit Score": "4/10",
        "Deck Path": "output/decks/Kaden_WebDev_AI_Automation_Audit_GenosAI.pptx",
        "Automation Opportunities": (
            "1. Bilingual (AR/EN) AI chatbot — 24/7 enquiry handling on kaden-sa.com\n"
            "2. Auto-generate bilingual project descriptions + AI photo tagging & virtual tour support\n"
            "3. Investor & tenant portal (Azure AD B2C) — lease docs, unit reports, payment history, maintenance\n"
            "4. Algolia bilingual property search with faceted filters (location, type, size, status)\n"
            "5. Saudi market intelligence: weekly AI-generated Vision 2030 alignment reports + predictive vacancy/pricing\n"
            "6. Lease renewal auto-drafts + WhatsApp/email follow-up sequences; maintenance ticket AI routing\n"
            "7. PDPL cookie consent + security headers (CSP, HSTS) + Arabic contract review & clause flagging\n"
            "8. LinkedIn + X auto-posting for new project launches; ERP / SAP integration layer"
        ),
        "Notes": (
            "Squarespace site, Arabic-first with incomplete EN parity; Vision 2030 aligned developer "
            "across Mixed-Use, Logistics, Retail, Waterfront and Commercial. Mobile PageSpeed 44/100, "
            "desktop 69, SEO 58 — partial WCAG AA, no AI/chatbot, limited social automation. "
            "Indicative Year-1 spend SAR 289K–501K against SAR 280K–500K+ savings/upside; breakeven 8–14 months. "
            "First-mover window in KSA prop-tech — Arabic LLMs now Gulf-market ready."
        ),
    },
    {
        "Company": "NewExGen (Pty) Ltd",
        "Contact Person": "Sales / Marketing Team",
        "Title": "Sales / Marketing Director (decision maker TBC)",
        "Email": None,
        "Phone": None,
        "Website": "https://www.newexgen.co.za",
        "LinkedIn": "https://www.linkedin.com/company/newexgen-pty-ltd/",
        "City": "Rivonia, Sandton",
        "State/Country": "South Africa",
        "Industry": "ICT / Digital Transformation & Managed Services",
        "Source": "Manual Research — Website + LinkedIn",
        "Status": "To Contact",
        "Audit Score": "5/10",
        "Deck Path": "output/decks/NewExGen_WebDev_AI_Automation_Audit_GenosAI.pptx",
        "Automation Opportunities": (
            "1. Zoho SalesIQ Zia AI chatbot — 24/7 lead qualification + auto demo booking on newexgen.co.za\n"
            "2. AI-scored leads in Zoho CRM by fit & intent + personalised Zoho Campaigns follow-up sequences\n"
            "3. AI ticket triage & predictive SLA-breach alerts in Zoho Desk for managed services clients\n"
            "4. AI field-technician dispatch optimisation + auto-generated client-facing service reports\n"
            "5. Tender monitoring (CIDB, e-Tender) + weekly AI-generated SA tech market digest + competitor gap analysis\n"
            "6. Invoice / AP automation via Zoho Books + contract & SLA clause flagging\n"
            "7. AI CV screening in Zoho Recruit + meeting transcription → action items\n"
            "8. POPIA cookie consent + CSP/HSTS hardening; LinkedIn thought-leadership content engine"
        ),
        "Notes": (
            "15 yrs in operation, HQ in Rivonia (Sandton). Already a Zoho partner — Zoho Bookings live, "
            "Zoho CRM/Desk/SalesIQ adoption is the natural automation path. Six segments: System Integration, "
            "ICT Field, Managed Services, BPO, SHERQ, Zoho. Mobile PageSpeed 51/100, desktop 74, SEO 61; "
            "no CDN, no chat widget, no schema markup, blog/thought-leadership absent. Demo booking flow "
            "breaks on mobile and Zoho subdomain not HSTS-enforced. Strong Zoho-native fast-win opportunity."
        ),
    },
    {
        "Company": "Realtor CoJo (RC Realtor)",
        "Contact Person": "General Enquiries",
        "Title": "Founder / Sales Lead (decision maker TBC)",
        "Email": None,
        "Phone": None,
        "Website": "https://rc-realtor.vercel.app",
        "LinkedIn": "https://www.linkedin.com/company/rc-realtor/",
        "City": "Lagos",
        "State/Country": "Nigeria",
        "Industry": "Real Estate — Luxury & Affordable Properties",
        "Source": "Manual Research — Website + LinkedIn",
        "Status": "To Contact",
        "Audit Score": "2/10",
        "Deck Path": "output/decks/RealtorCoJo_WebDev_AI_Automation_Audit_GenosAI.pptx",
        "Automation Opportunities": (
            "1. AI WhatsApp bot — 24/7 enquiries, auto-send photos/details, qualify leads (budget/timeline/location), book viewings\n"
            "2. AI-generate property descriptions from specs + auto-resize and tag listing photos\n"
            "3. Auto-post new listings to Instagram & LinkedIn; dynamic price-drop alerts to interested buyers\n"
            "4. AI lead scoring in CRM + WhatsApp + email follow-up sequences; cold-lead re-engagement campaigns\n"
            "5. Viewing reminders + confirmation automation across WhatsApp and email\n"
            "6. Lagos property price-trend AI digest + competitor monitoring on NPC & Jiji\n"
            "7. Weekly AI-generated Lagos market report; Instagram Reels auto-captioned from listing data\n"
            "8. SSR/SSG migration (Next.js 15) + Sanity CMS + WhatsApp click-to-chat on every listing; custom domain + branded email"
        ),
        "Notes": (
            "Early-stage Lagos luxury/affordable property brand still on rc-realtor.vercel.app staging URL — "
            "no custom domain, no branded email. Built as Next.js / React CSR-only = site is effectively "
            "invisible to Google (SEO ~0/100, site title is the only indexable content). No property listings "
            "rendering, no advanced search, no WhatsApp CTA, no testimonials, no NDPR consent. Custom domain "
            "+ SSR/SSG migration unblocks SEO entirely — single highest-ROI move on this account."
        ),
    },
    {
        "Company": "Idili Africa",
        "Contact Person": "General Enquiries",
        "Title": "Founder / Talent Director (decision maker TBC)",
        "Email": None,
        "Phone": None,
        "Website": "https://idili.africa",
        "LinkedIn": "https://www.linkedin.com/company/idili-africa/",
        "City": "Cape Town",
        "State/Country": "South Africa",
        "Industry": "International Talent Mobility & Graduate Recruitment",
        "Source": "Manual Research — Website",
        "Status": "To Contact",
        "Audit Score": "3/10",
        "Deck Path": "output/decks/IdiliAfrica_WebDev_AI_Automation_Audit_GenosAI.pptx",
        "Automation Opportunities": (
            "1. AI CV parser + skills extraction on upload; auto-score candidates against role requirements\n"
            "2. Psychometric pre-screening bot (pre-human review); AI-match candidates → best-fit partner companies\n"
            "3. WhatsApp candidate journey: eligibility check → document upload → assessment booking → status updates\n"
            "4. Auto-send assessment links + interview schedules; post-placement check-ins & feedback collection\n"
            "5. AI-generated shortlist summaries for partner hiring managers + weekly pipeline reports\n"
            "6. Auto-notify partners when new candidates qualify; interview debrief templates auto-generated\n"
            "7. SEO-optimised opportunity listings with JobPosting schema (Google-indexable) targeting 'graduate jobs SA → Sweden'\n"
            "8. AI-draft graduate success stories; auto-post placements to LinkedIn + Instagram; SA university grad-calendar digest"
        ),
        "Notes": (
            "Mission: 'Talent is global. Opportunity should be too.' Based at Innovation City Cape Town. "
            "Routes SA graduates (18–25) to Swedish / global companies — partners include Saab, Atlas Copco, "
            "Bitprop, REY17, Oakvale Invest; affiliated with Sweden Consulate Cape Town and Made with Sweden. "
            "Also seed-funds African AI founders via Oakvale Invest. Site is extremely thin (4 pages, fragile "
            "URLs with 404 subpages), no ATS / candidate portal, no partner company dashboard, no privacy / "
            "POPIA notice, no social links. Needs full talent platform build, not just a website refresh."
        ),
    },
    {
        "Company": "Shalom Park Estate (IFT Realty Ltd)",
        "Contact Person": "Oluremi Oshikanlu",
        "Title": "CEO, IFT Realty Ltd",
        "Email": None,
        "Phone": None,
        "Website": "https://www.shalomparknigeria.com",
        "LinkedIn": "https://www.linkedin.com/company/ift-realty/",
        "City": "Abijo, Ibeju Lekki, Lagos",
        "State/Country": "Nigeria",
        "Industry": "Real Estate — Luxury Estate Development",
        "Source": "Manual Research — Website + Social",
        "Status": "To Contact",
        "Audit Score": "4/10",
        "Deck Path": "output/decks/ShalomPark_WebDev_AI_Automation_Audit_GenosAI.pptx",
        "Automation Opportunities": (
            "1. WhatsApp AI bot — 24/7 enquiries, auto-send brochures + pricing, qualify buyer (budget/timeline/unit), book site visits or virtual tours\n"
            "2. Monthly construction progress auto-updates (WhatsApp + email) to subscribers; milestone messages (foundation, roofing)\n"
            "3. Payment-due reminders with Paystack link; post-purchase welcome sequence + handover pack\n"
            "4. Realtor partner portal (replace Google Meet onboarding) + auto-assign web leads to nearest realtor + commission tracker\n"
            "5. Weekly realtor leaderboard + performance nudges; marketing collateral downloads & referral programme\n"
            "6. AI-targeted diaspora email campaigns (US/UK Nigerians); Instagram Reels auto-captioned from site milestones\n"
            "7. MREIF fund: AI-generated quarterly investor reports + dedicated SEO-optimised landing page (NGN + USD)\n"
            "8. NDPR cookie consent + Sanity CMS migration + transparent NGN/USD pricing + Matterport 3D walkthroughs on every unit"
        ),
        "Notes": (
            "IFT Realty Ltd is a subsidiary of IFT Realty Development LLC, New York. 25-acre gated luxury "
            "community on the Lekki-Epe axis. Units: 2-bed condos, 4-bed and 5-bed semi / fully detached "
            "homes, plus serviced plots. Lagos State Governor's Consent + approved layout + Megacity Plan "
            "in place. WordPress site with mobile PageSpeed ~48/100; ZERO pricing displayed on any listing "
            "(direct conversion blocker for diaspora buyers transacting in USD); Google Meet used for realtor "
            "onboarding (poor luxury brand fit); MREIF investment fund has no dedicated landing page. "
            "Pricing + WhatsApp CTAs + buyer portal = three changes that directly unlock stalled sales."
        ),
    },
    {
        "Company": "Fortress Real Estate Investments Limited",
        "Contact Person": "Steven Brown",
        "Title": "Chief Executive Officer",
        "Email": None,
        "Phone": None,
        "Website": "https://fortressfund.co.za",
        "LinkedIn": "https://www.linkedin.com/company/fortress-income-fund/",
        "City": "Johannesburg",
        "State/Country": "South Africa (also CEE)",
        "Industry": "REIT — Logistics & Convenience Retail (JSE: FFA & FFB)",
        "Source": "Manual Research — Website + JSE listing",
        "Status": "To Contact",
        "Audit Score": "5/10",
        "Deck Path": "output/decks/Fortress_WebDev_AI_Automation_Audit_GenosAI.pptx",
        "Automation Opportunities": (
            "1. SSR migration (Next.js 15) — every SENS page, property listing & ESG report becomes Google-indexable for analysts and institutional screeners\n"
            "2. Live IR hub: auto-aggregated SENS feed, live FFA/FFB share price ticker, interactive dividend history, two-click annual / interim reports\n"
            "3. IR chatbot (Claude API) — 24/7 answers on dividends, NAV, SENS; earnings call transcripts auto-extracted to key metrics\n"
            "4. Logistics & retail availability portal — searchable GLA by province, spec, size; AI qualifies inbound enquiries before BD engages\n"
            "5. AI-powered heads-of-lease drafting assistant + vacancy alert emails to registered tenant prospects + auto-match GLA to tenant requirements\n"
            "6. Portfolio AI: monthly performance summaries, predictive lease-expiry & renewal-risk flagging, automated valuation variance reports\n"
            "7. ESG dashboard: TCFD-aligned KPIs auto-compiled from PMS, AI-drafted GRESB / JSE SRI submissions, monthly board ESG pack auto-generated\n"
            "8. CEE portfolio English showcase + EU logistics market intelligence digest; POPIA + GDPR cookie consent and security headers"
        ),
        "Notes": (
            "Internally managed REIT — 3rd largest in South Africa. JSE-listed FFA & FFB, R29.6B market cap, "
            "7.36% yield, 86.29 cps FFB H1 2025. Portfolio: 2.9M sqm GLA across 52 retail centres + logistics "
            "(SA + Central & Eastern Europe) + 23.9% stake in NEPI Rockcastle; 4.0% vacancy. CEO: Steven Brown, "
            "Johannesburg HQ. Site is a JS-rendered SPA on fortressfund.co.za — content effectively invisible "
            "to Google crawlers and to institutional ESG screeners that parse pages programmatically; mobile "
            "PageSpeed ~50/100. IR UX lags global REIT peers (Growthpoint, Redefine) — SENS buried, no live "
            "share-price widget, no logistics leasing portal. Highest-prestige logo in pipeline at R29.6B scale."
        ),
    },
]

wb = openpyxl.load_workbook(XLSX)
ws = wb[SHEET]

# Find last row with data
last_row = ws.max_row
while last_row > 1 and all(ws.cell(row=last_row, column=c).value in (None, "") for c in range(1, ws.max_column + 1)):
    last_row -= 1

# Copy style from a template row (row 2 = first lead) and apply to new rows
template_row = 2
headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

next_num = (ws.cell(row=last_row, column=1).value or 0) + 1
if not isinstance(next_num, int):
    try:
        next_num = int(next_num) + 1
    except Exception:
        next_num = last_row

for i, lead in enumerate(leads):
    target_row = last_row + 1 + i
    lead_full = {"#": next_num + i, **lead}
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=target_row, column=c_idx, value=lead_full.get(h))
        tmpl = ws.cell(row=template_row, column=c_idx)
        if tmpl.has_style:
            cell.font = copy(tmpl.font)
            cell.fill = copy(tmpl.fill)
            cell.border = copy(tmpl.border)
            cell.alignment = copy(tmpl.alignment)
            cell.number_format = tmpl.number_format
            cell.protection = copy(tmpl.protection)
    # row height matches template
    if template_row in ws.row_dimensions and ws.row_dimensions[template_row].height:
        ws.row_dimensions[target_row].height = ws.row_dimensions[template_row].height

wb.save(XLSX)
print(f"Appended {len(leads)} leads. New last row: {last_row + len(leads)}")
