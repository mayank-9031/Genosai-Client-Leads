from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy as cp
import os

ROOT = Path(__file__).parent.parent.parent
WB_DIR = ROOT / "output" / "workbooks"

import glob
files = glob.glob(str(WB_DIR / "GenosAI_RealEstate_Leads*.xlsx"))
src = next((f for f in files if "UPDATED" in f), str(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx"))
print(f"Loading: {src}")
wb = load_workbook(src)

lead = {
    "num": 92,
    "first": "Steve",
    "last": "Bernas",
    "full": "Steve Bernas",
    "title": "President & CEO",
    "company": "BBB of Chicago & Northern Illinois",
    "email": "sbernas@chicago.bbb.org",
    "email_status": "Valid",
    "website": "https://www.bbb.org/chicago",
    "phone": "(312) 832-0500",
    "city": "Chicago",
    "state": "IL",
    "country": "US",
    "employees": "11-50",
    "industry": "Nonprofit / Consumer Protection",
    "linkedin": "https://www.linkedin.com/in/stevebernas/",
    "ppt": "BBBChicago_Audit_GenosAI.pptx"
}

subject = "Your April column on AI scams - and a free BBB audit"

body = """\
Hi Steve,

Your April Daily Herald piece on moving scams - specifically the line "With AI, the fraudsters are very convincing" - that cuts both ways.

I'm Rohan Malik, CEO of Genos AI. We help organizations use the same tools the bad actors are using, but to move faster, serve more people, and stop things from falling through the cracks.

Ran a free audit of bbb.org/chicago and attached it. A few things stood out.

BBB Chicago serves 9,000 accredited businesses and fields thousands of consumer complaints a year. But most of the workflows still run manually - complaints triaged by hand, accreditation applications sitting in a queue, renewal reminders going out on a fixed schedule regardless of engagement. After-hours inquiries from businesses trying to get accredited go unanswered until Monday morning.

A few things worth a look:
- Complaint triage AI cuts resolution time and spots emerging fraud patterns before they become headlines
- Accreditation processing automation so your team reviews decisions, not paperwork
- Scam alert syndication publishes and distributes Steve's Daily Herald content across all channels in minutes
- Member renewal sequences keep 9,000 accredited businesses engaged before they quietly lapse
- AI chat captures business and consumer inquiries at 2am on a Saturday

On the website side - the centennial, Mayor Brandon Johnson's remarks, Lester Holt at Navy Pier - none of that shows up when a business owner lands on bbb.org/chicago at 9pm deciding whether to get accredited. 100 years of credibility that the local pages don't yet reflect.

Audit's attached, no strings. Happy to walk through it in 15 minutes.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi Steve, just checking the audit didn't get buried. The complaint triage and member renewal sections are probably most relevant for an org handling the volume you are. Happy to walk through it in 15 minutes. - Rohan\
"""

fu2 = """\
Hi Steve, last one from me. 100 years of marketplace trust in Chicago is a serious legacy - the digital side should match. Leaving the audit here in case timing is better later. If it ever makes sense to talk, you know where to find me. - Rohan\
"""

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = cp.copy(src_cell.font)
        dst_cell.fill = cp.copy(src_cell.fill)
        dst_cell.border = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

# ── 1. Leads ──
ws = wb["Leads"]
row = 93
vals = [lead["num"], lead["first"], lead["last"], lead["full"], lead["title"],
        lead["company"], lead["email"], lead["email_status"], lead["website"],
        lead["phone"], lead["city"], lead["state"], lead["country"],
        lead["employees"], lead["industry"], lead["linkedin"], lead["ppt"]]
for col, v in enumerate(vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# ── 2. Website Audits ──
ws2 = wb["Website Audits"]
audit_vals = [
    lead["num"], lead["company"], lead["website"],
    "Custom / National Platform (bbb.org)", "Yes", "Google Analytics", "Unknown",
    "No", "None Detected", "Unknown", "Unknown", "Unknown", 6, "Medium",
    "No live chat or AI chatbot for after-hours business/consumer inquiries; centennial milestone and media presence (Mayor Johnson, Lester Holt) absent from local Chicago pages; no personalization between accredited business vs consumer visitor; scam alert content not auto-syndicated; no video/YouTube content from Chicago team on local pages; limited local CTAs beyond accreditation button",
    "Complaint triage AI for volume routing and fraud pattern detection; accreditation processing automation (reduce manual queue); scam alert auto-syndication across Daily Herald, email, social; member renewal sequences for 9,000 accredited businesses; AI chat 24/7 for business and consumer inquiries; post-event content automation; donor/sponsor engagement sequences"
]
row2 = 93
for col, v in enumerate(audit_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# ── 3. Cold Emails ──
ws3 = wb["Cold Emails"]
email_vals = [lead["num"], lead["first"], lead["company"], lead["email"], subject, body, fu1, fu2]
row3 = 93
for col, v in enumerate(email_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

# ── 4. Audit PPT Content ──
ws4 = wb["Audit PPT Content"]
ppt_vals = [
    lead["num"], lead["company"], lead["full"], lead["email"], lead["website"], 6, "Medium",
    "BBB of Chicago & Northern Illinois - Website & AI Growth Audit | Prepared by Genos AI",
    "Founded 1926 | 100-year centennial celebrated April 22 2026 at Navy Pier | 9,000 accredited businesses | Steve Bernas: 35+ years tenure | President & CEO | Regular Daily Herald columnist | BA Psychology Loyola/Niles 1987 | Active media presence (NBC, ABC, CBS) | Nonprofit mission: advance marketplace trust",
    "Strong brand authority and 100-year credibility; active media presence (TV, radio, Daily Herald); Steve Bernas respected consumer protection voice; good national SEO structure; schema markup implemented; 9,000 member businesses = strong renewal revenue base; Mayor and media relationships; Chicago community events",
    "No live chat or AI chatbot for after-hours inquiries; centennial content and media achievements not reflected on local bbb.org/chicago pages; complaint triage is largely manual; accreditation applications sit in manual queue; scam alerts published one channel at a time; member renewal is non-personalized; no video content from Chicago team on local pages",
    "1. Complaint triage AI (route by type, flag patterns, speed resolution) | 2. Accreditation processing automation (reduce manual review queue) | 3. Scam alert auto-syndication (Steve's Daily Herald articles → email + social + press) | 4. Member renewal sequences (personalized by industry, engagement history) | 5. AI chat 24/7 for business accreditation and consumer complaint intake | 6. Post-event content automation for centennial legacy storytelling",
    "Scammers are using AI to create fake reviews and professional-looking phony sites (Steve said so himself in April 2026). BBB fighting them with manual workflows is an asymmetric battle. No chat = losing after-hours accreditation inquiries to competitors. Non-personalized renewals = lapsed accreditations that quietly don't renew. 100 years of credibility invisible on the local site.",
    "Local bbb.org/chicago rebuild: centennial credibility front and center, Mayor Johnson and Lester Holt moment highlighted, video content from Chicago team embedded, conversion-focused design for accreditation CTA | AI complaint triage system | Scam alert auto-syndication pipeline | Personalized member renewal sequences by industry segment | AI chat trained on BBB Chicago services and accreditation process | Accreditation application workflow automation",
    "Month 1: AI chat live for 24/7 inquiry capture, complaint triage AI deployed, scam alert syndication automated | Month 2: Accreditation processing automation live, member renewal sequences active, local Chicago pages rebuilt | Month 3: Renewal velocity up, complaint resolution time down, centennial content driving local SEO and media pickups",
    "Reply to this email or book a 15-min call | Rohan Malik, CEO | hello@genosai.tech | www.genosai.tech"
]
row4 = 93
for col, v in enumerate(ppt_vals, 1):
    cell = ws4.cell(row4, col)
    cell.value = v
    copy_style(ws4.cell(row4 - 1, col), cell)

# ── 5. Organic Leads ──
ws5 = wb["Organic Leads"]
thin = Side(style="thin", color="444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
row_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
row_font = Font(name="Calibri", size=10, color="1A1A2E")

organic_vals = [
    3, "Steve", "Bernas", "Steve Bernas", "President & CEO",
    "BBB of Chicago & Northern Illinois", "sbernas@chicago.bbb.org",
    "Valid", "https://www.bbb.org/chicago", "(312) 832-0500",
    "Chicago", "IL", "US", "11-50", "Nonprofit / Consumer Protection",
    "https://www.linkedin.com/in/stevebernas/", "Organic/Manual",
    "2026-05-06", "BBBChicago_Audit_GenosAI.pptx", "Email Sent - Pending Reply"
]
for col, v in enumerate(organic_vals, 1):
    cell = ws5.cell(4, col)
    cell.value = v
    cell.font = row_font
    cell.fill = row_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border

# Email body column
for col in range(1, ws5.max_column + 1):
    if ws5.cell(1, col).value == "Cold Email Body":
        cell = ws5.cell(4, col)
        cell.value = body
        cell.font = row_font
        cell.fill = row_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        break

ws5.row_dimensions[4].height = 180

wb.save(WB_DIR / "GenosAI_RealEstate_Leads_UPDATED.xlsx")
print("Excel updated - Steve Bernas added to all 5 sheets.")
