from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import copy as cp
import os

ROOT = Path(__file__).parent.parent.parent
WB_DIR = ROOT / "output" / "workbooks"

import glob
files = glob.glob(str(WB_DIR / "GenosAI_RealEstate_Leads*.xlsx"))
src = next((f for f in files if "UPDATED" in f), str(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx"))
print(f"Loading: {src}")
wb = load_workbook(src)

lead = {
    "num": 104,
    "first": "June",
    "last": "James",
    "full": "June James",
    "title": "Founder",
    "company": "Notary Workforce",
    "email": "june@notaryworkforce.com",
    "email_status": "Verified (RocketReach + ZoomInfo)",
    "website": "https://www.notaryworkforce.com/",
    "phone": "443-207-1278",
    "city": "Manassas",
    "state": "VA",
    "country": "US",
    "employees": "2-10",
    "industry": "Estate Planning / Notary Services / Legal",
    "linkedin": "https://www.linkedin.com/in/junejames/",
    "ppt": "NotaryWorkforce_Audit_GenosAI.pptx"
}

subject = "845 notarized trusts and a webinar that loses most of its leads at the end"

body = """\
Hi June,

Saw your LinkedIn post about attorneys being buried in admin while they should be practicing law. That framing is exactly right - and it also describes the gap I noticed on notaryworkforce.com from the client side.

You're running a free 25-minute webinar as your top-of-funnel. That's a smart move - it educates before asking for anything. But the moment someone finishes that webinar and doesn't immediately book a 15-minute call, there's nothing holding them. No automated follow-up. No email sequence. No chatbot to answer the questions they probably have at 11pm that made them register in the first place.

845 notarized trusts across all 50 states means you've already built the proof. The gap is the infrastructure between "watched the webinar" and "became a client" - especially for the 80% of registrants who don't convert immediately but aren't saying no.

Same thing on the attorney side - your LinkedIn post pitches mobile trust delivery and client onboarding as an outsourcing solution for estate planning attorneys. That's a strong B2B angle. But the website doesn't have a dedicated attorney landing page, a qualifying intake form, or any automation that moves an interested attorney from "saw the post" to "scheduled a demo."

I'm Rohan Malik, CEO of Genos AI. We build AI agents and website automation for service businesses - specifically the post-webinar nurture sequences, chatbots that pre-qualify at any hour, and intake flows that replace manual back-and-forth. Ran a quick audit of notaryworkforce.com and found a few specific things worth 15 minutes.

Happy to walk you through it whenever works.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi June, following up on my last note. The specific gap - webinar registrants who don't book immediately but aren't saying no - is where the most straightforward win is. Automated sequence after webinar, AI chatbot for late-night questions, attorney intake form. Happy to walk through the audit in 15 minutes. - Rohan | Genos AI\
"""

fu2 = """\
June, last one from me. The Notary Workforce audit is ready whenever the timing is right. The attorney B2B pipeline you're building on LinkedIn deserves website infrastructure to match it. Leaving it here in case it's useful. - Rohan | Genos AI\
"""

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = cp.copy(src_cell.font)
        dst_cell.fill = cp.copy(src_cell.fill)
        dst_cell.border = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

# Leads
ws = wb["Leads"]
row = 105
vals = [lead["num"], lead["first"], lead["last"], lead["full"], lead["title"],
        lead["company"], lead["email"], lead["email_status"], lead["website"],
        lead["phone"], lead["city"], lead["state"], lead["country"],
        lead["employees"], lead["industry"], lead["linkedin"], lead["ppt"]]
for col, v in enumerate(vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# Website Audits
ws2 = wb["Website Audits"]
audit_vals = [
    lead["num"], lead["company"], lead["website"],
    "WordPress", "Yes", "Partial (basic contact form)", "No",
    "No", "No", "No (845 trusts not visible on homepage)", "No", "No", 5, "High",
    "Site is functional but leaks the majority of its leads at every stage: free 25-min webinar is top-of-funnel but no automated post-webinar email sequence visible - 80%+ of registrants who don't immediately book a call are never re-engaged; no AI chatbot or live chat - prospects with questions at 10pm get nothing; no dedicated attorney landing page despite active B2B pivot on LinkedIn pitching mobile trust delivery outsourcing to estate planning firms; no qualifying intake for attorney prospects (no project size, state coverage, volume estimate captured before discovery call); 845 notarized trusts across 50 states = strong social proof completely buried - not visible on homepage; no scheduling automation - mobile trust delivery is logistics-heavy but site has no AI scheduling; no SEO content targeting estate planning by state despite 50-state coverage; no pricing page, no packages visible",
    "Deploy AI chatbot trained on estate planning FAQ, webinar content, and White-Glove Trust Delivery services - captures late-night questions and pre-qualifies without June's time; build automated post-webinar nurture sequence (Day 1, Day 3, Day 7, Day 14) converting registrants who watched but did not book; build dedicated attorney landing page with qualifying intake form (state coverage needed, trust volume/month, current signing process, pain point) routing to June with full context; surface 845 trusts stat above fold on homepage with testimonials and state coverage map; deploy AI scheduling automation for mobile trust delivery appointments; SEO content engine: estate planning guides by state (Virginia, Maryland, DC first then all 50); add probate cost calculator as lead magnet with email capture before results shown"
]
row2 = 105
for col, v in enumerate(audit_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# Cold Emails
ws3 = wb["Cold Emails"]
email_vals = [lead["num"], lead["first"], lead["company"], lead["email"], subject, body, fu1, fu2]
row3 = 105
for col, v in enumerate(email_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

# Audit PPT Content
ws4 = wb["Audit PPT Content"]
ppt_vals = [
    lead["num"], lead["company"], lead["full"], lead["email"], lead["website"], 5, "High",
    "Notary Workforce - Website & AI Growth Audit | Prepared by Genos AI",
    "June James, CNTDA, CNSA: Founder of Notary Workforce | Title: Legacy Curator-in-Chief | Location: Manassas, Virginia | Credentials: CNTDA (Certified Notary Trust Delivery Agent), CNSA (Certified Notary Signing Agent) | Also Managing Director at The Signing Workforce LLC | Notary since 2002 - originally commissioned at Wachovia/Wells Fargo as Financial Specialist (trust documents, refinance, HELOC) | 20+ years in notary, real estate, mortgage consulting | Email: june@notaryworkforce.com | Phone: 443-207-1278 | Notary Workforce: WordPress site, Manassas VA, all 50 states, 31 years combined experience, 845+ notarized trusts | Services: free 25-min estate planning webinar (ClickMeeting), free 15-min consultation, document planning with attorney advocates, notarization, estate plan maintenance, White-Glove Trust Delivery | B2B pivot: LinkedIn posts pitch mobile trust delivery outsourcing to estate planning attorneys - handles client onboarding, scheduling, signing, notarization so attorneys focus on billable work | Probate cost calculator tool on site | No AI features, no chatbot, no client portal, no scheduling automation",
    "June actively posts on LinkedIn pitching attorneys as B2B clients - the concept and angle are sharp; 845 notarized trusts across 50 states = strong social proof and operational track record; free webinar top-of-funnel is the right lead magnet for estate planning where education precedes purchase; CNTDA and CNSA credentials differentiate from generic notary services; White-Glove Trust Delivery is a premium, differentiated offering not common in the market; 20+ years starting at Wachovia/Wells Fargo gives credibility in financial/legal services; probate cost calculator exists as a lead tool - just needs email gate before results to build the list",
    "No post-webinar email sequence - 80% of registrants who don't immediately book are permanently lost; no AI chatbot - late-night estate planning questions get no response; no dedicated attorney landing page despite active B2B pivot; no attorney intake form - qualifying detail not captured before discovery call; 845 trusts buried - social proof not visible on homepage; no AI scheduling automation for mobile trust delivery logistics; no SEO content by state despite 50-state coverage; probate calculator shows results without capturing email first - missed list-building opportunity; no pricing page or package transparency",
    "1. AI chatbot trained on webinar content + estate planning FAQ + service details - captures and pre-qualifies 24/7 | 2. Post-webinar nurture sequence (Day 1/3/7/14) - re-engages the 80% who watched but did not immediately book | 3. Attorney B2B landing page with qualifying intake (state coverage, trust volume, current process, pain point) | 4. Surface 845 trusts + testimonials above fold on homepage - convert social proof into visible trust signal | 5. Gate probate calculator results behind email capture - turns calculator traffic into list | 6. AI scheduling automation for mobile trust delivery appointments",
    "June wrote on LinkedIn: attorneys became lawyers to practice law but estate planning admin is burying them. The same is true for Notary Workforce on the consumer side - June built a 20-year track record and 845 notarized trusts, but the website sends 80% of webinar registrants into silence. The gap between 'watched the webinar' and 'became a client' is exactly where Genos AI operates.",
    "Website: surface 845 trusts above fold, gate probate calculator, dedicated attorney page | AI automation: post-webinar nurture sequence, AI chatbot for 24/7 pre-qualification, attorney intake form with routing, AI scheduling for mobile trust delivery appointments | SEO: estate planning guides by state starting VA/MD/DC then all 50 | Implementation: chatbot and post-webinar sequence first (fastest revenue impact), attorney page and intake second, SEO content third",
    "Week 1: AI chatbot deployed trained on webinar content, service details, and estate planning FAQ; post-webinar email nurture sequence (Day 1/3/7/14) live and connected to ClickMeeting registration | Week 2: Attorney B2B landing page live with qualifying intake form routing to June; probate calculator gated behind email capture; 845 trusts surfaced on homepage | Month 2: AI scheduling automation for mobile trust delivery appointments; first 10 state-specific estate planning SEO articles published (VA, MD, DC, FL, TX) | Month 3: SEO articles indexing; attorney intake data informing B2B outreach targeting; full 50-state content calendar planned",
    "Reply to this email or book a 15-min call | Rohan Malik, CEO | hello@genosai.tech | www.genosai.tech"
]
row4 = 105
for col, v in enumerate(ppt_vals, 1):
    cell = ws4.cell(row4, col)
    cell.value = v
    copy_style(ws4.cell(row4 - 1, col), cell)

# Organic Leads
ws5 = wb["Organic Leads"]
thin = Side(style="thin", color="444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
row_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
row_font = Font(name="Calibri", size=10, color="1A1A2E")

organic_vals = [
    15, "June", "James", "June James", "Founder",
    "Notary Workforce", "june@notaryworkforce.com",
    "Verified (RocketReach + ZoomInfo)", "https://www.notaryworkforce.com/",
    "443-207-1278",
    "Manassas", "VA", "US", "2-10", "Estate Planning / Notary Services / Legal",
    "https://www.linkedin.com/in/junejames/", "Organic/Manual",
    "2026-05-11", "NotaryWorkforce_Audit_GenosAI.pptx", "Email Sent - Pending Reply"
]
for col, v in enumerate(organic_vals, 1):
    cell = ws5.cell(16, col)
    cell.value = v
    cell.font = row_font
    cell.fill = row_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border

for col in range(1, ws5.max_column + 1):
    if ws5.cell(1, col).value == "Cold Email Body":
        cell = ws5.cell(16, col)
        cell.value = body
        cell.font = row_font
        cell.fill = row_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        break

ws5.row_dimensions[16].height = 180

wb.save(WB_DIR / "GenosAI_RealEstate_Leads_UPDATED.xlsx")
print("Excel updated - June James / Notary Workforce added to all 5 sheets.")
