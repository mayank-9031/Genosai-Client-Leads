from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import copy as cp
import os

ROOT = Path(__file__).parent.parent.parent
WB_DIR = ROOT / "output" / "workbooks"

import glob
files = glob.glob(str(WB_DIR / "GenosAI_RealEstate_Leads*.xlsx"))
src = next((f for f in files if "UPDATED" in f), str(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx"))
print(f"Loading: {src}")
wb = load_workbook(src)

lead = {
    "num": 103,
    "first": "Andrea",
    "last": "Frymire",
    "full": "Andrea Frymire",
    "title": "President & CEO",
    "company": "Housing for Communities, Inc.",
    "email": "andrea@housingforcommunities.org",
    "email_status": "Verified (listed on board page)",
    "website": "https://www.housingforcommunities.org/",
    "phone": "(405) 417-1183",
    "city": "Oklahoma City",
    "state": "OK",
    "country": "US",
    "employees": "2-10",
    "industry": "Affordable Housing / Non-profit",
    "linkedin": "https://www.linkedin.com/company/housing-for-communities-inc/",
    "ppt": "HousingForCommunities_Audit_GenosAI.pptx"
}

subject = "Senior Housing conference in 8 days - one chatbot could free Madeline's inbox"

body = """\
Hi Andrea,

Saw the Senior Housing conference is May 19th - eight days out as I write this. Your team at Housing for Communities is four people running training, consulting, development work, AND a major event in a week. That's a lot of weight on Madeline and Bailey.

I looked at housingforcommunities.org. The contact form is four fields and a send button, no FAQ, no chatbot. Anyone landing on the site this week with a "is there still space," "do you have CEU paperwork," "what's the parking situation" question is generating an email that needs a human to answer it. Multiply that across the next eight days.

A simple AI agent on the site that knows the conference agenda, registration link, the AGING20 promo code, parking, CEU details, and your training and consulting service descriptions would absorb 70-80% of those questions in real time. Your team gets the inbox back. The 20% that need a human get routed with full context already captured.

Beyond the conference: same agent answers questions about your three program areas (training, consulting, development), captures qualified leads for consulting work, and surfaces them to Bailey instead of letting them sit in info@ until someone has bandwidth.

I'm Rohan Malik, CEO of Genos AI. We work with mission-driven non-profits and real estate organizations on website automation and AI agents. Cost is far lower than people assume because the work scales - one agent, every visitor.

I built a short audit deck for Housing for Communities specifically. Happy to send it across or walk you through it after the 19th when life calms down.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi Andrea, hope the conference went well on the 19th. Following up on the AI agent idea - now that the event noise has cleared, the same automation pays for itself across consulting intake and training enrollment for the rest of the year. 15 minutes whenever works. - Rohan | Genos AI\
"""

fu2 = """\
Andrea, last note from me. The audit for Housing for Communities is sitting ready - mostly focused on freeing your team from inbox triage so they can focus on the housing work. If the timing isn't right now, no worries. Wishing the Coalition's 10-year anniversary year goes well. - Rohan | Genos AI\
"""

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = cp.copy(src_cell.font)
        dst_cell.fill = cp.copy(src_cell.fill)
        dst_cell.border = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

# Leads
ws = wb["Leads"]
row = 104
vals = [lead["num"], lead["first"], lead["last"], lead["full"], lead["title"],
        lead["company"], lead["email"], lead["email_status"], lead["website"],
        lead["phone"], lead["city"], lead["state"], lead["country"],
        lead["employees"], lead["industry"], lead["linkedin"], lead["ppt"]]
for col, v in enumerate(vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# Website Audits
ws2 = wb["Website Audits"]
audit_vals = [
    lead["num"], lead["company"], lead["website"],
    "Wix (basic CMS)", "Yes", "Partial (4-field form only)", "No",
    "No", "No", "Partial (event mentions only)", "No", "No", 5, "High",
    "Site is functional but minimal for a 4-person non-profit running training, consulting, AND a 200+ attendee conference in 8 days: contact form is 4 fields with no qualification or routing - every inquiry lands in info@ inbox; no AI chatbot or live chat - all conference questions (registration, CEU paperwork, parking, agenda, AGING20 promo) require human email responses; no FAQ or self-service knowledge base; no donor portal or recurring giving CTA visible despite Walmart/Sam's Club Spark Good grant recognition; no separate intake forms for the three service lines (training enrollment, consulting inquiry, development partnership) - all funnel to one generic form; no email signup for newsletter or housing policy alerts; no automated training course enrollment or CEU certificate delivery; no Spanish-language version despite Oklahoma demographic need",
    "Deploy AI agent on homepage trained on conference details (May 19 Senior Housing, agenda, AGING20 promo, CEU, parking, registration link) and the three service lines - absorbs 70-80% of routine inquiries; build separate qualified intake forms for Training (course interest, CEU need), Consulting (project type, timeline, budget range), and Development (partnership type, geography) routing to Bailey/Madeline with full context; add donor-facing CTA with recurring giving option leveraging Spark Good grant credibility; deploy newsletter signup with automated welcome sequence covering Oklahoma housing policy updates; add Spanish-language site version using AI translation for Oklahoma's Spanish-speaking housing population; build automated post-event follow-up - certificate delivery, session recording access, next-event registration; add Andrea/team thought leadership content engine - she's CCIM, founder of OK Coalition, 2021 CRE Advocate of the Year, and her LinkedIn voice is currently underutilized as web content"
]
row2 = 104
for col, v in enumerate(audit_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# Cold Emails
ws3 = wb["Cold Emails"]
email_vals = [lead["num"], lead["first"], lead["company"], lead["email"], subject, body, fu1, fu2]
row3 = 104
for col, v in enumerate(email_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

# Audit PPT Content
ws4 = wb["Audit PPT Content"]
ppt_vals = [
    lead["num"], lead["company"], lead["full"], lead["email"], lead["website"], 5, "High",
    "Housing for Communities - Website & AI Audit | Prepared by Genos AI",
    "Andrea N. Frymire, CCIM: President & CEO of Housing for Communities, Inc. | 25+ years in commercial real estate, property management, and affordable housing | Education: BS Finance Radford University, Master's Real Estate (Commercial Real Estate Investment & Analysis), CCIM designation | Career: Partner at Welcome Home Management Services LLC; Dobson Mortgage Corp; Broker/Owner Oklahoma Investment Realty Inc; licensed real estate broker and instructor | Founder & Inaugural Board President, Oklahoma Coalition for Affordable Housing (Board President again 2025 for 10-year anniversary) | Assisted forming 8+ real estate non-profits and foundations | 2021 Commercial Real Estate Advocate of the Year | Direct: andrea@housingforcommunities.org / (405) 417-1183 | Housing for Communities, Inc.: founded 2023, OKC-based 501(c)(3); 4-person team (Andrea CEO, Madeline King Marketing Coord, Bailey Koewing Development Director, Kaley Wilson Accountant); 3 service pillars (Training, Consulting, Development); Senior Housing Conference May 19, 2026 at Metro Tech OKC; Walmart Neighborhood Market and Sam's Club Spark Good Local Grants recipient; partnerships with Earthly Dwelling, Housing Forward, National Council on Aging | Address: 500 N Broadway Ave Suite 175, OKC OK 73102",
    "Andrea is a recognized housing policy advocate with 25+ years of credibility, CCIM designation, and double-board-presidency standing - she opens doors that small non-profits cannot; the Senior Housing conference May 19 is 8 days out from outreach date - immediate AI agent value is undeniable for absorbing inquiry volume; 4-person team running 3 service lines plus a major event proves operational leverage need; existing Spark Good grant relationships with Walmart/Sam's Club show fundraising is working - donor-facing automation extends that; mission-driven org with measurable outputs (training delivered, consulting projects, units developed) means automation ROI is straightforward to articulate; OK Coalition for Affordable Housing co-presidency provides cross-organization credibility once one site is built well",
    "4-field contact form with no qualification or routing - all 3 service lines funnel to one info@ inbox; no AI chatbot to absorb conference inquiries (May 19 event = 8 days out); no FAQ or self-service knowledge base; no separate intake for Training enrollment, Consulting inquiry, Development partnership; no donor CTA or recurring giving despite Spark Good grant credibility; no newsletter or housing policy alert signup; no automated CEU certificate delivery for training attendees; no Spanish-language site for Oklahoma's Spanish-speaking population; thought leadership content underleveraged - Andrea's CCIM credibility and 25 years of expertise barely surfaced on the site",
    "1. Conference AI agent (week 1, immediate ROI for May 19 event) - trained on agenda, AGING20 promo, CEU paperwork, parking, registration; routes the 20% needing humans with full context | 2. Three-track intake forms - Training, Consulting, Development - each capturing qualifying detail and routing to Bailey/Madeline | 3. Donor and recurring-giving CTA leveraging Spark Good grant credibility | 4. Newsletter automation with Oklahoma housing policy welcome sequence | 5. Andrea thought leadership content engine - CCIM, founder, advocate-of-year credibility translated to indexed web content | 6. Spanish-language site version using AI translation for OK's Spanish-speaking housing population",
    "Andrea founded a non-profit with 25 years of commercial real estate credibility behind her, but the day-to-day reality is a 4-person team running training, consulting, development, AND a 200+ attendee conference in 8 days. The site they built is fine. It's not the bottleneck - their inbox is. Genos AI builds the AI agent that handles the inbox so Andrea can focus on the housing policy work that only she can do.",
    "Website: deploy conference AI agent immediately (8-day timeline), build 3-track intake forms, add donor CTA, add Spanish version | AI automation: chatbot trained on services and conference, automated CEU certificate delivery, post-event follow-up sequence, newsletter welcome sequence, automated consulting lead qualification routed to Bailey | Implementation: AI agent live before May 19 conference (rapid deployment), full website upgrade post-conference once event noise clears",
    "Week 1 (pre-conference May 11-19): AI agent deployed on homepage trained on Senior Housing conference details, AGING20 promo code, CEU info, parking, registration; agent absorbs majority of conference inquiries before May 19; team monitors and refines | Week 2-3 (post-conference): expand agent training to Training/Consulting/Development service lines; deploy 3-track intake forms with routing logic; automated CEU certificate delivery and post-event follow-up sequence go live | Month 2: donor CTA and recurring-giving flow live; newsletter automation with Oklahoma housing policy welcome sequence; thought leadership content engine producing first Andrea-voice articles indexed across affordable-housing keywords | Month 3: Spanish-language site version live; Coalition cross-promotion configured; conference 2027 early-interest pipeline already capturing leads",
    "Reply to this email or book a 15-min call | Rohan Malik, CEO | hello@genosai.tech | www.genosai.tech"
]
row4 = 104
for col, v in enumerate(ppt_vals, 1):
    cell = ws4.cell(row4, col)
    cell.value = v
    copy_style(ws4.cell(row4 - 1, col), cell)

# Organic Leads
ws5 = wb["Organic Leads"]
thin = Side(style="thin", color="444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
row_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
row_font = Font(name="Calibri", size=10, color="1A1A2E")

organic_vals = [
    14, "Andrea", "Frymire", "Andrea Frymire", "President & CEO",
    "Housing for Communities, Inc.", "andrea@housingforcommunities.org",
    "Verified (listed on board page)", "https://www.housingforcommunities.org/",
    "(405) 417-1183",
    "Oklahoma City", "OK", "US", "2-10", "Affordable Housing / Non-profit",
    "https://www.linkedin.com/company/housing-for-communities-inc/", "Organic/Manual",
    "2026-05-11", "HousingForCommunities_Audit_GenosAI.pptx", "Email Sent - Pending Reply"
]
for col, v in enumerate(organic_vals, 1):
    cell = ws5.cell(15, col)
    cell.value = v
    cell.font = row_font
    cell.fill = row_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border

for col in range(1, ws5.max_column + 1):
    if ws5.cell(1, col).value == "Cold Email Body":
        cell = ws5.cell(15, col)
        cell.value = body
        cell.font = row_font
        cell.fill = row_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        break

ws5.row_dimensions[15].height = 180

wb.save(WB_DIR / "GenosAI_RealEstate_Leads_UPDATED.xlsx")
print("Excel updated - Andrea Frymire / Housing for Communities added to all 5 sheets.")
