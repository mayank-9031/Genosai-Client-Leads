"""
Fill missing Cold Emails and Pipeline rows for leads 9-14.
These leads exist in Organic Leads (rows 10-15) but were never written to the other two sheets.
We insert 6 rows after row 9 in both sheets, then populate them.
"""
from pathlib import Path
import openpyxl
import copy as cp

ROOT = Path(__file__).parent.parent.parent
WB_PATH = ROOT / "output" / "workbooks" / "GenosAI_OrganicLeads.xlsx"

print(f"Loading: {WB_PATH}")
wb = openpyxl.load_workbook(WB_PATH)


def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = cp.copy(src_cell.font)
        dst_cell.fill = cp.copy(src_cell.fill)
        dst_cell.border = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


# ── Lead data ──────────────────────────────────────────────────────────────────
leads = [
    {
        "num": 9,
        "company": "Kaden Investment",
        "contact": "General Enquiries",
        "email": None,
        "industry": "Real Estate Development & Asset Management",
        "score": "4/10",
        "subject": "kaden-sa.com serves Saudi HNW investors in Squarespace Arabic — Vision 2030's buyers expect more",
        "body": """\
Hi team,

I went through kaden-sa.com this week. The portfolio positioning — Mixed-Use, Logistics, Retail, Waterfront, Commercial — maps directly to Saudi Arabia's Vision 2030 construction pipeline. The fundamentals are strong.

Three gaps that will cost leads as Vision 2030 accelerates:

1. Mobile performance is 44/100. Saudi HNW buyers and institutional partners browse on mobile first. A 4+ second page load on kaden-sa.com hands those visitors to competitors before the Arabic content even renders.

2. The Arabic-English parity is incomplete. Arabic pages carry the full story; English pages don't. Overseas investors — the exact profile that Vision 2030 is attracting — hit incomplete content and leave. That's a direct trust deficit for a SAR 289K+ project inquiry.

3. There is no AI chatbot, no bilingual enquiry handling, and no automated investor follow-up. A Saudi HNW buyer browsing at 11pm doesn't wait for an email reply the next morning. The developer that responds immediately in Arabic — through an AI agent — wins the appointment.

We build AI systems for exactly this: a bilingual (AR/EN) AI agent on kaden-sa.com, investor and tenant portals with Azure AD authentication, automated Vision 2030 market intelligence reports, and an Algolia-powered bilingual property search. PDPL cookie consent is a one-day fix that removes the data protection exposure.

I put together a 12-slide audit deck specific to Kaden — PageSpeed targets, bilingual AI agent spec, investor portal scope, and a phased roadmap with indicative SAR pricing.

Happy to send it over or take 15 minutes on a call.

Worth a conversation?

Rohan Malik
CEO, Genos AI
hello@genosai.tech | +91 638-714-2699
www.genosai.tech""",
        "fu1": """\
Hi team,

Following up on my message from last week.

One data point: Saudi Arabia's Vision 2030 construction pipeline is accelerating, and bilingual digital infrastructure is now a procurement-level requirement for major institutional partners. Developers with Arabic AI agents and investor portals are qualifying larger transactions — and faster.

The audit deck I mentioned covers exactly what's needed on kaden-sa.com. Happy to send it across.

Rohan
Genos AI | hello@genosai.tech""",
        "fu2": """\
Hi team,

Last note from me on this.

If the timing isn't right, no problem. The deck is ready whenever the conversation is — just reply and I'll send it across.

Rohan
Genos AI | hello@genosai.tech""",
        "next_action": "Source decision-maker email via LinkedIn; send audit deck + bilingual AI agent pitch",
        "deal_value": "$20,000–$45,000 (bilingual AI agent + investor portal + Vision 2030 market intelligence)",
    },
    {
        "num": 10,
        "company": "NewExGen (Pty) Ltd",
        "contact": "Sales / Marketing Team",
        "email": None,
        "industry": "ICT / Digital Transformation & Managed Services",
        "score": "5/10",
        "subject": "You're a Zoho partner — the AI layer on top of it is the fastest win on your roadmap",
        "body": """\
Hi team,

I spent time on newexgen.co.za this week. 15 years, six service segments, Zoho Bookings live, CRM and Desk in place — the operational foundation is genuine. For an ICT transformation consultancy, the website undersells what the business actually does.

Three things that stood out:

1. Mobile PageSpeed is 51/100. You're pitching digital transformation to corporate clients and government departments — and your own site loads slowly on mobile. That's a first-impression inconsistency that procurement teams notice.

2. The demo booking flow breaks on mobile. Zoho Bookings is live but the booking path misfires on small screens — the exact journey a prospective client would take after seeing a LinkedIn post or a referral.

3. Zoho SalesIQ's Zia AI is sitting unused. As a Zoho partner with CRM, Desk, and SalesIQ already deployed, you're one configuration away from a 24/7 lead qualification agent that responds to inbound enquiries, books demos automatically, and scores leads in CRM before a human sees them. That's not a new tool purchase — it's switching on what you already have.

We build AI systems for Zoho-native businesses: Zia AI chatbot activation, AI-scored lead routing in CRM, AI ticket triage and SLA-breach prediction in Zoho Desk, and automated field-technician dispatch for managed services. The POPIA cookie consent and mobile booking fix are both same-day items.

I put together a 12-slide audit deck specific to NewExGen — Zoho fast-wins, mobile UX fixes, and a phased roadmap with ZAR pricing.

Happy to send it over or take 15 minutes on a call.

Worth a conversation?

Rohan Malik
CEO, Genos AI
hello@genosai.tech | +91 638-714-2699
www.genosai.tech""",
        "fu1": """\
Hi team,

Following up on my message from last week.

The fastest win here isn't a new system — it's activating Zia AI on your existing Zoho stack. For a Zoho partner already running CRM, Desk, and SalesIQ, this typically takes under a week and immediately improves lead response time and demo booking conversion.

Audit deck is ready. Happy to send it across.

Rohan
Genos AI | hello@genosai.tech""",
        "fu2": """\
Hi team,

Last note from me.

If the timing isn't right, I'll leave the deck available whenever it's useful. Just reply and I'll send it over.

Rohan
Genos AI | hello@genosai.tech""",
        "next_action": "Source Sales/Marketing Director email via LinkedIn; send audit deck — Zoho AI fast-wins pitch",
        "deal_value": "$12,000–$28,000 (Zoho AI activation + mobile UX + managed services automation)",
    },
    {
        "num": 11,
        "company": "Realtor CoJo (RC Realtor)",
        "contact": "General Enquiries",
        "email": None,
        "industry": "Real Estate — Luxury & Affordable Properties",
        "score": "2/10",
        "subject": "rc-realtor.vercel.app is invisible to Google — a custom domain fixes it in one day",
        "body": """\
Hi team,

I came across RC Realtor while researching Lagos property platforms. The luxury and affordable property positioning — targeting the Lagos market across both segments — is a smart differentiation play.

One critical issue that will block all growth until it's fixed: rc-realtor.vercel.app is a staging URL. Google treats staging subdomains as non-indexable. Combined with the Next.js CSR-only rendering, the current build returns effectively zero SEO-indexable content — the site title is the only thing Google can read. A buyer searching "luxury apartments Lagos" will never find RC Realtor, regardless of how good the listings are.

Three things that unblock the business:

1. Custom domain + branded email. Moving off vercel.app to a proper domain costs under $20/year and immediately signals credibility to Lagos HNW buyers who check whether the agency exists before making a call.

2. Next.js SSR/SSG migration. Server-side rendering makes every property listing indexable by Google. This is the single highest-ROI move — it turns a site that ranks for nothing into one that can rank for "2-bed Lekki apartment" within 60 days.

3. WhatsApp CTA on every listing. Lagos buyers don't fill out contact forms — they tap WhatsApp. A click-to-chat button on each listing, backed by an AI agent that qualifies budget and books viewings, captures the leads the current site is losing every day.

We build AI systems for exactly this: custom domain setup, SSR migration via Next.js 15 with Sanity CMS, WhatsApp AI agent for 24/7 lead qualification, and automated listing posts to Instagram and LinkedIn.

I put together a 12-slide audit deck specific to RC Realtor — SSR migration scope, WhatsApp AI agent flow, and a phased roadmap with NGN/USD pricing.

Happy to send it over or take 15 minutes on a call.

Worth a conversation?

Rohan Malik
CEO, Genos AI
hello@genosai.tech | +91 638-714-2699
www.genosai.tech""",
        "fu1": """\
Hi team,

Following up on my message from last week.

The domain + SSR migration point is worth repeating: as long as RC Realtor is on rc-realtor.vercel.app, it cannot rank on Google — regardless of listing quality or marketing spend. That's a fixable constraint, typically resolved in under 2 weeks.

Audit deck is ready. Happy to send it over.

Rohan
Genos AI | hello@genosai.tech""",
        "fu2": """\
Hi team,

Last one from me.

If the timing isn't right, no problem. The deck is there when you need it — just reply and I'll send it across.

Rohan
Genos AI | hello@genosai.tech""",
        "next_action": "Source founder contact via RC Realtor LinkedIn; lead with domain + SSR migration pitch",
        "deal_value": "$8,000–$18,000 (Next.js SSR migration + WhatsApp AI agent + SEO rebuild)",
    },
    {
        "num": 12,
        "company": "Idili Africa",
        "contact": "General Enquiries",
        "email": None,
        "industry": "International Talent Mobility & Graduate Recruitment",
        "score": "3/10",
        "subject": "SA graduates placed in Sweden — and the candidate pipeline is a 4-page website",
        "body": """\
Hi team,

I spent time on idili.africa this week. The mission — routing South African graduates (18–25) into Saab, Atlas Copco, and Swedish partner companies — is a genuinely differentiated offer in the graduate recruitment space. The Oakvale Invest and Sweden Consulate Cape Town affiliations are strong proof of concept.

The infrastructure doesn't match the ambition.

Three things that are limiting growth right now:

1. The site is four pages with fragile URLs that 404 mid-navigation. A prospective graduate who clicks through from your LinkedIn post and hits a broken page doesn't apply — they close the tab. With no ATS or candidate portal, there's no other way to capture that interest.

2. There is no candidate tracking system. When a graduate submits interest, what happens next? Without an ATS, every step from application to shortlist to partner match to placement is manual. That process doesn't scale to the volume Idili Africa needs to be credible to partners like Saab.

3. The opportunity listings have no SEO structure. A graduate searching "graduate jobs South Africa Sweden" or "international placement programme SA" will never find idili.africa — the current site has no JobPosting schema and virtually no indexable content.

We build AI systems for exactly this: a candidate intake portal with CV parsing and AI skills extraction, an automated WhatsApp candidate journey (eligibility check → document upload → assessment booking → status updates), a partner dashboard for Saab and Atlas Copco, and SEO-optimised opportunity listings that rank on Google.

I put together a 12-slide proposal specific to Idili Africa — ATS scope, candidate AI flow, partner portal spec, and a phased roadmap with ZAR/USD pricing.

Happy to send it over or take 15 minutes on a call.

Worth a conversation?

Rohan Malik
CEO, Genos AI
hello@genosai.tech | +91 638-714-2699
www.genosai.tech""",
        "fu1": """\
Hi team,

Following up on my note from last week.

The ATS gap is the most immediate constraint — without a candidate tracking system, Idili Africa's growth is bottlenecked by manual coordination. An AI-assisted intake portal that parses CVs and auto-qualifies graduates against partner requirements is the single fastest unlock.

Proposal is ready. Happy to send it across.

Rohan
Genos AI | hello@genosai.tech""",
        "fu2": """\
Hi team,

Last note from me.

If the timing isn't right, no problem. The proposal is ready whenever the conversation is useful — just reply and I'll send it across.

Rohan
Genos AI | hello@genosai.tech""",
        "next_action": "Source founder/Talent Director contact via LinkedIn or idili.africa; send proposal deck — ATS + candidate AI pitch",
        "deal_value": "$15,000–$35,000 (candidate portal ATS + AI screening + partner dashboard + SEO)",
    },
    {
        "num": 13,
        "company": "Shalom Park Estate (IFT Realty Ltd)",
        "contact": "Oluremi Oshikanlu",
        "email": None,
        "industry": "Real Estate — Luxury Estate Development",
        "score": "4/10",
        "subject": "Shalom Park: luxury Lekki units with no pricing on the website — diaspora buyers won't call",
        "body": """\
Hi Oluremi,

I went through shalomparknigeria.com this week. 25 acres, Lekki-Epe axis, Lagos State Governor's Consent, Megacity Plan approval — the legal foundation for Shalom Park is stronger than most Nigerian luxury developments I've reviewed.

Three things are directly blocking diaspora conversions:

1. There is no pricing on the website. Not a starting figure, not a "from ₦X" marker, not a USD equivalent. A US or UK Nigerian browsing at 11pm New York time will not fill out a contact form to ask what a 4-bed semi-detached costs — they'll move to the next developer who shows numbers. In the diaspora market, hiding pricing signals uncertainty.

2. There is no WhatsApp CTA on any listing. Nigerian and diaspora buyers use WhatsApp before they use anything else. A "Chat on WhatsApp" button on every unit page — backed by an AI agent that sends brochures, answers questions on unit type, and books site visits — captures the inquiry before it goes to a competitor.

3. The realtor onboarding runs on Google Meet. For a developer positioning at luxury price points, sending partner realtors through a generic video call for onboarding is a brand mismatch. A dedicated realtor portal with commission tracking, marketing collateral downloads, and automated lead assignment signals the professionalism that attracts top Lagos agents.

We build AI systems for exactly this: a WhatsApp AI agent for 24/7 buyer qualification, automated diaspora investor sequences (US/UK Nigerians in USD), a partner realtor portal, and monthly construction progress automations that keep buyers warm from deposit to handover.

I put together a 12-slide audit deck specific to Shalom Park — pricing strategy, diaspora buyer journey, realtor portal scope, and a phased roadmap with NGN/USD pricing.

Happy to send it over or take 15 minutes on a call.

Worth a conversation?

Rohan Malik
CEO, Genos AI
hello@genosai.tech | +91 638-714-2699
www.genosai.tech""",
        "fu1": """\
Hi Oluremi,

Following up on my message from last week.

The pricing gap is the single fastest thing to fix — adding "from ₦X / from $X" to each unit page typically increases diaspora inquiry conversion by 30–40% within the first month. Everything else builds on that.

Audit deck is ready. Happy to send it across.

Rohan
Genos AI | hello@genosai.tech""",
        "fu2": """\
Hi Oluremi,

Last one from me.

If the timing isn't right, no problem. The deck covers pricing strategy, diaspora buyer flows, and the realtor portal in full — I'll leave it available whenever the conversation is useful.

Rohan
Genos AI | hello@genosai.tech""",
        "next_action": "Source Oluremi Oshikanlu email via IFT Realty website or LinkedIn; send audit deck — diaspora buyer flow + pricing pitch",
        "deal_value": "$12,000–$28,000 (WhatsApp AI + diaspora campaign + realtor portal + construction update automation)",
    },
    {
        "num": 14,
        "company": "Fortress Real Estate Investments Limited",
        "contact": "Steven Brown",
        "email": None,
        "industry": "REIT — Logistics & Convenience Retail (JSE: FFA & FFB)",
        "score": "5/10",
        "subject": "R29.6B REIT, JSE-listed — and fortressfund.co.za is invisible to institutional screeners",
        "body": """\
Hi Steven,

I spent time on fortressfund.co.za this week. R29.6B market cap, 2.9M sqm GLA, 52 retail centres, logistics across SA and Central Eastern Europe, 23.9% stake in NEPI Rockcastle — the portfolio is among the most diversified in the JSE REIT universe.

The digital infrastructure doesn't reflect the scale.

Three specific issues:

1. fortressfund.co.za is a JavaScript-rendered SPA. Institutional ESG screeners, Bloomberg terminal plugins, and JSE data aggregators that parse pages programmatically see nothing — the content is invisible until JavaScript executes, which automated tools typically don't wait for. Every SENS announcement, property listing, and ESG disclosure in the current build is effectively unindexed by the tools that institutional investors actually use.

2. The IR experience lags global REIT peers. Growthpoint, Redefine, and Emira all publish SENS feeds, live share-price widgets, and one-click dividend history on their IR pages. Fortress's equivalent requires investors to navigate multiple clicks with no live data surface. For FFA and FFB holders, that friction is a retention and engagement gap.

3. There is no AI layer on the investor or leasing side. An IR chatbot trained on Fortress SENS filings, annual reports, and dividend history would handle 80% of analyst inquiries — NAV per unit, yield calculations, lease expiry profiles — without management time. On the leasing side, a searchable GLA availability portal that qualifies inbound tenants before your BD team engages would materially reduce time-to-lease on logistics vacancies.

We build AI systems for institutional-grade real estate platforms: SSR migration to Next.js 15 (making every page Google and Bloomberg-indexable), a live IR hub with SENS aggregation and share-price tickers, an IR chatbot trained on Fortress filings, a logistics and retail leasing portal, and ESG dashboard automation for TCFD/GRESB submissions.

I put together a 14-slide proposal specific to Fortress — SSR migration scope, IR hub spec, AI chatbot training data plan, and a phased roadmap with indicative ZAR pricing.

Happy to send it over or take 20 minutes on a call.

Worth a conversation?

Rohan Malik
CEO, Genos AI
hello@genosai.tech | +91 638-714-2699
www.genosai.tech""",
        "fu1": """\
Hi Steven,

Following up on my message from last week.

The SSR migration point is worth underlining: as long as fortressfund.co.za runs as a JS-rendered SPA, every piece of content — SENS filings, property data, ESG disclosures — is invisible to the programmatic tools institutional investors and screeners rely on. That's a resolvable infrastructure gap, typically addressed in 6–8 weeks for a site of this scope.

Proposal is ready. Happy to send it across.

Rohan
Genos AI | hello@genosai.tech""",
        "fu2": """\
Hi Steven,

Last note from me.

If the timing isn't right, no problem. The proposal covers the IR hub, AI chatbot, leasing portal, and ESG automation in full — available whenever the conversation is useful.

Rohan
Genos AI | hello@genosai.tech""",
        "next_action": "Source Steven Brown email via JSE filing / Fortress investor relations page; send proposal deck — IR hub + SSR migration pitch",
        "deal_value": "$50,000–$120,000+ (SSR migration + IR hub + AI chatbot + leasing portal — flagship JSE engagement)",
    },
]

# ── Insert 6 blank rows at row 10 in Cold Emails and Pipeline ─────────────────
ws_email = wb["Cold Emails"]
ws_pipe = wb["Pipeline"]

ws_email.insert_rows(10, 6)
ws_pipe.insert_rows(10, 6)

# ── Write Cold Emails rows 10-15 ───────────────────────────────────────────────
for i, lead in enumerate(leads):
    row = 10 + i
    style_src_row = 9  # copy style from row 9 (lead 8 / Tokuti)
    vals = [
        lead["num"], lead["company"], lead["contact"],
        lead["email"], lead["subject"], lead["body"],
        lead["fu1"], lead["fu2"],
    ]
    for col, v in enumerate(vals, 1):
        cell = ws_email.cell(row, col)
        cell.value = v
        copy_style(ws_email.cell(style_src_row, col), cell)

# ── Write Pipeline rows 10-15 ─────────────────────────────────────────────────
for i, lead in enumerate(leads):
    row = 10 + i
    style_src_row = 9
    vals = [
        lead["num"], lead["company"], lead["contact"],
        lead["email"], lead["industry"], lead["score"],
        "To Contact", "—", lead["next_action"], lead["deal_value"],
    ]
    for col, v in enumerate(vals, 1):
        cell = ws_pipe.cell(row, col)
        cell.value = v
        copy_style(ws_pipe.cell(style_src_row, col), cell)

wb.save(WB_PATH)
print("Done — leads 9-14 inserted into Cold Emails and Pipeline sheets.")
