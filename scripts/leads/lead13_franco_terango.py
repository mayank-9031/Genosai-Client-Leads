from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import copy as cp
import os

ROOT = Path(__file__).parent.parent.parent
WB_DIR = ROOT / "output" / "workbooks"

import glob
files = glob.glob(str(WB_DIR / "GenosAI_RealEstate_Leads*.xlsx"))
src = next((f for f in files if "UPDATED" in f), str(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx"))
print(f"Loading: {src}")
wb = load_workbook(src)

lead = {
    "num": 102,
    "first": "Franco",
    "last": "Terango",
    "full": "Franco Terango",
    "title": "CEO",
    "company": "Certainty Home Lending",
    "email": "franco.terango@certaintyhomelending.com",
    "email_status": "Estimated",
    "website": "https://www.certaintyhomelending.com/",
    "phone": "",
    "city": "Los Angeles",
    "state": "CA",
    "country": "US",
    "employees": "201-500",
    "industry": "Mortgage Lending / Home Loans",
    "linkedin": "https://www.linkedin.com/company/certaintyhomelending/",
    "ppt": "CertaintyHomeLending_Audit_GenosAI.pptx"
}

subject = "25 years building Certainty, 10,671 five-star reviews, and a homepage that shows none of them"

body = """\
Hi Franco,

Read your HousingWire piece on AI changing the LO role - and the Scotsman Guide follow-up on reducing administrative burden through technology. The thinking is sharp and clearly informed by 25 years of watching how the industry actually operates rather than how it's supposed to.

One thing stood out when I looked at certaintyhomelending.com. You've built 10,671 reviews at 4.9 stars on RaveCapture - that's a genuinely exceptional trust signal. None of it is visible to a prospect landing on your homepage at 9pm on a Sunday. They see a Find a Loan Officer button and a calculator. The reviews that would convert them are sitting on a third-party platform they'll never visit.

There's a related gap in the same place. A prospect landing after hours has no way to engage, qualify themselves, or get a useful response until someone picks up on Monday. For a company covering 34 states with 200-plus employees, that's a lot of Sunday evening interest going cold before a competitor's AI chatbot answers it in 60 seconds.

I'm Rohan Malik, CEO of Genos AI. We work with mortgage lenders on two things: website development (surfacing social proof correctly, building conversion flows that work without an LO in the room) and AI automation (chatbots that pre-qualify, rate alert sequences, post-close review and referral workflows). Ran a quick audit for Certainty and thought it might be worth 15 minutes.

Happy to walk through it whenever suits you.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi Franco, just checking my last message didn't get buried. The specific gap I flagged - 10,671 reviews invisible to homepage visitors - is a quick fix that moves conversion on existing traffic. Happy to walk through the full audit in 15 minutes. - Rohan | Genos AI\
"""

fu2 = """\
Franco, last one from me. The audit for Certainty is ready whenever the timing is right. The AI pre-qual chatbot and rate alert automation are probably the fastest wins for a 34-state operation. Leaving it here in case it's useful down the road. - Rohan | Genos AI\
"""

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = cp.copy(src_cell.font)
        dst_cell.fill = cp.copy(src_cell.fill)
        dst_cell.border = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

# Leads
ws = wb["Leads"]
row = 103
vals = [lead["num"], lead["first"], lead["last"], lead["full"], lead["title"],
        lead["company"], lead["email"], lead["email_status"], lead["website"],
        lead["phone"], lead["city"], lead["state"], lead["country"],
        lead["employees"], lead["industry"], lead["linkedin"], lead["ppt"]]
for col, v in enumerate(vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# Website Audits
ws2 = wb["Website Audits"]
audit_vals = [
    lead["num"], lead["company"], lead["website"],
    "Custom / Guaranteed Rate Platform", "Yes", "Yes", "Partial",
    "No", "RaveCapture (offsite)", "Yes (10,671 reviews - offsite only)", "No", "Partial (press only)", 6, "High",
    "Website is professionally designed but underperforms on conversion and social proof: 10,671 five-star RaveCapture reviews are not embedded on homepage or loan pages - invisible to prospects; no AI chatbot or live chat means after-hours traffic (Sunday evening, weekend) gets zero response while competitors answer in under 60 seconds; no rate quote or rate alert tool visible to top-of-funnel visitors; Find a Loan Officer CTA is a directory dead end with no intelligent routing by loan type, geography, or borrower profile; no SEO content strategy - no blog, no pillar pages, no long-tail keyword targeting despite covering 34 states; no video content or personalized LO video messaging; no visible lead nurture email capture for not-ready-yet prospects",
    "Embed RaveCapture reviews above the fold on homepage and loan type pages; replace Find a Loan Officer directory with AI chatbot that pre-qualifies by loan type, geography, credit range, and timeline before routing to specific LO with context; deploy rate alert subscription for prospects not ready to transact - automated email/SMS when rates drop to their threshold; build SEO content layer targeting state-specific and loan-type long-tail keywords across 34 licensed states; implement post-close review and referral automation sequence (D+3, D+7, D+30) to systematically amplify review volume and generate warm referrals at the emotional high point of closing; add LO-level video messaging capability for personalized prospect and pre-approval communication"
]
row2 = 103
for col, v in enumerate(audit_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# Cold Emails
ws3 = wb["Cold Emails"]
email_vals = [lead["num"], lead["first"], lead["company"], lead["email"], subject, body, fu1, fu2]
row3 = 103
for col, v in enumerate(email_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

# Audit PPT Content
ws4 = wb["Audit PPT Content"]
ppt_vals = [
    lead["num"], lead["company"], lead["full"], lead["email"], lead["website"], 6, "High",
    "Certainty Home Lending - Website & AI Growth Audit | Prepared by Genos AI",
    "Franco Terango: CEO of Certainty Home Lending since April 2023 | 35+ years in financial services, 25 years mortgage banking | BA Economics, George Mason University 1987 | Greater Los Angeles area | Board of Trustees, Greater Los Angeles Zoo Association (GLAZA) since March 2019 | Published author: HousingWire (The Role of the LO Will Soon Be Very Different), Scotsman Guide (Jan 2024 - The Seismic Shift Occurring in Your Workplace; Oct 2025 - A Primer for Success by Empowering Your Lending Team), MBA NewsLink (Dec 2024) | Scotsman Guide Author Showcase Podcast co-host with Jim Clapp (March 2024) | Certainty Home Lending: national mortgage lender, DBA of Guaranteed Rate (NMLS #2611, #2 US retail mortgage lender), founded 2000 as WR Starkey Mortgage, rebranded circa 2016 | 201-500 employees | Licensed 34 states + DC | 10,671 reviews at 4.9 stars (RaveCapture) | MortgageCX Best-In-Class Lender 5 consecutive years (2018-2022) | Loan types: Conventional, FHA, VA, Jumbo, ARMs, HELOC, Renovation, Construction, Down Payment Assistance | Spanish-language digital mortgage launched March 2026 | Parent Rate Intelligence AI platform launched Oct 2024 ($100M investment, reduces approval time from weeks to minutes)",
    "Published thought leader on AI reducing LO administrative burden - Franco has already done the ideological work; Certainty benefits from parent Rate's $100M Rate Intelligence AI infrastructure investment (real-time underwriting, SmartUnderwrite income/employment verification); 10,671 five-star reviews at 4.9 on RaveCapture is a trust signal that most mortgage lenders cannot match - the asset exists and is not being deployed on the website; Spanish-language digital mortgage launch (March 2026) demonstrates active investment in UX and borrower experience innovation; 34-state licensing and 200-plus employee team means operational scale exists to handle automation-generated volume; MortgageCX Best-In-Class 5 consecutive years demonstrates sustained service quality that automation would amplify rather than compromise; Jim Clapp as President with 30 years experience provides institutional continuity; actively growing executive team (5 senior hires in 2024-2025) signals expansion mode",
    "10,671 five-star reviews invisible on homepage - prospects see Find a Loan Officer and a calculator, social proof sits on RaveCapture offsite; no AI chatbot or live chat - after-hours prospects (Sunday evening, weekends) get zero response; competitors using AI lead response tools answer in under 90 seconds; no rate alert or rate watch tool for top-of-funnel visitors who are not yet ready to transact; Find a Loan Officer is a directory dead end with no intelligent routing by loan type, geography, credit score, or timeline; no SEO content strategy - no blog, no pillar pages, no long-tail keyword targeting across 34 licensed states; no video content or LO personalized video messaging (industry standard at top-producing offices); no visible lead nurture email capture for prospects who research but do not apply; Agent Advantage portal for Realtor partners exists but has no AI features for co-marketing or lead sharing",
    "1. Embed RaveCapture reviews above fold on homepage and loan pages - convert existing trust signal into visible conversion asset | 2. AI chatbot for pre-qualification and intelligent LO routing - 5-7 qualifying questions, routes to right LO with context, captures all after-hours traffic | 3. Rate alert subscription automation - builds prospect pipeline for when rates hit borrower thresholds | 4. Post-close review and referral sequence (D+3, D+7, D+30) - amplifies review volume and generates warm referrals at emotional close high point | 5. SEO content engine - state-specific and loan-type long-tail keyword targeting across 34 licensed states | 6. LO video messaging capability - personalized video for prospect outreach and pre-approval delivery",
    "Franco has written publicly that user-friendly tools reducing administrative burden are the defining factor in LO recruitment and retention - and that the LO role will soon be fundamentally different because of technology. The website of the company he leads sends after-hours prospects to a loan officer directory and a mortgage calculator. 10,671 five-star reviews sit on a third-party platform that prospects will never visit. The gap between what Franco writes and what certaintyhomelending.com does today is where Genos AI operates.",
    "Website: embed reviews above fold, add rate alert subscription, replace LO directory with AI routing chatbot, SEO content layer for 34 licensed states | AI automation: pre-qualification chatbot with intelligent LO routing, post-close review and referral sequence, rate drop alert system for pipeline prospects, LO prospecting automation for Realtor relationship management | Implementation: pilot one market or loan type first (Franco's stated strategic preference), measure conversion lift and LO time saved, scale to full 34-state operation",
    "Month 1: Reviews embedded above fold on homepage and top loan pages; AI chatbot deployed with pre-qualification flow and LO routing logic trained on loan types and 34-state geography; rate alert subscription capture live on homepage | Month 2: Post-close review and referral automation sequences active for all closed loans; SEO content engine producing first state-specific and loan-type articles; LO video messaging capability live for 10 pilot LOs | Month 3: SEO articles indexing across target keywords; chatbot pre-qualification data informing LO routing improvements; referral sequence generating measurable warm referral volume; pilot results packaged for full-team rollout presentation",
    "Reply to this email or book a 15-min call | Rohan Malik, CEO | hello@genosai.tech | www.genosai.tech"
]
row4 = 103
for col, v in enumerate(ppt_vals, 1):
    cell = ws4.cell(row4, col)
    cell.value = v
    copy_style(ws4.cell(row4 - 1, col), cell)

# Organic Leads
ws5 = wb["Organic Leads"]
thin = Side(style="thin", color="444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
row_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
row_font = Font(name="Calibri", size=10, color="1A1A2E")

organic_vals = [
    13, "Franco", "Terango", "Franco Terango", "CEO",
    "Certainty Home Lending", "franco.terango@certaintyhomelending.com",
    "Estimated", "https://www.certaintyhomelending.com/", "",
    "Los Angeles", "CA", "US", "201-500", "Mortgage Lending / Home Loans",
    "https://www.linkedin.com/company/certaintyhomelending/", "Organic/Manual",
    "2026-05-10", "CertaintyHomeLending_Audit_GenosAI.pptx", "Email Sent - Pending Reply"
]
for col, v in enumerate(organic_vals, 1):
    cell = ws5.cell(14, col)
    cell.value = v
    cell.font = row_font
    cell.fill = row_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border

for col in range(1, ws5.max_column + 1):
    if ws5.cell(1, col).value == "Cold Email Body":
        cell = ws5.cell(14, col)
        cell.value = body
        cell.font = row_font
        cell.fill = row_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        break

ws5.row_dimensions[14].height = 180

wb.save(WB_DIR / "GenosAI_RealEstate_Leads_UPDATED.xlsx")
print("Excel updated - Franco Terango / Certainty Home Lending added to all 5 sheets.")
