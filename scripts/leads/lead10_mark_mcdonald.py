from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import copy as cp
import os

ROOT = Path(__file__).parent.parent.parent
WB_DIR = ROOT / "output" / "workbooks"

import glob
files = glob.glob(str(WB_DIR / "GenosAI_RealEstate_Leads*.xlsx"))
src = next((f for f in files if "UPDATED" in f), str(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx"))
print(f"Loading: {src}")
wb = load_workbook(src)

lead = {
    "num": 99,
    "first": "Mark",
    "last": "McDonald",
    "full": "Mark McDonald",
    "title": "President",
    "company": "CoStar Real Estate Manager",
    "email": "mark.mcdonald@costar.com",
    "email_status": "Verify",
    "website": "https://costarmanager.com/",
    "phone": "",
    "city": "Atlanta",
    "state": "GA",
    "country": "US",
    "employees": "201-500",
    "industry": "Real Estate Tech - Lease Management / SaaS",
    "linkedin": "https://www.linkedin.com/in/mark-mcdonald-77b5887/",
    "ppt": "CoStarManager_Audit_GenosAI.pptx"
}

subject = "15 years from Virtual Premise to Fortune 500 compliance, and a demo that still asks them to fill out a form"

body = """\
Hi Mark,

From joining Virtual Premise in the startup days through the hyper-growth phase to President of CoStar Real Estate Manager, guiding Fortune 500 companies through ASC 842 and IFRS 16 compliance is not a simple product category. 15 years of accumulated expertise in a market where the compliance clock is real and the switching cost is high.

I'm Rohan Malik, CEO of Genos AI. We help B2B platforms like yours turn that kind of product depth into more inbound pipeline through AI tools and a website that does something when a corporate real estate director lands on it at 9pm.

Ran a free audit of costarmanager.com and attached it. A few things stood out.

A VP of Corporate Real Estate at a 200-location retail chain is evaluating lease management platforms in January, under pressure to close ASC 842 compliance before the fiscal year audit. They land on costarmanager.com, read about unified lease accounting and administration, want to know what it costs for their portfolio size, and hit a form. Demo request. Someone will be in touch. They close the tab and request a demo from LeaseQuery in the same session.

The product has the CoStar data moat, the Visual Lease acquisition, 15 years of implementation depth, and Fortune 500 references. None of that is on the page when a prospect needs to make a shortlist at 11pm.

A few things worth a look:
- ROI calculator on the homepage. Prospects input portfolio size, lease count, and current compliance workflow. The calculator returns estimated savings vs. manual process and time-to-close-advantage for ASC 842 audit cycle. Captures the lead and pre-qualifies by deal size before any sales conversation
- AI chat trained on CoStar Real Estate Manager's product, ASC 842/IFRS 16 requirements, and integration landscape. Handles prospect questions after hours, books demos, routes by accounting team vs. real estate team vs. retail tenant persona
- Interactive product tour without form gating. Let a VP of Corporate Real Estate walk through the lease accounting module and the CoStar market data layer before committing to a 45-minute demo call
- Role-specific landing pages for accounting teams, real estate teams, and retail tenants. Each persona has a different pain, a different compliance timeline, and a different ROI story. One homepage doesn't convert all three
- Quantified case studies. Current testimonials on site have no numbers. The Hertz reference is there but John Grotto's quote doesn't say what it saved or how fast. Every Fortune 500 reference needs a before-and-after metric
- Compliance deadline nurture sequences. ASC 842 and IFRS 16 adoption windows are finite but new companies come into scope, get acquired, or change auditors. A prospect in month one of compliance research is a different conversation than one six months from audit. Automated sequences that route by compliance stage keep CoStar Real Estate Manager in front of the right buyers at the right time

On the website side, we'd rebuild to reflect what the CoStar data moat and the Visual Lease acquisition actually mean for a compliance-pressured buyer. ROI as a homepage CTA, product tour before the demo gate, role-based pages that speak to the specific pain accounting versus real estate teams bring to the platform.

Audit's attached, no strings. Happy to walk through it in 15 minutes.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi Mark, just making sure the audit didn't get buried. The ROI calculator and the role-specific landing pages are probably the most relevant for converting compliance-pressured buyers who land on costarmanager.com before they've scheduled a demo. Happy to walk through it in 15 minutes. - Rohan\
"""

fu2 = """\
Hi Mark, last one from me. ASC 842 brought a wave of new buyers into lease management software. The next wave is companies that adopted compliance tools in 2022 and 2023 that are now re-evaluating at contract renewal. The nurture sequences and product tour automation we'd build are what capture that cohort before they shortlist. Leaving the audit here in case timing improves. - Rohan\
"""

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = cp.copy(src_cell.font)
        dst_cell.fill = cp.copy(src_cell.fill)
        dst_cell.border = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

# Leads
ws = wb["Leads"]
row = 100
vals = [lead["num"], lead["first"], lead["last"], lead["full"], lead["title"],
        lead["company"], lead["email"], lead["email_status"], lead["website"],
        lead["phone"], lead["city"], lead["state"], lead["country"],
        lead["employees"], lead["industry"], lead["linkedin"], lead["ppt"]]
for col, v in enumerate(vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# Website Audits
ws2 = wb["Website Audits"]
audit_vals = [
    lead["num"], lead["company"], lead["website"],
    "Custom SaaS / React (costarmanager.com)", "Yes", "Unknown", "Unknown",
    "No", "None Detected", "Unknown", "Unknown", "Yes (blog, webinars, podcast)", 5, "High",
    "No pricing page or pricing transparency - demo request required to learn cost for any portfolio size; no ROI calculator or self-assessment tool for compliance savings; demo is fully form-gated with no interactive product tour or instant walkthrough; no AI chat for 24/7 prospect or customer questions; no role-specific landing pages (accounting team vs. real estate team vs. retail tenant have different pains, one homepage does not convert all three); testimonials have no quantified results (Hertz reference present but no before/after metrics); no customer count or 'trusted by X companies' statement; no Gartner/Forrester analyst recognition surfaced; no compliance deadline nurture sequences visible; no interactive product demo or sandbox access; no FAQ addressing objections or competitor comparison",
    "ROI calculator on homepage: prospect inputs portfolio size, lease count, compliance workflow and gets estimated savings vs. manual process + ASC 842 audit cycle advantage; AI chat trained on platform, ASC 842/IFRS 16, integration landscape for 24/7 prospect routing and demo booking; interactive product tour without form gating - let buyer walk through lease accounting + CoStar market data layer before committing to a sales call; role-specific landing pages for accounting teams, real estate teams, and retail tenants with persona-specific ROI stories; quantified case studies with before/after metrics for every Fortune 500 reference; compliance deadline nurture sequences routing by compliance stage; buyer alert sequences for ASC 842 contract renewal cohort (2022-2023 adopters coming up for re-evaluation); site rebuild to surface CoStar data moat and Visual Lease acquisition value for compliance-pressured buyers"
]
row2 = 100
for col, v in enumerate(audit_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# Cold Emails
ws3 = wb["Cold Emails"]
email_vals = [lead["num"], lead["first"], lead["company"], lead["email"], subject, body, fu1, fu2]
row3 = 100
for col, v in enumerate(email_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

# Audit PPT Content
ws4 = wb["Audit PPT Content"]
ppt_vals = [
    lead["num"], lead["company"], lead["full"], lead["email"], lead["website"], 5, "High",
    "CoStar Real Estate Manager - Website & AI Growth Audit | Prepared by Genos AI",
    "Mark McDonald: President at CoStar Real Estate Manager (formerly Virtual Premise) | 15-year tenure from startup professional services through hyper-growth to President | Previous roles: VP Strategy and Customer Success, Sales, Implementation, Product | Georgia Tech BS Mechanical Engineering + MIT Sloan MBA | Guides Fortune 500 companies through digital transformation of real estate portfolios and lease accounting compliance | Speaker at CoreNet Global and National Retail Tenants Association on ASC 842/IFRS 16 | Contributor to Accounting Today and Propmodo | Also oversees Visual Lease (acquired by CoStar) | Focus: profitable growth of CoStar Real Estate Manager customer base + product leadership in lease administration and accounting markets | Atlanta, GA | Platform serves corporate real estate accounting teams, real estate teams, and retail tenants managing office and retail portfolios",
    "15-year institutional depth from Virtual Premise startup through CoStar acquisition - rare in enterprise SaaS; CoStar data moat is a genuine differentiator - only lease management platform with integrated proprietary market data and analytics; Visual Lease acquisition expands addressable market and deepens accounting compliance specialisation; strong compliance-driven demand: ASC 842/IFRS 16 mandates create finite buying windows with real deadline pressure; Fortune 500 customer base with Hertz reference and enterprise implementation track record; multi-persona platform covering accounting teams, real estate teams, and retail tenants; thought leadership presence: CoreNet Global, National Retail Tenants Association speaking, Accounting Today and Propmodo publishing; podcast (The Lease Alert) and webinar programme demonstrates content engine",
    "No pricing page or transparency: demo required to access any cost information, creating friction for compliance-pressured buyers who need to shortlist quickly; no ROI calculator or self-assessment tool despite savings being a core product value proposition; demo fully form-gated with no interactive product tour or instant walkthrough; no AI chat or live chat for 24/7 prospect or customer questions; no role-specific landing pages despite accounting teams, real estate teams, and retail tenants having fundamentally different buying triggers and ROI stories; testimonials have no quantified results (Hertz reference present, no before/after metrics on time-to-close, savings, or audit cycle improvement); no customer count or analyst recognition (Gartner/Forrester) surfaced on homepage; no compliance deadline nurture sequences or re-engagement for 2022-2023 ASC 842 adopters approaching contract renewal",
    "1. ROI calculator: portfolio size + lease count + current compliance workflow = savings vs. manual and ASC 842 audit cycle advantage | 2. AI chat trained on platform features, ASC 842/IFRS 16 requirements, integration landscape - 24/7 demo booking and persona routing | 3. Interactive product tour without form gate: lease accounting + CoStar market data layer walkthrough before sales call | 4. Role-specific landing pages: accounting team (compliance + audit close), real estate team (portfolio visibility + transaction workflow), retail tenant (lease optimisation + renewal) | 5. Quantified case studies: every Fortune 500 reference gets a before/after metric | 6. Compliance deadline nurture sequences routing by ASC 842 compliance stage | 7. Re-engagement sequences for 2022-2023 adopters at contract renewal window",
    "A VP of Corporate Real Estate at a 200-location retail chain is evaluating lease management platforms in January under audit pressure. Lands on costarmanager.com, reads about unified lease accounting and CoStar data integration, wants to know what it costs for their portfolio size, hits a form. Demo request. Someone will be in touch. They request a demo from LeaseQuery in the same session. The CoStar data moat, the Visual Lease acquisition, 15 years of Fortune 500 implementation depth - none of it is on the page when a compliance-deadline buyer needs to make a shortlist at 11pm.",
    "ROI calculator on homepage: portfolio inputs return estimated compliance savings and ASC 842 audit cycle advantage - captures lead and pre-qualifies by deal size | AI chat trained on CoStar Real Estate Manager's platform, compliance standards, and integration landscape - 24/7 routing, demo booking, persona identification | Interactive product tour: lease accounting + CoStar market data walkthrough without form gate | Role-specific landing pages: accounting (audit close, ASC 842), real estate (portfolio, transactions), retail tenant (lease optimisation, renewals) | Quantified case studies: before/after metrics on every Fortune 500 reference | Compliance nurture sequences by stage: research, evaluation, adoption, renewal | Re-engagement for 2022-2023 ASC 842 cohort approaching contract renewal | Site rebuild: CoStar data moat and Visual Lease acquisition surfaced as homepage differentiators",
    "Month 1: ROI calculator live on costarmanager.com homepage and pricing page, AI chat deployed for 24/7 prospect routing and demo booking, interactive product tour launched without form gate, site rebuild begins with role-specific landing pages | Month 2: Accounting team, real estate team, and retail tenant landing pages live with persona-specific ROI stories; quantified case studies published for Hertz and top Fortune 500 references; compliance deadline nurture sequences active | Month 3: Re-engagement sequences running for 2022-2023 ASC 842 adopters at contract renewal window; buyer alert sequences for new compliance scope entrants; measurable inbound demo requests from ROI calculator; full site reflecting CoStar data moat and Visual Lease acquisition value for compliance-pressured buyers",
    "Reply to this email or book a 15-min call | Rohan Malik, CEO | hello@genosai.tech | www.genosai.tech"
]
row4 = 100
for col, v in enumerate(ppt_vals, 1):
    cell = ws4.cell(row4, col)
    cell.value = v
    copy_style(ws4.cell(row4 - 1, col), cell)

# Organic Leads
ws5 = wb["Organic Leads"]
thin = Side(style="thin", color="444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
row_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
row_font = Font(name="Calibri", size=10, color="1A1A2E")

organic_vals = [
    10, "Mark", "McDonald", "Mark McDonald", "President",
    "CoStar Real Estate Manager", "mark.mcdonald@costar.com",
    "Verify", "https://costarmanager.com/", "",
    "Atlanta", "GA", "US", "201-500", "Real Estate Tech - Lease Management / SaaS",
    "https://www.linkedin.com/in/mark-mcdonald-77b5887/", "Organic/Manual",
    "2026-05-10", "CoStarManager_Audit_GenosAI.pptx", "Email Sent - Pending Reply"
]
for col, v in enumerate(organic_vals, 1):
    cell = ws5.cell(11, col)
    cell.value = v
    cell.font = row_font
    cell.fill = row_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border

for col in range(1, ws5.max_column + 1):
    if ws5.cell(1, col).value == "Cold Email Body":
        cell = ws5.cell(11, col)
        cell.value = body
        cell.font = row_font
        cell.fill = row_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        break

ws5.row_dimensions[11].height = 180

wb.save(WB_DIR / "GenosAI_RealEstate_Leads_UPDATED.xlsx")
print("Excel updated - Mark McDonald / CoStar Real Estate Manager added to all 5 sheets.")
