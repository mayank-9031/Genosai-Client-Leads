from pathlib import Path
import copy as cp
from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent.parent
WB_PATH = ROOT / "output" / "workbooks" / "GenosAI_OrganicLeads.xlsx"

print(f"Loading: {WB_PATH}")
wb = load_workbook(WB_PATH)

lead = {
    "num":   19,
    "first": "Jason",
    "last":  "Clark",
    "full":  "Jason Clark",
    "title": "Director & Co-Founder",
    "company": "AndCo Realty Group",
    "email": "jason@andcorealty.nz",
    "email_status": "Direct founder email — verified on website",
    "website":  "https://www.andcorealty.nz/",
    "phone":    "+64 274 838 808",
    "city":     "New Zealand",
    "state":    "New Zealand",
    "industry": "Real Estate",
    "linkedin": "https://www.linkedin.com/company/andco-realty-group/about/",
    "ppt":   "AndCoRealty_Pitch_GenosAI.pptx",
    "score": 6.0,
    "automation": (
        "Agent/franchisee onboarding automation via WhatsApp (compliance orientation, marketing setup, "
        "milestone check-ins for new office operators); WhatsApp AI for buyer inquiry routing to the "
        "right regional office by area and property type; vendor nurture sequences for sellers who "
        "requested appraisal but haven't listed; REAA compliance and CPD tracking per agent across "
        "17 offices with automated renewal reminders; post-settlement review collection automation; "
        "agent performance reporting delivered monthly to each franchisee automatically"
    ),
    "notes": (
        "Fixed-fee commission model. 17 independent offices across NZ (Dunedin, Queenstown, Taranaki, "
        "Bay of Plenty, Whanganui, Wellington). Agent-owned franchise model — 'Your company, you call "
        "the shots.' Residential focus $500K-$5M+. REAA 2008 licensed. "
        "Co-founders: Jason Clark (Director), Kelly Clark (Business Development), Craig Lowe (Director). "
        "Personal LinkedIn not found — ice-breaker from website model/scale."
    ),
}

subject = "17 independent offices — onboarding, compliance, and lead flow for each without scaling your team"

body = """\
Hi Jason,

The AndCo model — fixed-fee, agent-owned, 17 offices across New Zealand — is a structurally different proposition to any traditional franchise. The operational challenge it creates is specific: every new office operator needs the same quality of onboarding, compliance guidance, and lead flow, but the team delivering that support doesn't scale linearly with the number of offices.

Genos AI builds the automation layer for real estate groups at this stage. For AndCo specifically:

Agent onboarding automation — new office operators go through a structured WhatsApp-based flow: compliance orientation, marketing setup, listing platform training, and milestone check-ins. You set the content once; the AI runs it for every new operator from day one.

Buyer inquiry routing — a WhatsApp AI that qualifies buyer inquiries from the AndCo website and routes them to the right regional office by area and property type. Each agent receives a pre-qualified lead with budget, property type, and timeline already captured.

Vendor nurture sequences — sellers who request an appraisal but don't proceed immediately receive an automated follow-up over 6-12 weeks. The agent stays top of mind without tracking every warm prospect manually.

REAA compliance and CPD tracking — license renewal dates and CPD requirements tracked per agent across all 17 offices, with automated reminders at 90, 30, and 7 days. No agent falls out of compliance because a reminder was missed.

Post-settlement review collection — automated Google and Trade Me review requests at settlement date. Across 17 offices, that's consistent social proof built without any admin overhead.

I'm Rohan Malik, CEO of Genos AI. We build AI automation for real estate groups. Happy to show you how it maps to AndCo's model in 20 minutes.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi Jason, following up on my last message. The fastest fix for a group at AndCo's stage — automating the onboarding flow for new office operators so your team isn't manually running the same process every time — is typically live in under 2 weeks. Happy to walk through it in 20 minutes. — Rohan | Genos AI\
"""

fu2 = """\
Jason, last one from me. A fixed-fee model with 17 independent offices scales on agent trust and operational consistency. AI automation ensures every new office operator gets the same quality of onboarding and support — regardless of how many join in a given month. Leaving this here if the timing isn't right. — Rohan | Genos AI\
"""


def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font      = cp.copy(src_cell.font)
        dst_cell.fill      = cp.copy(src_cell.fill)
        dst_cell.border    = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


# ── Organic Leads ──────────────────────────────────────────────────────────────
ws  = wb["Organic Leads"]
row = ws.max_row + 1
organic_vals = [
    lead["num"], lead["company"], lead["full"], lead["title"],
    lead["email"], lead["phone"], lead["website"], lead["linkedin"],
    lead["city"], lead["state"], lead["industry"],
    "Organic/Manual", "Pending Send", lead["score"],
    lead["ppt"], lead["automation"], lead["notes"],
]
for col, v in enumerate(organic_vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# ── Cold Emails ────────────────────────────────────────────────────────────────
ws2  = wb["Cold Emails"]
row2 = ws2.max_row + 1
email_vals = [
    lead["num"], lead["company"], lead["full"],
    lead["email"], subject, body, fu1, fu2,
]
for col, v in enumerate(email_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# ── Pipeline ───────────────────────────────────────────────────────────────────
ws3  = wb["Pipeline"]
row3 = ws3.max_row + 1
pipeline_vals = [
    lead["num"], lead["company"], lead["full"],
    lead["email"], lead["industry"], lead["score"],
    "Email Pending", "2026-05-17", "Send cold email — pitch AI automation",
    "",
]
for col, v in enumerate(pipeline_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

wb.save(WB_PATH)
print(f"OrganicLeads.xlsx updated — Jason Clark / AndCo Realty Group added to all 3 sheets (row {row}).")
