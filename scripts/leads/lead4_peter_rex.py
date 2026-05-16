from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy as cp
import os

ROOT = Path(__file__).parent.parent.parent
WB_DIR = ROOT / "output" / "workbooks"

import glob
files = glob.glob(str(WB_DIR / "GenosAI_RealEstate_Leads*.xlsx"))
src = next((f for f in files if "UPDATED" in f), str(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx"))
print(f"Loading: {src}")
wb = load_workbook(src)

lead = {
    "num": 93,
    "first": "Peter",
    "last": "Rex",
    "full": "Peter Rex",
    "title": "Founder & CEO",
    "company": "REX Companies",
    "email": "peter@rex.com",
    "email_status": "Verify",
    "website": "https://www.rex.com/real-estate",
    "phone": "",
    "city": "Austin",
    "state": "TX",
    "country": "US",
    "employees": "501-1000",
    "industry": "Real Estate - Multifamily / PropTech",
    "linkedin": "https://www.linkedin.com/in/peterex/",
    "ppt": "REX_Audit_GenosAI.pptx"
}

subject = "Disagree with everyone's approach to real estate tech - quick thought"

body = """\
Hi Peter,

Your X post: "I disagree with almost everyone's approach to tech in real estate." Curious what you'd say about ours.

I'm Rohan Malik, CEO of Genos AI. We work on the commercial layer - leasing funnels, investor acquisition, resident retention. Not a product company. Specifically using AI to drive revenue from existing real estate portfolios.

Ran a free audit of rex.com and attached it. One thing stood out.

Nine patents, 10 AI companies, GetDone using ML for maintenance, AI R&D going back to 2017. The tech side is real. What caught my eye in the audit is the gap between that depth and what the revenue-facing digital layer is doing.

Rex.com/real-estate reads like a brand deck, not an investor conversion funnel. No performance data, no fund structure, no team bios. OwnProp investors researching fractional real estate at 11pm find very little to act on. For 19,000 apartments across 19 metros, the leasing funnel at the individual property level may have the same gap - well-built product, underserved top of funnel.

A few things from the audit:
- Leasing lead sequences for 19,000 units - automated follow-up, tour confirmation, post-tour nudges recover the leads that go cold over weekends
- OwnProp investor nurture - fractional real estate has a longer consideration cycle. Education sequences and re-engagement flows for prospects who land but don't convert
- Resident retention - lease renewal sequences starting 90 days out, personalized by tenure and unit type, with measurable vacancy impact
- Investment pages rebuild - returns data, deal history, team bios - converting site visitors into fund inquiries, not just brand impressions

Audit's attached, no strings. Given what you're already building, there's probably a 15-minute conversation worth having.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi Peter, just checking the audit didn't get buried. The OwnProp investor nurture and leasing automation sections are probably the most relevant to what you're scaling. Happy to walk through it in 15 minutes. - Rohan\
"""

fu2 = """\
Hi Peter, last one from me. 19,000 apartments and 10 AI companies is a serious operation - leaving the audit here in case timing is better later. If it ever makes sense to compare notes on the revenue-layer AI stuff, you know where to find me. - Rohan\
"""

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = cp.copy(src_cell.font)
        dst_cell.fill = cp.copy(src_cell.fill)
        dst_cell.border = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

# ── 1. Leads ──
ws = wb["Leads"]
row = 94
vals = [lead["num"], lead["first"], lead["last"], lead["full"], lead["title"],
        lead["company"], lead["email"], lead["email_status"], lead["website"],
        lead["phone"], lead["city"], lead["state"], lead["country"],
        lead["employees"], lead["industry"], lead["linkedin"], lead["ppt"]]
for col, v in enumerate(vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# ── 2. Website Audits ──
ws2 = wb["Website Audits"]
audit_vals = [
    lead["num"], lead["company"], lead["website"],
    "Custom / Proprietary (Rex tech stack)", "Yes", "Google Analytics", "Unknown",
    "No", "None Detected", "Unknown", "Unknown", "Unknown", 5, "High",
    "Rex.com/real-estate reads as brand deck not investor conversion funnel; no performance data or fund structure; no team bios; OwnProp landing page lacks investor education funnel; no live chat for leasing or investor inquiries; no testimonials or deal case studies; individual apartment community pages unclear on conversion optimization",
    "Leasing lead automation for 19,000 unit portfolio (follow-up, tour scheduling, post-tour nudge sequences); OwnProp investor nurture (education + re-engagement sequences for fractional RE consideration cycle); resident retention sequences (90-day pre-renewal, personalized by tenure and unit); investment pages rebuild with returns data and fund detail; AI chat for investor and leasing inquiries 24/7; Peter thought leadership content auto-syndication pipeline"
]
row2 = 94
for col, v in enumerate(audit_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# ── 3. Cold Emails ──
ws3 = wb["Cold Emails"]
email_vals = [lead["num"], lead["first"], lead["company"], lead["email"], subject, body, fu1, fu2]
row3 = 94
for col, v in enumerate(email_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

# ── 4. Audit PPT Content ──
ws4 = wb["Audit PPT Content"]
ppt_vals = [
    lead["num"], lead["company"], lead["full"], lead["email"], lead["website"], 5, "High",
    "REX Companies - Website & AI Growth Audit | Prepared by Genos AI",
    "Founded by Peter Rex | Georgetown BA (Philosophy/Govt), Harvard Law JD, CPA | AI R&D since 2017 | 9 patents pending in real estate | 10 revenue-generating AI tech companies | 19,000 apartments in 19 metros across 6 states | $5B+ AUM | OwnProp (fractional tokenized RE) | GetDone (maintenance AI/ML) | 600+ teammates | Austin TX | WSJ op-ed 2020 on relocating from Seattle",
    "Deep AI/tech infrastructure (9 patents, AI R&D since 2017); vertically integrated operation across investment, asset management, construction; 10 revenue-generating tech products; premium modern brand design on rex.com; strong thought leadership presence (X, podcast, press); OwnProp is first-mover in tokenized fractional RE; Fortune top places to work (twice); proven capital discipline ($1.5B+ sold in 2019-2022)",
    "Rex.com/real-estate is brand deck not investor funnel - no performance data, fund structure, or team; OwnProp lacks investor education funnel (fractional RE has long consideration cycle); no live chat for leasing or investor inquiries; no testimonials or deal case studies; individual property marketing automation at scale unclear; no returns track record visible to accredited investors researching the firm",
    "1. Leasing lead automation for 19K units (follow-up, tour, post-tour, retention sequences) | 2. OwnProp investor acquisition funnel (education sequences, re-engagement, fractional RE nurture) | 3. Resident retention (90-day pre-renewal, personalized by tenure and unit type) | 4. Investment pages with returns data and deal case studies | 5. AI chat for leasing and investor inquiries 24/7 | 6. Peter thought leadership auto-syndication (X to email to press)",
    "Brand deck investment pages = accredited investors who researched but found nothing to act on. No OwnProp nurture sequence = fractional investors who visited, didn't convert, and never heard back. No leasing automation = leads going cold between Friday PM and Monday AM across 19,000 units. Nine patents and AI R&D since 2017 on the product side - but the revenue-facing digital layer is running manual.",
    "Rex.com/real-estate rebuild: returns data, fund structure, team bios, deal case studies - converts brand impressions into investor inquiries | OwnProp investor acquisition funnel: education sequences, fractional RE content, re-engagement flows | Leasing automation layer: lead capture, tour scheduling, post-tour follow-up, renewal sequences across all 19 metros | AI chat for leasing and investor inquiries 24/7 | Peter's thought leadership auto-syndicated from X to email to press",
    "Month 1: AI chat live for leasing and investor inquiries, OwnProp nurture sequences active, rex.com investment rebuild underway | Month 2: Leasing automation deployed across portfolio, investor pages live with returns data and team, first OwnProp nurture campaigns measured | Month 3: Resident retention sequences reducing vacancy at scale, OwnProp conversion rates up, leasing lead recovery from weekend gaps measurable",
    "Reply to this email or book a 15-min call | Rohan Malik, CEO | hello@genosai.tech | www.genosai.tech"
]
row4 = 94
for col, v in enumerate(ppt_vals, 1):
    cell = ws4.cell(row4, col)
    cell.value = v
    copy_style(ws4.cell(row4 - 1, col), cell)

# ── 5. Organic Leads ──
ws5 = wb["Organic Leads"]
thin = Side(style="thin", color="444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
row_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
row_font = Font(name="Calibri", size=10, color="1A1A2E")

organic_vals = [
    4, "Peter", "Rex", "Peter Rex", "Founder & CEO",
    "REX Companies", "peter@rex.com",
    "Verify", "https://www.rex.com/real-estate", "",
    "Austin", "TX", "US", "501-1000", "Real Estate - Multifamily / PropTech",
    "https://www.linkedin.com/in/peterex/", "Organic/Manual",
    "2026-05-06", "REX_Audit_GenosAI.pptx", "Email Sent - Pending Reply"
]
for col, v in enumerate(organic_vals, 1):
    cell = ws5.cell(5, col)
    cell.value = v
    cell.font = row_font
    cell.fill = row_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border

for col in range(1, ws5.max_column + 1):
    if ws5.cell(1, col).value == "Cold Email Body":
        cell = ws5.cell(5, col)
        cell.value = body
        cell.font = row_font
        cell.fill = row_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        break

ws5.row_dimensions[5].height = 180

wb.save(WB_DIR / "GenosAI_RealEstate_Leads_UPDATED.xlsx")
print("Excel updated - Peter Rex added to all 5 sheets.")
