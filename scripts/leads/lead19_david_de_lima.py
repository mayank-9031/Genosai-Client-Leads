from pathlib import Path
import copy as cp
from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent.parent
WB_PATH = ROOT / "output" / "workbooks" / "GenosAI_OrganicLeads.xlsx"

print(f"Loading: {WB_PATH}")
wb = load_workbook(WB_PATH)

lead = {
    "num":   17,
    "first": "David",
    "last":  "De Lima",
    "full":  "David De Lima",
    "title": "Managing Director",
    "company": "24/7 Security Services",
    "email": "info@24-7security.co.za",
    "email_status": "General contact — personal email not publicly listed",
    "website":  "https://24-7security.co.za/",
    "phone":    "+27 011 444 2237",
    "city":     "Johannesburg",
    "state":    "Gauteng, South Africa",
    "industry": "Security Services",
    "linkedin": "https://www.linkedin.com/in/david-de-lima-696ba99/",
    "ppt":   "24-7Security_Audit_GenosAI.pptx",
    "score": 3.5,
    "automation": (
        "WhatsApp AI for prospect inquiries (24/7 quote capture + site assessment scheduling); "
        "automated quote generation from security site assessment inputs; "
        "guard shift scheduling and attendance automation via WhatsApp; "
        "PSIRA certification renewal reminders per guard; "
        "automated monthly security activity and incident reports to clients; "
        "client onboarding automation (contract to deployment to handover); "
        "lead nurture sequences for prospects who inquired but didn't convert"
    ),
    "notes": (
        "SASA Gold, ISO 9001, Level 1 B-BBEE certified. 30+ years in market. "
        "Services: Armed Reaction, Physical Guarding, Remote Monitoring, Smart Surveillance, Drone Force. "
        "Also runs 24/7 Drone Force — autonomous drones for residential estate perimeter security. "
        "PSIRA registration: 70152. Offices: JHB, Tshwane, Cape Town. "
        "Ice-breaker: 24/7 Drone Force expansion into autonomous perimeter security."
    ),
}

subject = "Security quotes via contact form while residential estates move to instant WhatsApp response"

body = """\
Hi David,

Saw the 24/7 Drone Force expansion — autonomous drones reshaping perimeter security in residential estates is a different category from the standard guarding pitch. That kind of operational innovation is hard to replicate.

One gap that shows up when you look at 24-7security.co.za from a prospect's perspective: someone managing a residential estate or commercial property contacts you for a quote, fills in a form, and waits. Meanwhile they've sent the same inquiry to three other security companies. The first one to respond on WhatsApp — even if it's an AI asking qualifying questions — controls that conversation.

What Genos AI builds for security companies like 24/7:

WhatsApp AI for prospect inquiries — responds 24/7, asks qualifying questions (property type, size, current security setup, risk level), routes commercial to commercial and residential to residential, schedules a site assessment automatically.

Automated quote generation — site assessment inputs trigger a structured quote to your ops team. No manual data entry from the client side.

Guard attendance and shift automation — shift confirmations via WhatsApp, exception alerts when a guard doesn't check in. PSIRA certification renewal dates tracked per guard with automated reminders — no admin overhead.

Client communication automation — monthly activity reports and incident summaries delivered automatically to each client account. Makes your service value visible without your team writing reports manually.

I'm Rohan Malik, CEO of Genos AI. We build AI automation for security companies and managed service businesses. Happy to walk through the inquiry flow in a 20-minute call.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi David, following up on my last message. The fastest fix — a WhatsApp AI that captures and qualifies security quote inquiries 24/7 — is typically live in under 2 weeks. Happy to walk through it in 20 minutes. — Rohan | Genos AI\
"""

fu2 = """\
David, last one from me. For a company operating armed reaction across Johannesburg, Tshwane, and Cape Town simultaneously, the complexity of guard scheduling, client reporting, and prospect follow-up adds up fast. AI automation addresses all three without adding headcount. Leaving this here if the timing isn't right. — Rohan | Genos AI\
"""


def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font      = cp.copy(src_cell.font)
        dst_cell.fill      = cp.copy(src_cell.fill)
        dst_cell.border    = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


# ── Organic Leads ──────────────────────────────────────────────────────────────
ws  = wb["Organic Leads"]
row = ws.max_row + 1
organic_vals = [
    lead["num"], lead["company"], lead["full"], lead["title"],
    lead["email"], lead["phone"], lead["website"], lead["linkedin"],
    lead["city"], lead["state"], lead["industry"],
    "Organic/Manual", "Pending Send", lead["score"],
    lead["ppt"], lead["automation"], lead["notes"],
]
for col, v in enumerate(organic_vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# ── Cold Emails ────────────────────────────────────────────────────────────────
ws2  = wb["Cold Emails"]
row2 = ws2.max_row + 1
email_vals = [
    lead["num"], lead["company"], lead["full"],
    lead["email"], subject, body, fu1, fu2,
]
for col, v in enumerate(email_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# ── Pipeline ───────────────────────────────────────────────────────────────────
ws3  = wb["Pipeline"]
row3 = ws3.max_row + 1
pipeline_vals = [
    lead["num"], lead["company"], lead["full"],
    lead["email"], lead["industry"], lead["score"],
    "Email Pending", "2026-05-17", "Send cold email — pitch AI automation",
    "",
]
for col, v in enumerate(pipeline_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

wb.save(WB_PATH)
print(f"OrganicLeads.xlsx updated — David De Lima / 24/7 Security Services added to all 3 sheets (row {row}).")
