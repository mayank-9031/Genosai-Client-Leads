from pathlib import Path
import copy as cp
from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent.parent
WB_PATH = ROOT / "output" / "workbooks" / "GenosAI_OrganicLeads.xlsx"

print(f"Loading: {WB_PATH}")
wb = load_workbook(WB_PATH)

lead = {
    "num":   20,
    "first": "Eric",
    "last":  "Dahlstrom",
    "full":  "Eric Finnas Dahlstrom",
    "title": "CEO",
    "company": "JamesEdition",
    "email": "info@jamesedition.com",
    "email_status": "General contact — personal email not publicly listed",
    "website":  "https://www.jamesedition.com/",
    "phone":    "+31 20 241 7944",
    "city":     "Amsterdam",
    "state":    "North Holland, Netherlands",
    "industry": "Luxury Marketplace / PropTech",
    "linkedin": "https://www.linkedin.com/company/jamesedition/",
    "ppt":   "JamesEdition_Pitch_GenosAI.pptx",
    "score": 8.0,
    "automation": (
        "Seller onboarding automation (structured WhatsApp/email sequences for new sellers — "
        "listing optimization, platform guidance, milestone check-ins); "
        "churn prevention AI (flag sellers with no leads in 14+ days and trigger re-engagement "
        "before cancellation); upsell trigger sequences (basic to premium based on lead volume); "
        "buyer inquiry pre-qualification for $500K+ listings; "
        "tier-1 support automation (account, billing, listing queries handled by AI); "
        "seller performance report delivery (automated monthly dashboards to each seller)"
    ),
    "notes": (
        "Global luxury marketplace. 770K+ listings, 40K+ sellers, 21K+ trusted sellers, "
        "120+ countries, 35 employees. Subscription model: $299-$1,199/month. "
        "Min listing price $500K USD. Founder: Noam Perski. CEO: Eric Finnas Dahlstrom "
        "(prev. Head of General Classifieds at Avito). CTO: Alexander Lomakin. "
        "Categories: real estate, yachts, jets, cars, watches. "
        "Tech stack: Pipedrive CRM, Hotjar, Mixpanel, Tableau. "
        "Ice-breaker: 35-person team managing 40K+ sellers across 120 countries — "
        "operational ratio only works with automation. Personal LinkedIn not found."
    ),
}

subject = "35 employees, 40,000 sellers, 120 countries — the operational ratio that only works with automation"

body = """\
Hi Eric,

JamesEdition runs 40,000+ sellers across 120 countries with a 35-person team. That ratio only works because the platform handles discovery — but the seller-side operational layer (onboarding, engagement, churn prevention, upsell) scales differently from the marketplace itself.

The gap at JamesEdition's current stage: a seller paying $299/month who doesn't see inquiries in week one is likely to cancel. A seller on a basic plan who's getting strong lead volume has no automated prompt to upgrade to premium. Neither problem requires a human — both require a system.

What Genos AI builds for marketplace businesses at JamesEdition's scale:

Seller onboarding automation — new sellers receive a structured sequence: listing optimisation guidance, platform tips, performance benchmarks, and milestone check-ins. With 40,000 sellers and 35 employees, that process doesn't happen manually — and most sellers who leave in month one leave because nobody helped them get started.

Churn prevention AI — sellers whose lead volume drops below a threshold, or who haven't logged in for 14 days, are flagged and receive an automated re-engagement sequence before they cancel. At $299–$1,199/month per seller, each prevented cancellation is material to revenue.

Upsell trigger sequences — sellers on basic plans who've received strong lead volume receive an automated upgrade prompt with specific ROI context from their own data. No sales rep required.

Buyer pre-qualification — when a buyer inquires on a $5M+ listing, an AI runs an initial qualification (budget confirmed, timeline, finance or cash) before connecting to the seller. Reduces seller time spent on unqualified inquiries — a common complaint on luxury platforms.

Tier-1 support automation — listing issues, account questions, billing queries handled by AI before escalating to your team. With 21,000+ trusted sellers, the inbound support volume doesn't scale to 35 people without automation.

I'm Rohan Malik, CEO of Genos AI. We build AI automation for marketplace businesses. Happy to show you how it maps to JamesEdition's operational model in 15 minutes.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi Eric, following up on my last message. The highest-impact fix — churn prevention sequences for sellers who go quiet in the first 30 days — is typically live in under 2 weeks and directly protects subscription revenue. Happy to walk through it in 15 minutes. — Rohan | Genos AI\
"""

fu2 = """\
Eric, last one from me. A 35-person team running a $299–$1,199/month subscription marketplace at 40,000+ sellers across 120 countries only stays profitable if the seller success layer is automated. That's exactly what we build. Leaving this here if the timing isn't right. — Rohan | Genos AI\
"""


def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font      = cp.copy(src_cell.font)
        dst_cell.fill      = cp.copy(src_cell.fill)
        dst_cell.border    = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


# ── Organic Leads ──────────────────────────────────────────────────────────────
ws  = wb["Organic Leads"]
row = ws.max_row + 1
organic_vals = [
    lead["num"], lead["company"], lead["full"], lead["title"],
    lead["email"], lead["phone"], lead["website"], lead["linkedin"],
    lead["city"], lead["state"], lead["industry"],
    "Organic/Manual", "Pending Send", lead["score"],
    lead["ppt"], lead["automation"], lead["notes"],
]
for col, v in enumerate(organic_vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# ── Cold Emails ────────────────────────────────────────────────────────────────
ws2  = wb["Cold Emails"]
row2 = ws2.max_row + 1
email_vals = [
    lead["num"], lead["company"], lead["full"],
    lead["email"], subject, body, fu1, fu2,
]
for col, v in enumerate(email_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# ── Pipeline ───────────────────────────────────────────────────────────────────
ws3  = wb["Pipeline"]
row3 = ws3.max_row + 1
pipeline_vals = [
    lead["num"], lead["company"], lead["full"],
    lead["email"], lead["industry"], lead["score"],
    "Email Pending", "2026-05-17", "Send cold email — pitch AI automation",
    "",
]
for col, v in enumerate(pipeline_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

wb.save(WB_PATH)
print(f"OrganicLeads.xlsx updated — Eric Dahlstrom / JamesEdition added to all 3 sheets (row {row}).")
