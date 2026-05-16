from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy as cp
import os

ROOT = Path(__file__).parent.parent.parent
WB_DIR = ROOT / "output" / "workbooks"

# Load whichever file is available
import glob
files = glob.glob(str(WB_DIR / "GenosAI_RealEstate_Leads*.xlsx"))
# prefer _UPDATED, else original
src = next((f for f in files if "UPDATED" in f), str(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx"))
print(f"Loading: {src}")
wb = load_workbook(src)

# ── Lead data ──
lead = {
    "num": 91,
    "first": "Rich",
    "last": "DiGiovanna",
    "full": "Rich DiGiovanna",
    "title": "Associate Broker / Team Leader",
    "company": "DiGiovanna Realty Group (iSellVirginia)",
    "email": "richd@isellvirginia.com",
    "email_status": "Valid",
    "website": "https://isellvirginia.com/",
    "phone": "(571) 655-3753",
    "city": "Burke",
    "state": "VA",
    "country": "US",
    "employees": "2-10",
    "industry": "Real Estate - Residential",
    "linkedin": "https://www.linkedin.com/in/richdigiovanna/",
    "ppt": "iSellVirginia_Audit_GenosAI.pptx"
}

subject = "39 years, 2,000 families - free iSellVirginia audit attached"

body = """\
Hi Rich,

39 years in Northern Virginia, 2,000+ families helped, and a YouTube channel you're actually posting on. The brand is real.

I'm Rohan Malik, CEO of Genos AI. We help real estate teams turn that kind of track record into more inbound through AI tools and a website that works between 5pm and 9am.

Ran a free audit of iSellVirginia.com and attached it. A few things stood out.

The testimonials section is empty, your YouTube videos aren't surfacing on the homepage, and there's no live chat. When a seller quietly compares agents on a Sunday night before calling Monday morning, those gaps decide who gets the call. We'd rebuild the site to actually look like what 39 years and 2,000 families means - clean, modern, YouTube content embedded, real community pages that rank.

A few AI things worth a look for a team your size:
- Lead follow-up sequences so no inquiry goes cold over a weekend
- Automated market reports to your farm areas, keeping you top of mind before someone lists
- AI chat capturing buyers and sellers whenever you're not in the office
- Post-closing review automation to fill that empty testimonials section
- Property management automation for Superior Real Estate Services

Audit's attached, no strings. Happy to walk through it in 15 minutes.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi Rich, just checking the audit didn't get buried. The lead follow-up and review automation sections are probably the most relevant for a team your size. Happy to walk through it in 15 minutes. - Rohan\
"""

fu2 = """\
Hi Rich, last one from me. Dropping the audit here in case timing is better later. 39 years and 2,000 families is a serious body of work - the digital side should match. If it ever makes sense to talk, you know where to find me. - Rohan\
"""

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = cp.copy(src_cell.font)
        dst_cell.fill = cp.copy(src_cell.fill)
        dst_cell.border = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

# ── 1. Leads ──
ws = wb["Leads"]
row = 92
vals = [lead["num"], lead["first"], lead["last"], lead["full"], lead["title"],
        lead["company"], lead["email"], lead["email_status"], lead["website"],
        lead["phone"], lead["city"], lead["state"], lead["country"],
        lead["employees"], lead["industry"], lead["linkedin"], lead["ppt"]]
for col, v in enumerate(vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# ── 2. Website Audits ──
ws2 = wb["Website Audits"]
audit_vals = [
    lead["num"], lead["company"], lead["website"],
    "WordPress / IDXBroker", "Yes", "Google Analytics", "LeadConnectorHQ",
    "No", "None Detected", "CloudFlare", "Unknown", "Unknown", 5, "High",
    "Empty testimonials section; YouTube content not embedded on homepage; no live chat; template-heavy design lacks personality; no sticky header; limited blog; CTAs lack urgency; duplicate nav in markup",
    "AI lead follow-up sequences; automated seller market reports to farm areas; AI chat 24/7; post-closing review automation; property management automation (Superior RE Services); listing description generator; social media content automation"
]
row2 = 92
for col, v in enumerate(audit_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# ── 3. Cold Emails ──
ws3 = wb["Cold Emails"]
email_vals = [lead["num"], lead["first"], lead["company"], lead["email"], subject, body, fu1, fu2]
row3 = 92
for col, v in enumerate(email_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

# ── 4. Audit PPT Content ──
ws4 = wb["Audit PPT Content"]
ppt_vals = [
    lead["num"], lead["company"], lead["full"], lead["email"], lead["website"], 5, "High",
    "iSellVirginia / DiGiovanna Realty Group - Website & AI Growth Audit | Prepared by Genos AI",
    "39 years in Northern Virginia | 2,000+ families moved | $500M+ career sales | 14 years U.S. Coast Guard | CRS, GRI, ABR certified | Team: 5 members | RE/MAX Allegiance | YouTube channel active",
    "Strong brand credibility and long tenure; active YouTube presence; IDX platform live; good local SEO structure; CloudFlare CDN; mortgage partner (Intercoastal); property management arm (Superior RE Services)",
    "Testimonials section empty; YouTube not embedded on homepage; no live chat or AI bot; template-heavy design lacks personality; duplicate nav in markup; no sticky header; limited blog content; CTAs lack urgency; no remarketing setup",
    "1. AI lead follow-up (no lead goes cold over a weekend) | 2. Automated seller market reports to farm areas | 3. AI chat 24/7 for buyer/seller capture | 4. Post-closing review automation | 5. Property management automation for Superior RE Services | 6. Listing description generator | 7. Social content automation (YouTube, Instagram, Facebook)",
    "No follow-up automation = leads going cold every Friday evening. Empty testimonials = social proof gap at the exact moment a seller is deciding who to call. No chat = losing after-hours inquiries to Zillow and Redfin who have bots running 24/7.",
    "Modern site rebuild: YouTube content embedded, community pages that rank, clean premium design reflecting 39 years | AI chat widget trained on Rich's services and farm areas | Lead nurture sequences (30/60/90 day) | Automated CMA reports to prospect list | Review automation post-closing | Property management workflow automation for Superior RE Services",
    "Month 1: new site live, AI chat capturing leads 24/7, first automated market reports sent | Month 2: lead sequences active, testimonials section filling, YouTube embedding driving engagement | Month 3: Superior RE Services workflows automated, 2x review velocity, SEO content ranking for Burke/Woodbridge keywords",
    "Reply to this email or book a 15-min call | Rohan Malik, CEO | hello@genosai.tech | www.genosai.tech"
]
row4 = 92
for col, v in enumerate(ppt_vals, 1):
    cell = ws4.cell(row4, col)
    cell.value = v
    copy_style(ws4.cell(row4 - 1, col), cell)

# ── 5. Organic Leads ──
ws5 = wb["Organic Leads"]
thin = Side(style="thin", color="444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
row_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
row_font = Font(name="Calibri", size=10, color="1A1A2E")

organic_vals = [
    2, "Rich", "DiGiovanna", "Rich DiGiovanna", "Associate Broker / Team Leader",
    "DiGiovanna Realty Group (iSellVirginia)", "richd@isellvirginia.com",
    "Valid", "https://isellvirginia.com/", "(571) 655-3753",
    "Burke", "VA", "US", "2-10", "Real Estate - Residential",
    "https://www.linkedin.com/in/richdigiovanna/", "Organic/Manual",
    "2026-05-06", "iSellVirginia_Audit_GenosAI.pptx", "Email Sent - Pending Reply"
]
for col, v in enumerate(organic_vals, 1):
    cell = ws5.cell(3, col)
    cell.value = v
    cell.font = row_font
    cell.fill = row_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border

# Email body column
for col in range(1, ws5.max_column + 1):
    if ws5.cell(1, col).value == "Cold Email Body":
        cell = ws5.cell(3, col)
        cell.value = body
        cell.font = row_font
        cell.fill = row_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        break

ws5.row_dimensions[3].height = 180

wb.save(WB_DIR / "GenosAI_RealEstate_Leads_UPDATED.xlsx")
print("Excel updated - Rich DiGiovanna added to all 5 sheets.")
