from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy as cp
import os

ROOT = Path(__file__).parent.parent.parent
WB_DIR = ROOT / "output" / "workbooks"

import glob
files = glob.glob(str(WB_DIR / "GenosAI_RealEstate_Leads*.xlsx"))
src = next((f for f in files if "UPDATED" in f), str(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx"))
print(f"Loading: {src}")
wb = load_workbook(src)

lead = {
    "num": 95,
    "first": "Ofer",
    "last": "Alon",
    "full": "Ofer Alon",
    "title": "CEO & Founder",
    "company": "Quality Acceptance / Nexa Optimum Solutions",
    "email": "oalon@qualityacceptance.com",
    "email_status": "Verify",
    "website": "https://www.qualityacceptance.com/",
    "phone": "818.503.1322",
    "city": "Van Nuys",
    "state": "CA",
    "country": "US",
    "employees": "51-200",
    "industry": "Auto Finance / Subprime Lending",
    "linkedin": "https://www.linkedin.com/in/ofer-alon-34776123/",
    "ppt": "QualityAcceptance_Audit_GenosAI.pptx"
}

subject = "You wrote 'technology is not failing; implementation is failing' - agree completely"

body = """\
Hi Ofer,

Your Auto Remarketing commentary said something most lenders won't say out loud: the problem isn't the technology, it's the gap between deploying tools and actually transforming operations. That's the gap we work on.

I'm Rohan Malik, CEO of Genos AI. We help financial services companies build the automation layer that connects existing systems to actual revenue outcomes, not just activity metrics.

Ran a free audit of qualityacceptance.com and attached it. A few things stood out.

Quality Acceptance serves two audiences with different needs. Dealers need a clear enrollment funnel and ongoing engagement to keep submitting deals. Consumers need self-service that doesn't depend on business hours. Right now, both experiences start and end with a phone number.

The same implementation gap you described for debt collection applies here. Dealer reactivation sequences, consumer payment reminder flows built around actual behavioral triggers, insurance lapse alerts before they become compliance issues. These aren't technology problems. They're implementation problems. You named the gap. We close it.

A few things from the audit:
- Dealer acquisition funnel: no conversion path from site visit to enrolled dealer, just a contact number
- Consumer self-service: AI chat handles 60-70% of inbound call volume (balances, payment options, insurance questions) without touching your servicing team
- Payment reminder sequences: behavioral-trigger flows reduce early-stage delinquency more than static reminder schedules
- Insurance management automation: automated alerts when coverage lapses, before it becomes a loss exposure
- Dealer reactivation: dealers who go quiet aren't always done - automated sequences at 30/60/90 days recover deals that drop off without requiring a call

Audit's attached, no strings. Given what you're building with Nexa as well, there's probably a useful 15 minutes here.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi Ofer, just checking the audit didn't get buried. The dealer acquisition funnel and consumer self-service sections are probably the most relevant to what you're running at both Quality Acceptance and Nexa. Happy to walk through it in 15 minutes. - Rohan\
"""

fu2 = """\
Hi Ofer, last one from me. Your point about implementation being the failure mode, not the technology, is exactly why I reached out. Leaving the audit here in case timing is better down the road. - Rohan\
"""

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = cp.copy(src_cell.font)
        dst_cell.fill = cp.copy(src_cell.fill)
        dst_cell.border = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

# ── 1. Leads ──
ws = wb["Leads"]
row = 96
vals = [lead["num"], lead["first"], lead["last"], lead["full"], lead["title"],
        lead["company"], lead["email"], lead["email_status"], lead["website"],
        lead["phone"], lead["city"], lead["state"], lead["country"],
        lead["employees"], lead["industry"], lead["linkedin"], lead["ppt"]]
for col, v in enumerate(vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# ── 2. Website Audits ──
ws2 = wb["Website Audits"]
audit_vals = [
    lead["num"], lead["company"], lead["website"],
    "Custom / Legacy Stack", "Yes", "Unknown", "Unknown",
    "Yes", "None Detected", "Unknown", "Unknown", "Unknown", 5, "High",
    "No dealer acquisition funnel on site - just a phone number, no online enrollment path; no AI chat or consumer self-service for payment/balance/insurance questions (heavy reliance on 1-800 number); no behavioral-trigger payment reminder sequences visible; insurance lapse management appears manual; no dealer reactivation automation for dormant dealer relationships; dated site design does not reflect 25+ years and $1B+ in funded transactions; no dealer or consumer testimonials; no case studies",
    "Dealer acquisition funnel with online enrollment CTA and automated onboarding sequences; AI chat for consumer self-service (balances, payment options, insurance questions) - reduces 60-70% of inbound call volume; behavioral-trigger payment reminder sequences replacing static schedules; insurance lapse alert automation before compliance exposure; dealer reactivation sequences at 30/60/90 days for dormant dealers; site rebuild reflecting institutional credibility ($1B+ funded, A+ BBB, 25+ years)"
]
row2 = 96
for col, v in enumerate(audit_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# ── 3. Cold Emails ──
ws3 = wb["Cold Emails"]
email_vals = [lead["num"], lead["first"], lead["company"], lead["email"], subject, body, fu1, fu2]
row3 = 96
for col, v in enumerate(email_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

# ── 4. Audit PPT Content ──
ws4 = wb["Audit PPT Content"]
ppt_vals = [
    lead["num"], lead["company"], lead["full"], lead["email"], lead["website"], 5, "High",
    "Quality Acceptance / Nexa Optimum Solutions - Website & AI Growth Audit | Prepared by Genos AI",
    "Ofer Alon: CEO Quality Acceptance since 2002+ | Founded Nexa Optimum Solutions 2025 | Harvard Business School | University of Bradford | Tel Aviv University CPA | Former Ernst & Young manager | Former high-tech CEO at Buildcom.com | AFSA Board of Directors | 1st Franklin Financial Board (April 2026) | Quality Acceptance: founded 1998, Van Nuys CA, $1B+ funded transactions, 180K+ consumers served, CA/NV/TX/AZ, A+ BBB, IDB Bank-backed credit facility | Auto Remarketing commentary author: 'technology is not failing; implementation is failing'",
    "25+ years in subprime auto finance with institutional backing (IDB Bank); $1B+ funded transactions across 4 states; Ofer's rare background - CPA + Ernst & Young + Harvard + tech founder (Buildcom.com); AFSA board member and 1st Franklin Financial board (April 2026) - deep industry credibility; A+ BBB rating; Nexa Optimum Solutions shows proactive thinking about where auto finance is heading; mobile app for consumers already deployed",
    "No dealer acquisition funnel on site - just a phone number, no online enrollment path; no AI chat or consumer self-service (heavy 1-800 reliance for balances/payments/insurance); no behavioral-trigger payment reminder sequences; insurance lapse management appears manual; no dealer reactivation automation for dormant relationships; dated site design understates 25+ years of institutional credibility; no dealer or consumer testimonials visible",
    "1. Dealer acquisition funnel - online enrollment path with automated onboarding sequences | 2. Consumer AI chat - 24/7 self-service for payment, balance, insurance questions (60-70% call deflection) | 3. Behavioral-trigger payment reminder sequences replacing static schedules | 4. Insurance lapse automation - alerts and follow-up before compliance exposure | 5. Dealer reactivation - 30/60/90 day sequences for dormant dealers | 6. Site rebuild reflecting institutional scale ($1B+ funded, A+ BBB, 25+ years)",
    "Ofer wrote publicly that technology is not failing - implementation is. Quality Acceptance has the volume and track record. The gap is the automation layer: no self-service for the consumer calling about a balance, no sequence for the dealer who stopped submitting deals, no behavioral trigger for the insurance lapse that becomes a loss before anyone notices.",
    "Dealer acquisition page with online enrollment form and automated onboarding sequence | Consumer AI chat handling balance/payment/insurance questions 24/7 without touching servicing staff | Behavioral-trigger payment reminder flows segmented by days-since-last-payment and communication history | Insurance lapse alert automation triggered immediately on lapse notification | Dealer reactivation sequences at 30/60/90 day dormancy marks | Site rebuild with institutional credibility signals: $1B+ funded, case studies, dealer and consumer testimonials",
    "Month 1: AI chat live on qualityacceptance.com for consumer self-service, dealer acquisition funnel with online enrollment active, insurance lapse alert automation deployed | Month 2: Behavioral-trigger payment reminder sequences replacing static outreach, dealer reactivation sequences running, site rebuild underway with credibility signals | Month 3: Consumer call volume measurably reduced, dealer reactivation recovering dormant relationships, site live with testimonials and case studies, full implementation gap closed",
    "Reply to this email or book a 15-min call | Rohan Malik, CEO | hello@genosai.tech | www.genosai.tech"
]
row4 = 96
for col, v in enumerate(ppt_vals, 1):
    cell = ws4.cell(row4, col)
    cell.value = v
    copy_style(ws4.cell(row4 - 1, col), cell)

# ── 5. Organic Leads ──
ws5 = wb["Organic Leads"]
thin = Side(style="thin", color="444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
row_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
row_font = Font(name="Calibri", size=10, color="1A1A2E")

organic_vals = [
    6, "Ofer", "Alon", "Ofer Alon", "CEO & Founder",
    "Quality Acceptance / Nexa Optimum Solutions", "oalon@qualityacceptance.com",
    "Verify", "https://www.qualityacceptance.com/", "818.503.1322",
    "Van Nuys", "CA", "US", "51-200", "Auto Finance / Subprime Lending",
    "https://www.linkedin.com/in/ofer-alon-34776123/", "Organic/Manual",
    "2026-05-06", "QualityAcceptance_Audit_GenosAI.pptx", "Email Sent - Pending Reply"
]
for col, v in enumerate(organic_vals, 1):
    cell = ws5.cell(7, col)
    cell.value = v
    cell.font = row_font
    cell.fill = row_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border

for col in range(1, ws5.max_column + 1):
    if ws5.cell(1, col).value == "Cold Email Body":
        cell = ws5.cell(7, col)
        cell.value = body
        cell.font = row_font
        cell.fill = row_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        break

ws5.row_dimensions[7].height = 180

wb.save(WB_DIR / "GenosAI_RealEstate_Leads_UPDATED.xlsx")
print("Excel updated - Ofer Alon added to all 5 sheets.")
