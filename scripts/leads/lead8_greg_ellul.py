from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import copy as cp
import os

ROOT = Path(__file__).parent.parent.parent
WB_DIR = ROOT / "output" / "workbooks"

import glob
files = glob.glob(str(WB_DIR / "GenosAI_RealEstate_Leads*.xlsx"))
src = next((f for f in files if "UPDATED" in f), str(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx"))
print(f"Loading: {src}")
wb = load_workbook(src)

lead = {
    "num": 97,
    "first": "Greg",
    "last": "Ellul",
    "full": "Greg Ellul",
    "title": "Director, MMJ North / Licensee",
    "company": "MMJ North Real Estate (Corrimal)",
    "email": "greg.ellul@mmj.com.au",
    "email_status": "Valid",
    "website": "https://www.mmj.com.au/about/offices/north-corrimal/",
    "phone": "0412 421 214",
    "city": "Corrimal",
    "state": "NSW",
    "country": "AU",
    "employees": "11-50",
    "industry": "Real Estate - Residential / Commercial / PM",
    "linkedin": "https://www.linkedin.com/company/linkedin-mmj-north/",
    "ppt": "MMJNorth_Audit_GenosAI.pptx"
}

subject = "30 years on the Illawarra coast, 6 REINSW finalists, and a website that doesn't ask for the lead"

body = """\
Hi Greg,

64 years of MMJ in Australian real estate, six REINSW Awards finalists in 2025, Tiana McGrath taking the John Hill Award for property management, and your own 30+ years across the Illawarra coast from Towradgi to Stanwell Park. The track record speaks for itself.

I'm Rohan Malik, CEO of Genos AI. We help established agencies turn that kind of credential into more inbound through AI tools and a website that does something at 11pm when a vendor is comparing agents on the internet.

Ran a free audit of mmj.com.au with a focus on the North Corrimal office page and attached it. A few things stood out.

A vendor thinking about listing in Woonona or Bulli does the same thing every other vendor does first. Searches at 9pm, lands on three or four agency sites, decides who to call in the morning. The agencies with instant valuation tools and clear sold data capture that lead. The ones with a contact form and a phone number don't. MMJ has the track record. The site doesn't ask for the lead.

Same goes for the Corrimal office page. 30+ agents, paginated 6 deep, no agent matching tool, no recent sales by suburb, no neighborhood data. A buyer trying to figure out which agent specialises in Tarrawanna versus Corrimal is doing the matching themselves, manually, while comparing you to the next agency.

A few things worth a look:
- Instant property valuation tool on homepage and office page. Highest-converting lead magnet in residential real estate, with automated follow-up routed to the right agent based on suburb
- AI chat for vendor, buyer, and tenant questions after hours. Handles routine questions and books appraisals or inspections without anyone in the office picking up
- Vendor nurture sequences with weekly suburb-specific market reports keeping MMJ top of mind in the 6 to 12 month window before someone lists
- Office-level pages with recent sales by suburb, agent matching by property type, neighborhood data. Lead generation pages, not just team rosters
- Tenant automation for property management: maintenance requests, lease renewal sequences starting 90 days out, rent reminders triggered by behaviour
- Past-client re-engagement so the buyer from 2018 becomes the appraisal request in 2026, because MMJ stayed in their inbox

On the website side, we'd rebuild to reflect what 64 years actually means. National awards surfaced on every primary page, instant valuation as a homepage CTA, office pages that act as suburb authorities and not just team rosters.

Audit's attached, no strings. Happy to walk through it in 15 minutes.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi Greg, just making sure the audit didn't get buried. The instant valuation tool and the suburb-level office pages are probably the most relevant for North Corrimal's vendor pipeline this year. Happy to walk through it in 15 minutes. - Rohan\
"""

fu2 = """\
Hi Greg, last one from me. 30 years in this market is rare. Most of the people thinking about selling in Woonona, Bulli, or Stanwell Park in the next year haven't heard from MMJ yet. The vendor nurture and past-client re-engagement automation we'd build is exactly what closes that window. Leaving the audit here in case timing improves. - Rohan\
"""

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = cp.copy(src_cell.font)
        dst_cell.fill = cp.copy(src_cell.fill)
        dst_cell.border = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

# Leads
ws = wb["Leads"]
row = 98
vals = [lead["num"], lead["first"], lead["last"], lead["full"], lead["title"],
        lead["company"], lead["email"], lead["email_status"], lead["website"],
        lead["phone"], lead["city"], lead["state"], lead["country"],
        lead["employees"], lead["industry"], lead["linkedin"], lead["ppt"]]
for col, v in enumerate(vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# Website Audits
ws2 = wb["Website Audits"]
audit_vals = [
    lead["num"], lead["company"], lead["website"],
    "WordPress / Custom (mmj.com.au)", "Yes", "Unknown", "Unknown",
    "No", "None Detected", "Unknown", "Unknown", "Yes (newsletter)", 5, "High",
    "No instant property valuation tool on homepage or office pages despite this being the #1 lead magnet in residential RE; no AI chat for vendor/buyer/tenant after-hours questions; office pages are paginated agent rosters (6 pages, 30+ agents) with no agent matching tool, no recent sold data by suburb, no neighborhood market data; no online appraisal request form; no online rental application; no tenant portal for maintenance/rent/lease renewal; no video tours or VR tours; only 4 Google testimonials surfaced on homepage despite 64-year track record; REINSW awards and 6 finalist nominations not surfaced on primary pages; only digital lead capture is generic newsletter signup + property alerts",
    "Instant property valuation tool with suburb-routed follow-up sequences; AI chat trained on MMJ's appraisal process, market data, and PM workflows for 24/7 vendor/buyer/tenant self-service; vendor nurture sequences with weekly suburb-specific market reports targeting 6-12 month listing window; office-level suburb authority pages with recent sales, agent-suburb matching, neighborhood data; tenant portal automation (maintenance, lease renewals starting 90 days out, behavioural rent reminders); past-client re-engagement sequences (anniversary of purchase = appraisal opportunity); buyer alerts segmented by suburb and property type; awards/credibility wall on homepage and office pages"
]
row2 = 98
for col, v in enumerate(audit_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# Cold Emails
ws3 = wb["Cold Emails"]
email_vals = [lead["num"], lead["first"], lead["company"], lead["email"], subject, body, fu1, fu2]
row3 = 98
for col, v in enumerate(email_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

# Audit PPT Content
ws4 = wb["Audit PPT Content"]
ppt_vals = [
    lead["num"], lead["company"], lead["full"], lead["email"], lead["website"], 5, "High",
    "MMJ North Real Estate (Corrimal) - Website & AI Growth Audit | Prepared by Genos AI",
    "Greg Ellul: North Director / Licensee at MMJ North Real Estate, Corrimal NSW | 30+ years in real estate (since 1992) | Construction industry background informing property valuation and investment evaluation | Specialises in commercial and residential RE, property management, business sales | Coverage: Illawarra coastline from Towradgi to Stanwell Park (Corrimal, Woonona, Tarrawanna, Bulli, Stanwell Park) | Office: 251-253 Princes Highway, Corrimal NSW 2518 | 30+ agents across 6 paginated team pages | Parent company MMJ Real Estate founded 1960 (64+ years) | 11-50 employees at North office | 6 REINSW Awards finalists in 2025 | Tiana McGrath won John Hill Award (residential property management) | Recent activity: 11 listings for sale, 15 recently sold",
    "64-year MMJ track record going back to 1960 - rare longevity in Australian real estate; 6 REINSW Awards finalists in 2025 + John Hill Award winner shows team-wide quality; Greg's 30+ years and construction background = unusual combo of negotiation and property valuation depth; deep Illawarra coast specialisation (Towradgi to Stanwell Park) - the kind of geographic authority national portals can't fake; multi-service: residential sales + commercial + property management + business sales + town planning + strata; active community presence (food drives, client appreciation events); annual residential retreats for staff",
    "No instant property valuation tool despite being the highest-converting lead magnet in residential RE; no AI chat for after-hours vendor/buyer/tenant questions; office pages are paginated agent rosters with no agent matching, recent sold data, or neighborhood data; no online appraisal request form; no tenant portal for property management; no online rental application; only 4 Google testimonials on homepage despite 64-year history; REINSW awards not surfaced on primary pages; no past-client re-engagement automation; the only digital lead capture is generic newsletter + property alerts",
    "1. Instant property valuation tool on homepage + office page with suburb-routed follow-up | 2. AI chat for vendor/buyer/tenant 24/7 self-service - books appraisals and inspections without staff | 3. Vendor nurture sequences with weekly suburb-specific market reports (6-12 month listing window) | 4. Office-level suburb authority pages with recent sales, agent matching, neighborhood data | 5. Tenant portal automation: maintenance, lease renewals 90 days out, behavioural rent reminders | 6. Past-client re-engagement sequences (purchase anniversary = appraisal opportunity) | 7. Awards and credibility wall surfaced on homepage and office pages",
    "A vendor in Woonona or Bulli searches at 9pm, lands on three or four agency sites, and decides who to call in the morning. The agencies with instant valuation and clear sold data capture that lead. MMJ's contact form doesn't. Buyer trying to figure out which agent specialises in Tarrawanna versus Corrimal is doing manual matching while comparing MMJ to the next agency. 64 years of credibility, 6 REINSW finalists, and 30+ years from Greg himself - none of it surfaced where it would close the trust gap before the first call.",
    "Instant valuation tool on homepage and office page with suburb-routed automated follow-up | AI chat trained on MMJ's appraisal process, market data, PM workflows - 24/7 routing | Vendor nurture: weekly suburb-specific market reports automated by past behaviour | Office pages rebuilt as suburb authorities: recent sales, agent-suburb matching, neighborhood data | Tenant portal: online rental applications, maintenance requests, lease renewal sequences 90 days out, behavioural rent reminders | Past-client re-engagement: 1/3/5/7-year anniversary sequences with appraisal CTA | Awards wall on homepage + office pages | Site rebuild reflecting 64-year credibility and national award footprint",
    "Month 1: Instant valuation tool live on mmj.com.au + North Corrimal office page, AI chat deployed for vendor/buyer/tenant questions, office page rebuilt with recent sold data and agent matching, awards/credibility wall added | Month 2: Vendor nurture sequences running with suburb-specific weekly market reports, tenant portal live (maintenance, applications, renewals), past-client re-engagement sequences active | Month 3: Buyer alerts segmented by suburb and property type running, office pages acting as suburb authorities (Corrimal, Woonona, Tarrawanna, Bulli, Stanwell Park), measurable inbound appraisal requests from valuation tool",
    "Reply to this email or book a 15-min call | Rohan Malik, CEO | hello@genosai.tech | www.genosai.tech"
]
row4 = 98
for col, v in enumerate(ppt_vals, 1):
    cell = ws4.cell(row4, col)
    cell.value = v
    copy_style(ws4.cell(row4 - 1, col), cell)

# Organic Leads
ws5 = wb["Organic Leads"]
thin = Side(style="thin", color="444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
row_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
row_font = Font(name="Calibri", size=10, color="1A1A2E")

organic_vals = [
    8, "Greg", "Ellul", "Greg Ellul", "Director, MMJ North / Licensee",
    "MMJ North Real Estate (Corrimal)", "greg.ellul@mmj.com.au",
    "Valid", "https://www.mmj.com.au/about/offices/north-corrimal/", "0412 421 214",
    "Corrimal", "NSW", "AU", "11-50", "Real Estate - Residential / Commercial / PM",
    "https://www.linkedin.com/company/linkedin-mmj-north/", "Organic/Manual",
    "2026-05-07", "MMJNorth_Audit_GenosAI.pptx", "Email Sent - Pending Reply"
]
for col, v in enumerate(organic_vals, 1):
    cell = ws5.cell(9, col)
    cell.value = v
    cell.font = row_font
    cell.fill = row_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border

for col in range(1, ws5.max_column + 1):
    if ws5.cell(1, col).value == "Cold Email Body":
        cell = ws5.cell(9, col)
        cell.value = body
        cell.font = row_font
        cell.fill = row_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        break

ws5.row_dimensions[9].height = 180

wb.save(WB_DIR / "GenosAI_RealEstate_Leads_UPDATED.xlsx")
print("Excel updated - Greg Ellul / MMJ North added to all 5 sheets.")
