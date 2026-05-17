from pathlib import Path
import copy as cp
from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent.parent
WB_PATH = ROOT / "output" / "workbooks" / "GenosAI_OrganicLeads.xlsx"

print(f"Loading: {WB_PATH}")
wb = load_workbook(WB_PATH)

lead = {
    "num":   22,
    "first": "Amel",
    "last":  "Karboul",
    "full":  "Dr. Amel Karboul",
    "title": "CEO",
    "company": "Education Outcomes Fund (EOF)",
    "email": "amel.karboul@educationoutcomesfund.org",
    "email_status": "Inferred from domain pattern -- not publicly listed; contact form at educationoutcomesfund.org/get-in-touch",
    "website":  "https://www.educationoutcomesfund.org/",
    "phone":    "",
    "city":     "London",
    "state":    "England, UK",
    "industry": "Education / Impact Finance / NGO",
    "linkedin": "https://www.linkedin.com/company/the-education-outcomes-fund/",
    "ppt":   "EducationOutcomesFund_Audit_GenosAI.pptx",
    "score": 4.0,
    "automation": (
        "Partner and implementer intake AI (smart triage and routing for NGO/government/investor "
        "proposals currently submitted via a generic Google Form); "
        "donor reporting automation (auto-generate verified outcomes reports per donor "
        "from program data -- CIFF, UBS Optimus, Japan MoFA each have different requirements); "
        "stakeholder communications AI (segmented newsletter/update sequences by audience: "
        "government partner, donor, implementer, researcher); "
        "knowledge management RAG system (searchable AI layer over EOF's published practice notes, "
        "case studies, and technical briefs); "
        "country team coordination AI (async update summaries and scheduling automation "
        "for Sierra Leone, Nigeria, Rwanda, South Africa, Tunisia, Namibia teams); "
        "meeting scheduling automation (Calendly-style booking replacing single contact form "
        "for inbound donor and partner calls)"
    ),
    "notes": (
        "Founded 2018. Hosted by UNICEF. Founded by Education Commission + Global Steering Group "
        "for Impact Investing. Mission: outcomes-based education financing -- payment tied to "
        "verified learning/enrollment/employment results. ~45 staff, London HQ. "
        "Active programs: Sierra Leone (SLEIC, $18M, 100K+ children), Nigeria LEAF ($25M, "
        "200K children, launched March 2026), Rwanda (ECD), South Africa, Tunisia (youth employment). "
        "Namibia in development (government committed $6M+). $130M+ mobilized. 500K+ beneficiaries. "
        "2030 goal: 15 partnerships, 2M children, $350M. "
        "Board: Sir Ronald Cohen (Chair), Jakaya Kikwete, David Sengeh. "
        "Deputy CEO: Max McCabe. Chief Programmes Officer: Milena Castellnou. "
        "Head of Development: Eugenio Donadio. 19,581 LinkedIn followers. "
        "Website built on Wix -- limited automation capability. "
        "Ice-breaker: LEAF Nigeria launch quote (March 2026): 'Real change happens when we stop "
        "paying for promises and start paying for results.' 5 active programs managed by 45 people "
        "-- operational ratio only works with automation."
    ),
}

subject = "5 programs, 45 people, $130 million -- the operational layer that does not scale without automation"

body = """\
Hi Amel,

EOF launched the Lagos Education Access Fund in March 2026 -- the fifth active Outcomes Partnership in four years, with Namibia in development. That is a 5x growth in programme complexity managed by roughly the same team of 45 people. Your quote at the LEAF launch captures the philosophy precisely: real change happens when you stop paying for promises and start paying for results.

The principle applies to your operations too. Right now, every inbound partner proposal, donor inquiry, and government engagement arrives through a single generic contact form and is handled manually. At SLEIC scale, that works. At five simultaneous programmes across six countries -- with different donor reporting requirements for CIFF, UBS Optimus, and the Japan MoFA -- it becomes the bottleneck.

What Genos AI builds for organisations at EOF's stage:

Partner and implementer intake AI -- inbound proposals from NGOs, governments, and investors currently flow through a Google Form with no automated triage. An AI intake system scores and routes each proposal by programme type, geography, and organisational capacity before a programme manager is involved. Cuts weeks off onboarding.

Donor reporting automation -- each of your five active programmes has different reporting requirements across multiple funders. An AI layer that aggregates verified outcomes data and auto-generates funder-specific report drafts eliminates the analyst hours currently spent compiling from raw verification data.

Stakeholder communications AI -- 19,581 LinkedIn followers and a newsletter database with no visible segmentation or automation. Government partners, donors, implementers, and researchers each need different content. Automated sequences for each audience, triggered by role and geography, multiply EOF's reach without adding headcount.

Knowledge management -- EOF publishes practice notes, case studies, and technical briefs that represent irreplaceable institutional knowledge. A RAG-based AI system over that content base means any team member -- from a new Programme Analyst in Rwanda to a Policy Advisor in London -- can query it in seconds.

Country team coordination -- with programme leads in Sierra Leone, Nigeria, Rwanda, South Africa, and Tunisia all reporting to London, async update synthesis and automated scheduling reduces coordination overhead for you and Max significantly.

I'm Rohan Malik, CEO of Genos AI. We build AI automation for mission-driven organisations that need to scale impact without scaling headcount. Happy to show you how it maps to EOF's operational model in 20 minutes.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi Amel, following up on my note. The fastest win for EOF -- an AI intake and routing system for the partner and implementer proposals currently coming through a generic Google Form -- is typically live in under 2 weeks and directly reduces programme team time on unqualified enquiries. Happy to walk through it in 20 minutes. -- Rohan | Genos AI\
"""

fu2 = """\
Amel, last one from me. Five active Outcomes Partnerships across six countries, managed by 45 people, with Namibia in development -- that ratio only holds if the operational layer is automated. That is exactly what we build. Leaving this here if the timing is not right. -- Rohan | Genos AI\
"""


def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font      = cp.copy(src_cell.font)
        dst_cell.fill      = cp.copy(src_cell.fill)
        dst_cell.border    = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


# -- Organic Leads -------------------------------------------------------------
ws  = wb["Organic Leads"]
row = ws.max_row + 1
organic_vals = [
    lead["num"], lead["company"], lead["full"], lead["title"],
    lead["email"], lead["phone"], lead["website"], lead["linkedin"],
    lead["city"], lead["state"], lead["industry"],
    "Organic/Manual", "Pending Send", lead["score"],
    lead["ppt"], lead["automation"], lead["notes"],
]
for col, v in enumerate(organic_vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# -- Cold Emails ---------------------------------------------------------------
ws2  = wb["Cold Emails"]
row2 = ws2.max_row + 1
email_vals = [
    lead["num"], lead["company"], lead["full"],
    lead["email"], subject, body, fu1, fu2,
]
for col, v in enumerate(email_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# -- Pipeline ------------------------------------------------------------------
ws3  = wb["Pipeline"]
row3 = ws3.max_row + 1
pipeline_vals = [
    lead["num"], lead["company"], lead["full"],
    lead["email"], lead["industry"], lead["score"],
    "Email Pending", "2026-05-17", "Send cold email -- pitch AI automation for NGO operations",
    "",
]
for col, v in enumerate(pipeline_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

wb.save(WB_PATH)
print(f"OrganicLeads.xlsx updated -- Amel Karboul / Education Outcomes Fund added to all 3 sheets (row {row}).")
