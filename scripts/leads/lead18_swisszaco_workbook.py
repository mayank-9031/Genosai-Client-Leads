"""
Add Swisszaco Estates (lead #16) to GenosAI_OrganicLeads.xlsx.
Writes to all three sheets: Organic Leads, Cold Emails, Pipeline.
"""
from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import copy as cp

ROOT   = Path(__file__).parent.parent.parent
WB_PATH = ROOT / "output" / "workbooks" / "GenosAI_OrganicLeads.xlsx"

print(f"Loading: {WB_PATH}")
wb = load_workbook(WB_PATH)

# ── Lead data ─────────────────────────────────────────────────────────────────
lead = {
    "company":    "Swisszaco Estates",
    "contact":    "TBD",
    "title":      "TBD",
    "email":      "TBD",
    "phone":      "+2348101658350",
    "website":    "https://swisszacoestates.com/",
    "linkedin":   "https://www.linkedin.com/company/swisscazo-estate/about/",
    "city":       "TBD",
    "country":    "Nigeria",
    "industry":   "Real Estate Development",
    "score":      3,
    "deck":       "output/decks/SwisszacoEstates_Audit_GenosAI.pptx",
    "automation": (
        "WhatsApp AI buyer agent (property inquiry + site visit booking); "
        "agent network recruitment automation (intake, onboarding, nurture); "
        "lead nurture sequences (re-engagement + new listing alerts); "
        "post-handover review and testimonial collection; "
        "diaspora investor nurture (UK/US/Canada) via WhatsApp"
    ),
    "notes": "Contact TBD — outreach via company LinkedIn page. Score 3/10, HIGH priority.",
}

subject = "Swisszaco's agent network and a property inquiry process still stuck on contact forms"

body = """\
Hi,

Came across Swisszaco Estates while researching real estate developers in Nigeria. CAC-registered since March 2012 — 13 years in the market across construction, estate development, architectural design, and urban development. The agent network program stood out as an indicator the business is actively scaling.

One gap that showed up immediately on swisszacoestates.com: there's no way for a property buyer or investor to get a fast response outside business hours. The contact form is the only entry point — no chat widget, no WhatsApp integration, no booking flow for property viewings. In the Nigerian market, where most buyers are comparing 3-4 developers simultaneously and expect a WhatsApp reply within the hour, a contact form with a 24-48 hour turnaround quietly hands those leads to whoever responds first.

The same applies to your agent network. Agents researching whether to join need immediate answers on commission structure, available projects, and onboarding — a contact form puts that decision on hold.

What we build for real estate developers at Genos AI: WhatsApp AI agents that qualify buyers (property type, location, budget, timeline), book site visits automatically, and route hot leads to the right agent — 24/7, no manual handling. We also build agent recruitment flows that capture and nurture new agent sign-ups without your team chasing each one.

I'm Rohan Malik, CEO of Genos AI. Did a full audit of swisszacoestates.com — happy to share it or walk through it in 15 minutes.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi, following up on my last message. The core gap — property inquiries hitting a static contact form while competitors respond on WhatsApp in minutes — is the fastest thing to fix. A WhatsApp AI agent for Swisszaco is typically live within 2 weeks. Happy to show you exactly how it works in 15 minutes. — Rohan | Genos AI\
"""

fu2 = """\
Last one from me. The Swisszaco audit covers the WhatsApp buyer qualification flow, agent recruitment automation, and the property viewing booking system specifically. If the timing isn't right, no problem — leaving it here in case it's useful later. — Rohan | Genos AI\
"""


def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font      = cp.copy(src_cell.font)
        dst_cell.fill      = cp.copy(src_cell.fill)
        dst_cell.border    = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


# ── 1. Organic Leads ──────────────────────────────────────────────────────────
ws1      = wb["Organic Leads"]
next_row = ws1.max_row + 1
last_num = ws1.cell(ws1.max_row, 1).value or 0
lead_num = last_num + 1

# Map headers → values
headers = {ws1.cell(1, c).value: c for c in range(1, ws1.max_column + 1)}

organic_data = {
    "#":                   lead_num,
    "Company":             lead["company"],
    "Contact Person":      lead["contact"],
    "Title":               lead["title"],
    "Email":               lead["email"],
    "Phone":               lead["phone"],
    "Website":             lead["website"],
    "LinkedIn":            lead["linkedin"],
    "City":                lead["city"],
    "State/Country":       lead["country"],
    "Industry":            lead["industry"],
    "Source":              "Organic/LinkedIn",
    "Status":              "Contact TBD — LinkedIn Pending",
    "Audit Score":         lead["score"],
    "Deck Path":           lead["deck"],
    "Automation Opportunities": lead["automation"],
    "Notes":               lead["notes"],
}

for header, value in organic_data.items():
    if header in headers:
        col  = headers[header]
        cell = ws1.cell(next_row, col)
        cell.value = value
        copy_style(ws1.cell(next_row - 1, col), cell)

print(f"Organic Leads: wrote row {next_row} (lead #{lead_num})")


# ── 2. Cold Emails ────────────────────────────────────────────────────────────
ws2       = wb["Cold Emails"]
next_row2 = ws2.max_row + 1
headers2  = {ws2.cell(1, c).value: c for c in range(1, ws2.max_column + 1)}

email_data = {
    "#":           lead_num,
    "Company":     lead["company"],
    "Contact":     lead["contact"],
    "Email":       lead["email"],
    "Subject":     subject,
    "Body":        body,
    "Follow-Up 1": fu1,
    "Follow-Up 2": fu2,
}

for header, value in email_data.items():
    if header in headers2:
        col  = headers2[header]
        cell = ws2.cell(next_row2, col)
        cell.value = value
        copy_style(ws2.cell(next_row2 - 1, col), cell)

print(f"Cold Emails: wrote row {next_row2}")


# ── 3. Pipeline ───────────────────────────────────────────────────────────────
ws3       = wb["Pipeline"]
next_row3 = ws3.max_row + 1
headers3  = {ws3.cell(1, c).value: c for c in range(1, ws3.max_column + 1)}

pipeline_data = {
    "#":              lead_num,
    "Company":        lead["company"],
    "Contact":        lead["contact"],
    "Email":          lead["email"],
    "Industry":       lead["industry"],
    "Audit Score":    lead["score"],
    "Status":         "Contact TBD",
    "Last Touchpoint": "2026-05-17",
    "Next Action":    "Source contact via LinkedIn — send cold email + deck",
    "Est. Deal Value": "$3,000 - $6,000",
}

for header, value in pipeline_data.items():
    if header in headers3:
        col  = headers3[header]
        cell = ws3.cell(next_row3, col)
        cell.value = value
        copy_style(ws3.cell(next_row3 - 1, col), cell)

print(f"Pipeline: wrote row {next_row3}")


# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(WB_PATH)
print(f"\nDone — Swisszaco Estates (#{lead_num}) added to all 3 sheets in GenosAI_OrganicLeads.xlsx")
