from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import copy as cp
import glob

ROOT   = Path(__file__).parent.parent.parent
WB_DIR = ROOT / "output" / "workbooks"

files = glob.glob(str(WB_DIR / "GenosAI_RealEstate_Leads*.xlsx"))
src   = next((f for f in files if "UPDATED" in f), str(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx"))
print(f"Loading: {src}")
wb = load_workbook(src)

lead = {
    "num":          107,
    "first":        "TBD",
    "last":         "TBD",
    "full":         "TBD",
    "title":        "TBD",
    "company":      "Swisszaco Estates",
    "email":        "TBD",
    "email_status": "Not yet sourced — outreach via company LinkedIn page",
    "website":      "https://swisszacoestates.com/",
    "phone":        "+2348101658350",
    "city":         "TBD",
    "state":        "TBD",
    "country":      "Nigeria",
    "employees":    "TBD",
    "industry":     "Real Estate Development",
    "linkedin":     "https://www.linkedin.com/company/swisscazo-estate/about/",
    "ppt":          "SwisszacoEstates_Audit_GenosAI.pptx",
}

subject = "Swisszaco's agent network and a property inquiry process still stuck on contact forms"

body = """\
Hi,

Came across Swisszaco Estates while researching real estate developers in Nigeria. CAC-registered since March 2012 — 13 years in the market across construction, estate development, architectural design, and urban development. The agent network program stood out as an indicator the business is actively scaling.

One gap that showed up immediately on swisszacoestates.com: there's no way for a property buyer or investor to get a fast response outside business hours. The contact form is the only entry point — no chat widget, no WhatsApp integration, no booking flow for property viewings. In the Nigerian market, where most buyers are comparing 3–4 developers simultaneously and expect a WhatsApp reply within the hour, a contact form with a 24–48 hour turnaround quietly hands those leads to whoever responds first.

The same applies to your agent network. Agents researching whether to join need immediate answers on commission structure, available projects, and onboarding — a contact form puts that decision on hold.

What we build for real estate developers at Genos AI: WhatsApp AI agents that qualify buyers (property type, location, budget, timeline), book site visits automatically, and route hot leads to the right agent — 24/7, no manual handling. We also build agent recruitment flows that capture and nurture new agent sign-ups without your team chasing each one.

I'm Rohan Malik, CEO of Genos AI. Did a full audit of swisszacoestates.com — happy to share it or walk through it in 15 minutes.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi, following up on my last message. The core gap — property inquiries hitting a static contact form while competitors respond on WhatsApp in minutes — is the fastest thing to fix. A WhatsApp AI agent for Swisszaco is typically live within 2 weeks. Happy to show you exactly how it works in 15 minutes. — Rohan | Genos AI\
"""

fu2 = """\
Last one from me. The Swisszaco audit covers the WhatsApp buyer qualification flow, agent recruitment automation, and the property viewing booking system specifically. If the timing isn't right, no problem — leaving it here in case it's useful later. — Rohan | Genos AI\
"""


def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font      = cp.copy(src_cell.font)
        dst_cell.fill      = cp.copy(src_cell.fill)
        dst_cell.border    = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


# ── Leads ──────────────────────────────────────────────────────────────────
ws   = wb["Leads"]
row  = 108
vals = [
    lead["num"], lead["first"], lead["last"], lead["full"], lead["title"],
    lead["company"], lead["email"], lead["email_status"], lead["website"],
    lead["phone"], lead["city"], lead["state"], lead["country"],
    lead["employees"], lead["industry"], lead["linkedin"], lead["ppt"],
]
for col, v in enumerate(vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# ── Website Audits ─────────────────────────────────────────────────────────
ws2 = wb["Website Audits"]
audit_vals = [
    lead["num"], lead["company"], lead["website"],
    "WordPress (Astra theme)", "Yes", "Yes (basic GTM/GA)", "No",
    "No", "No", "No", "No", "Yes (remarketing pixel)", 3, "High",
    (
        "No WhatsApp integration — critical failure for Nigerian real estate market where buyers and investors "
        "expect an immediate WhatsApp response; contact form is the only inquiry channel with no automated "
        "follow-up; no live chat or AI chatbot — after-hours and weekend inquiries go unanswered; no online "
        "property viewing / site visit booking system; no email marketing or lead nurture sequence for buyers "
        "who express interest; no review or testimonial collection system beyond a single static quote; no "
        "property search or filter functionality visible to site visitors; no CDN — potential performance "
        "issues for diaspora audience outside Nigeria; no mortgage calculator or investment ROI tool; "
        "agent network recruitment has no automated onboarding or follow-up flow"
    ),
    (
        "1. WhatsApp AI agent: qualifies buyers (property type, budget, location, timeline), books site visits "
        "automatically, routes hot leads to sales agents — 24/7 response | "
        "2. Agent recruitment automation: capture agent sign-ups, deliver onboarding info, follow up with "
        "uncommitted candidates via WhatsApp | "
        "3. Lead nurture email + WhatsApp sequences for buyers who expressed interest but didn't commit | "
        "4. Automated review/testimonial request flow post-purchase or post-site-visit | "
        "5. Property inquiry routing: international/diaspora leads handled separately with timezone-aware "
        "follow-up | "
        "6. Investment calculator or ROI estimator as a lead magnet with WhatsApp capture"
    ),
]
row2 = 108
for col, v in enumerate(audit_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# ── Cold Emails ────────────────────────────────────────────────────────────
ws3       = wb["Cold Emails"]
email_vals = [lead["num"], lead["first"], lead["company"], lead["email"], subject, body, fu1, fu2]
row3      = 108
for col, v in enumerate(email_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

# ── Audit PPT Content ──────────────────────────────────────────────────────
ws4 = wb["Audit PPT Content"]
ppt_vals = [
    lead["num"], lead["company"], lead["full"], lead["email"], lead["website"], 3, "High",
    "Swisszaco Estates - Website & AI Automation Audit | Prepared by Genos AI",
    (
        "Swisszaco Estates (Swisszaco Ltd) | Associate of Ogorstar Conglomerates Inc | "
        "Registered in Nigeria, C.A.C. 13 March 2012 | Multi-national real estate developer | "
        "Phone: +2348101658350 | Website: swisszacoestates.com | LinkedIn: swisscazo-estate | "
        "Services: Construction & Sales, Architectural Designs, Branding & Renovations, "
        "Estate Development, Urban Development, In-House Surveyors | "
        "Agent Network Program active — actively recruiting real estate agents | "
        "13 years in the Nigerian real estate market"
    ),
    (
        "13 years of market presence — strong brand credibility and operational track record; "
        "multi-national structure (Ogorstar Conglomerates) signals serious scaling infrastructure; "
        "diverse service offering across the full real estate value chain (construction, design, renovation, "
        "development, surveying); active agent network = scalable distribution model; "
        "remarketing pixel in place = some digital marketing awareness; "
        "mobile-responsive site with analytics tracking = foundation exists to build on"
    ),
    (
        "No WhatsApp channel — critical gap in Nigerian market; contact form only with no automated response; "
        "no live chat or AI chatbot for after-hours inquiries; no online property viewing/site visit booking; "
        "no email marketing or lead nurture automation; no review collection system beyond one static quote; "
        "no property search or filter tool for visitors; no CDN affecting load time for diaspora audience; "
        "agent recruitment has no automated onboarding or follow-up; no mortgage or ROI calculator as lead magnet"
    ),
    (
        "1. WhatsApp AI agent: buyer qualification, viewing booking, agent routing | "
        "2. Agent recruitment automation: sign-up capture, onboarding delivery, nurture sequences | "
        "3. Post-inquiry lead nurture: WhatsApp + email sequences for warm but unconverted prospects | "
        "4. Automated review requests post-sale and post-site-visit | "
        "5. International/diaspora lead routing with timezone-aware follow-up | "
        "6. Investment ROI calculator as lead magnet with contact capture"
    ),
    (
        "Swisszaco Estates has 13 years of real estate experience and a multi-national holding structure. "
        "The operational infrastructure is real. The gap is entirely on the digital front door: a property "
        "buyer comparing three Nigerian developers on a Sunday evening gets an immediate WhatsApp reply from "
        "two of them and a contact form from Swisszaco. Genos AI builds the automation layer that ensures "
        "Swisszaco responds first, qualifies the lead automatically, and books the site visit — without "
        "adding headcount."
    ),
    (
        "Website: add WhatsApp chat widget (immediate), property viewing booking flow, investment ROI "
        "calculator, agent network landing page with automated onboarding | "
        "AI automation: WhatsApp AI agent for buyer qualification and visit booking, agent recruitment flow, "
        "post-inquiry nurture sequences, diaspora/international lead routing | "
        "Marketing: remarketing pixel already in place — connect to lead capture and retarget warm visitors; "
        "diaspora-focused campaigns via WhatsApp | "
        "Priority order: WhatsApp agent first (fastest ROI), agent recruitment flow second, nurture + "
        "retargeting third"
    ),
    (
        "Week 1: WhatsApp AI agent live — trained on property types, locations, pricing tiers, and site "
        "visit booking; contact form replaced with WhatsApp widget on all key pages | "
        "Week 2: Agent recruitment landing page with automated onboarding WhatsApp sequence; "
        "post-inquiry lead nurture flow active (3-touch WhatsApp + email) | "
        "Month 2: Investment ROI calculator deployed as lead magnet; diaspora inquiry routing live "
        "(UK, US, Canada numbers handled); automated review request sequence post-sale | "
        "Month 3: Full remarketing integration — retarget site visitors who didn't convert; "
        "agent performance dashboard for tracking leads sourced per agent"
    ),
    "Reply to this email or WhatsApp | Rohan Malik, CEO | hello@genosai.tech | www.genosai.tech | +91 638-714-2699",
]
row4 = 108
for col, v in enumerate(ppt_vals, 1):
    cell = ws4.cell(row4, col)
    cell.value = v
    copy_style(ws4.cell(row4 - 1, col), cell)

# ── Organic Leads ──────────────────────────────────────────────────────────
ws5       = wb["Organic Leads"]
thin      = Side(style="thin", color="444444")
border    = Border(left=thin, right=thin, top=thin, bottom=thin)
row_fill  = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
row_font  = Font(name="Calibri", size=10, color="1A1A2E")

organic_vals = [
    18, "TBD", "TBD", "TBD", "TBD",
    "Swisszaco Estates", "TBD",
    "Not yet sourced — outreach via company LinkedIn page",
    "https://swisszacoestates.com/",
    "+2348101658350",
    "TBD", "TBD", "Nigeria", "TBD", "Real Estate Development",
    "https://www.linkedin.com/company/swisscazo-estate/about/",
    "Organic/Manual",
    "2026-05-17", "SwisszacoEstates_Audit_GenosAI.pptx", "Contact TBD — LinkedIn Pending",
]
for col, v in enumerate(organic_vals, 1):
    cell             = ws5.cell(19, col)
    cell.value       = v
    cell.font        = row_font
    cell.fill        = row_fill
    cell.alignment   = Alignment(vertical="center", wrap_text=True)
    cell.border      = border

for col in range(1, ws5.max_column + 1):
    if ws5.cell(1, col).value == "Cold Email Body":
        cell             = ws5.cell(19, col)
        cell.value       = body
        cell.font        = row_font
        cell.fill        = row_fill
        cell.alignment   = Alignment(vertical="top", wrap_text=True)
        cell.border      = border
        break

ws5.row_dimensions[19].height = 180

wb.save(WB_DIR / "GenosAI_RealEstate_Leads_UPDATED.xlsx")
print("Done — Swisszaco Estates added to all 5 sheets (lead #107, organic #18).")
