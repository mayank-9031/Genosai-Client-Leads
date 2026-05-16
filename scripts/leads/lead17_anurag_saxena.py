from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import copy as cp
import os

ROOT = Path(__file__).parent.parent.parent
WB_DIR = ROOT / "output" / "workbooks"

import glob
files = glob.glob(str(WB_DIR / "GenosAI_RealEstate_Leads*.xlsx"))
src = next((f for f in files if "UPDATED" in f), str(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx"))
print(f"Loading: {src}")
wb = load_workbook(src)

lead = {
    "num": 106,
    "first": "Anurag",
    "last": "Saxena",
    "full": "Anurag Saxena",
    "title": "CEO",
    "company": "Aqua Technic",
    "email": "enquiry@aquatechnic.co.in",
    "email_status": "Business inquiry email (listed on site)",
    "website": "https://www.aquatechnic.co.in/",
    "phone": "+91-9391355379",
    "city": "Hyderabad",
    "state": "Telangana",
    "country": "India",
    "employees": "11-50",
    "industry": "Swimming Pool Construction & Maintenance",
    "linkedin": "https://in.linkedin.com/company/aquatechnic",
    "ppt": "AquaTechnic_Audit_GenosAI.pptx"
}

subject = "505 pool projects and weekend inquiries hitting a static contact page"

body = """\
Hi Anurag,

Came across Aqua Technic while researching the Indian pool construction market. 505 projects since 2009 with operations across India, Gulf, and Australia - that is a serious track record.

One thing stood out on aquatechnic.co.in. The most common question a hotel developer or homeowner has - "how much will a pool cost?" - lands on a contact form with no immediate response. In India's market, that prospect is also sending the same message to 3 other builders on WhatsApp while waiting for your reply.

The businesses winning high-value pool construction contracts right now are the ones responding in minutes, not days. A WhatsApp AI agent on your site that answers cost questions, qualifies the project (size, type, hotel or residential, location), and books a site visit call would handle that gap automatically - 24/7, including Sundays when most developers are actually browsing.

Beyond new inquiries: 490 maintenance clients is a significant AMC portfolio to manage manually. Automated AMC renewal reminders, service scheduling, and client communication via WhatsApp reduces churn and frees your team for the actual technical work.

I'm Rohan Malik, CEO of Genos AI. We build AI automation specifically for construction and service businesses - WhatsApp agents, lead qualification flows, and maintenance scheduling automation. Did a quick audit of aquatechnic.co.in. Happy to share it or walk through it in 15 minutes.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi Anurag, following up on my last message. The specific point - weekend inquiries hitting a static form while WhatsApp competitors respond in minutes - is the fastest gap to close. A WhatsApp AI agent is typically live in under 2 weeks. Happy to show you how it works in 15 minutes. - Rohan | Genos AI\
"""

fu2 = """\
Anurag, last one from me. The Aqua Technic audit covers the WhatsApp automation, AMC management, and lead qualification flows specifically. If the timing isn't right now, no problem - leaving it here in case it's useful. - Rohan | Genos AI\
"""

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = cp.copy(src_cell.font)
        dst_cell.fill = cp.copy(src_cell.fill)
        dst_cell.border = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

ws = wb["Leads"]
row = 107
vals = [lead["num"], lead["first"], lead["last"], lead["full"], lead["title"],
        lead["company"], lead["email"], lead["email_status"], lead["website"],
        lead["phone"], lead["city"], lead["state"], lead["country"],
        lead["employees"], lead["industry"], lead["linkedin"], lead["ppt"]]
for col, v in enumerate(vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

ws2 = wb["Website Audits"]
audit_vals = [
    lead["num"], lead["company"], lead["website"],
    "Custom PHP / WordPress", "Yes", "Partial (basic form)", "No",
    "No", "No", "Partial (stats only - no testimonials system)", "No", "No", 5, "High",
    "Site is functional but leaks high-value leads at every stage: no WhatsApp chat widget (critical failure for Indian B2B/B2C market - hotel developers and homeowners expect instant WhatsApp response); contact form with no automated follow-up - inquiry sits until someone manually replies; no AI cost estimator or ballpark quote tool - the most common question ('how much does a pool cost?') gets no immediate answer; placeholder/test entries visible in navigation indicate site maintenance issues affecting credibility; no automated AMC renewal reminders for 490 maintenance clients; no online maintenance scheduling or client portal; no project gallery with filterable search by type (hotel, residential, sports); no post-project review collection automation despite 505 completed projects; no lead nurture sequence for hotel/resort developer prospects; international inquiries (Gulf, Australia) get same generic form as local",
    "Deploy WhatsApp AI agent that answers cost estimation questions, qualifies project (pool type, size, hotel or residential, city, timeline) and books site visit - responds in minutes 24/7; build automated AMC renewal workflow: 30-day, 7-day, and 1-day WhatsApp reminders to 490 maintenance clients with payment link; add pool cost estimator tool on homepage as lead magnet with WhatsApp/email capture before showing estimate; fix navigation placeholder issues (immediate credibility improvement); build automated post-project review request sequence via WhatsApp at project completion; add filterable project gallery by category (hotels, residential, sports, water features); separate landing pages for hotel/resort clients vs residential with tailored qualification flows; international inquiry routing for Gulf and Australia with timezone-aware follow-up"
]
row2 = 107
for col, v in enumerate(audit_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

ws3 = wb["Cold Emails"]
email_vals = [lead["num"], lead["first"], lead["company"], lead["email"], subject, body, fu1, fu2]
row3 = 107
for col, v in enumerate(email_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

ws4 = wb["Audit PPT Content"]
ppt_vals = [
    lead["num"], lead["company"], lead["full"], lead["email"], lead["website"], 5, "High",
    "Aqua Technic - Website & AI Automation Audit | Prepared by Genos AI",
    "Anurag Saxena: CEO of Aqua Technic | Location: Hyderabad, Telangana | Phone: +91-9391355379 | Email: enquiry@aquatechnic.co.in | Company founded 2009 | Self-described as 'India's Largest Pool Builder' | 505 projects, 490 happy clients, 5 top pool builder awards, 12+ years experience | Certified Technician Based Pool Service Company | Industries served: hospitality, leisure, sports, wellness | International presence: Gulf and Australia | Services: turnkey pool construction, renovation, maintenance, electrical, mechanical, aesthetics, repair, waterproofing, water features, hydrotherapy equipment | Franchise operations mentioned",
    "505 completed projects = strong operational credibility and track record; 490 active maintenance clients = large recurring revenue base worth protecting with automation; international presence in Gulf and Australia shows scaling ambition; 5 top pool builder awards provide third-party validation; 12+ years in market = established brand in Hyderabad with growth potential nationally; certified technician-based service = quality positioning that automation amplifies rather than compromises; franchise model = automation infrastructure can be replicated across franchise network",
    "No WhatsApp chat widget - critical gap for Indian market where hotel developers and homeowners expect instant WhatsApp response; contact form only with no automated follow-up - inquiries sit until manual reply; no AI cost estimator - 'how much does a pool cost?' gets no immediate answer; navigation has placeholder/test entries - active credibility damage; 490 AMC clients managed with no automated renewal reminders or scheduling; no post-project review automation despite 505 completed projects; no filterable project gallery; no separate qualification flows for hotel vs residential vs international clients",
    "1. WhatsApp AI agent: cost qualification, project scoping, site visit booking - responds in minutes 24/7 | 2. AMC renewal automation: WhatsApp reminders at 30/7/1 days with payment link for 490 clients | 3. Pool cost estimator with WhatsApp/email capture before showing estimate | 4. Post-project review request via WhatsApp at completion | 5. Fix navigation placeholders (immediate credibility win) | 6. Hotel vs residential separate landing pages with tailored intake",
    "Aqua Technic has 505 projects and 490 maintenance clients built over 12 years. The operational track record is real. The gap is that a hotel developer comparing 4 pool builders on a Sunday afternoon gets an immediate WhatsApp reply from 3 of them and a contact form from Aqua Technic. Genos AI builds the automation layer that ensures Aqua Technic is always the first to respond - and the best qualified.",
    "Website: fix navigation placeholders, add WhatsApp widget, pool cost estimator, filterable project gallery | AI automation: WhatsApp AI agent for lead qualification, AMC renewal workflow, post-project review sequence, hotel vs residential separate intake flows | SEO: pool construction Hyderabad, hotel pool builders India, swimming pool maintenance Hyderabad | Implementation: WhatsApp agent + AMC reminders first (fastest ROI), cost estimator + gallery second, SEO third",
    "Week 1: WhatsApp AI agent deployed trained on pool types, cost ranges, project qualification; navigation placeholders fixed; AMC renewal WhatsApp automation live for 490 clients | Week 2: Pool cost estimator live with WhatsApp/email capture gate; post-project review request sequence active; hotel vs residential landing pages built | Month 2: First 10 SEO articles (pool construction Hyderabad, hotel pool India, swimming pool maintenance); filterable project gallery with 50+ past project images | Month 3: Gulf and Australia international inquiry routing; franchise network automation template; chatbot data informing ICP refinement",
    "Reply to this email or WhatsApp | Rohan Malik, CEO | hello@genosai.tech | www.genosai.tech | +91 638-714-2699"
]
row4 = 107
for col, v in enumerate(ppt_vals, 1):
    cell = ws4.cell(row4, col)
    cell.value = v
    copy_style(ws4.cell(row4 - 1, col), cell)

ws5 = wb["Organic Leads"]
thin = Side(style="thin", color="444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
row_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
row_font = Font(name="Calibri", size=10, color="1A1A2E")

organic_vals = [
    17, "Anurag", "Saxena", "Anurag Saxena", "CEO",
    "Aqua Technic", "enquiry@aquatechnic.co.in",
    "Business inquiry email (listed on site)", "https://www.aquatechnic.co.in/",
    "+91-9391355379",
    "Hyderabad", "Telangana", "India", "11-50", "Swimming Pool Construction & Maintenance",
    "https://in.linkedin.com/company/aquatechnic", "Organic/Manual",
    "2026-05-11", "AquaTechnic_Audit_GenosAI.pptx", "Email Sent - Pending Reply"
]
for col, v in enumerate(organic_vals, 1):
    cell = ws5.cell(18, col)
    cell.value = v
    cell.font = row_font
    cell.fill = row_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border

for col in range(1, ws5.max_column + 1):
    if ws5.cell(1, col).value == "Cold Email Body":
        cell = ws5.cell(18, col)
        cell.value = body
        cell.font = row_font
        cell.fill = row_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        break

ws5.row_dimensions[18].height = 180

wb.save(WB_DIR / "GenosAI_RealEstate_Leads_UPDATED.xlsx")
print("Excel updated - Anurag Saxena / Aqua Technic added to all 5 sheets.")
