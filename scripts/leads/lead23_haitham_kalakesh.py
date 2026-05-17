from pathlib import Path
import copy as cp
from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent.parent
WB_PATH = ROOT / "output" / "workbooks" / "GenosAI_OrganicLeads.xlsx"

print(f"Loading: {WB_PATH}")
wb = load_workbook(WB_PATH)

lead = {
    "num":   21,
    "first": "Haitham",
    "last":  "Kalakesh",
    "full":  "Haitham Kalakesh",
    "title": "Founder & CEO",
    "company": "Reve Real Estate / Riversong Technology",
    "email": "info@riversongtechnology.com",
    "email_status": "General contact — personal email not publicly listed",
    "website":  "https://reverealestates.com/home",
    "phone":    "+971 4 583 9860",
    "city":     "Dubai",
    "state":    "Dubai, UAE",
    "industry": "Real Estate / Consumer Electronics",
    "linkedin": "https://www.linkedin.com/in/haithamkalakesh/",
    "ppt":   "HaithamKalakesh_Audit_GenosAI.pptx",
    "score": 4.5,
    "automation": (
        "Reve Real Estate: property inquiry AI via WhatsApp and web chat (capture name, budget, "
        "property type, timeline before agent involvement); investor nurture sequences for leads "
        "who enquired but did not book a viewing; Golden Visa consultation funnel (education "
        "sequence to qualified call); developer launch alerts to buyer database. "
        "Riversong Technology: distributor and retailer B2B inquiry routing AI; "
        "customer support automation for product queries, warranty, and returns; "
        "post-purchase review and warranty registration sequences; B2B partnership "
        "qualification flow for new market entry. Cross-business: unified CRM pipeline "
        "and executive reporting AI covering both companies for Haitham."
    ),
    "notes": (
        "Dual-founder profile: Reve Real Estate (reverealestates.com) — Dubai HQ, offices in "
        "London, Beirut, Amman. Services: property sales, management, investment, Golden Visa "
        "consultation. Developer partners: Emaar, Damac, Meraas, Select Group, Ellington, "
        "Majid Al Futtaim. Website score 3/10 (CRITICAL — no chat, no booking, no lead capture, "
        "no reviews, no visible contact). "
        "Riversong Technology (riversongtechnology.com) — consumer electronics: smart watches, "
        "earbuds, phones, smart TVs, ACs. Dubai JLT X3 Tower, London, Cairo. "
        "VP Levant: Mutasem Shehadeh. Website score 6/10 (no chat, no WhatsApp, no reviews). "
        "Ice-breaker: running two distinct businesses (luxury real estate + consumer electronics) "
        "from Dubai across 5+ markets is an unusual operational profile. "
        "Personal LinkedIn found but inaccessible (HTTP 999)."
    ),
}

subject = "Two Dubai businesses, one missing layer — AI automation for Reve Real Estate and Riversong Technology"

body = """\
Hi Haitham,

Running two distinct businesses from Dubai -- a luxury real estate portal operating across Dubai, London, Beirut, and Amman, and a consumer electronics brand distributed across the Middle East and beyond -- means two separate inquiry pipelines, two support channels, and two sets of leads that currently depend on manual follow-up.

The gap in both businesses is the same: someone inquires, and without a structured automation layer, the follow-up quality depends entirely on who picks it up and when.

What Genos AI builds for each:

For Reve Real Estate:

Property inquiry AI -- a WhatsApp and web chat flow that captures budget, property type, location preference, and timeline before the agent is involved. Emaar, Damac, and Meraas developers bring serious buyer traffic; that traffic needs qualification before it reaches your team.

Investor nurture sequences -- leads who submitted an inquiry but didn't book a viewing receive a structured follow-up over 4--6 weeks: market data, featured projects, and a re-engagement prompt. Most real estate deals close on the fourth or fifth follow-up, not the first.

Golden Visa consultation funnel -- an automated education sequence for international investors asking about UAE residency through property investment. Qualified leads routed to a consultation call; unqualified leads nurtured until ready.

For Riversong Technology:

Distributor and retailer inquiry routing -- a B2B inquiry AI that qualifies new distribution or retail partnership enquiries by region, category, and order volume before connecting to your sales team.

Customer support automation -- product questions, warranty claims, and return requests handled by AI for the consumer side. With a product range across smart watches, earbuds, TVs, and ACs, the support query volume doesn't scale manually.

I'm Rohan Malik, CEO of Genos AI. We build AI automation for businesses operating at this kind of operational scale. Happy to show you how it maps to both companies in 20 minutes.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi Haitham, following up on my last message. The fastest win for Reve Real Estate -- a WhatsApp inquiry AI that pre-qualifies buyers before they reach your agents -- is typically live in under 2 weeks and directly protects agent time on your highest-value listings. Happy to walk through it in 20 minutes. -- Rohan | Genos AI\
"""

fu2 = """\
Haitham, last one from me. Two businesses across real estate and consumer electronics, both operating across multiple markets, both with inquiry pipelines that scale only as far as the team managing them. That is exactly the problem AI automation solves. Leaving this here if the timing is not right. -- Rohan | Genos AI\
"""


def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font      = cp.copy(src_cell.font)
        dst_cell.fill      = cp.copy(src_cell.fill)
        dst_cell.border    = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


# -- Organic Leads -------------------------------------------------------------
ws  = wb["Organic Leads"]
row = ws.max_row + 1
organic_vals = [
    lead["num"], lead["company"], lead["full"], lead["title"],
    lead["email"], lead["phone"], lead["website"], lead["linkedin"],
    lead["city"], lead["state"], lead["industry"],
    "Organic/Manual", "Pending Send", lead["score"],
    lead["ppt"], lead["automation"], lead["notes"],
]
for col, v in enumerate(organic_vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# -- Cold Emails ---------------------------------------------------------------
ws2  = wb["Cold Emails"]
row2 = ws2.max_row + 1
email_vals = [
    lead["num"], lead["company"], lead["full"],
    lead["email"], subject, body, fu1, fu2,
]
for col, v in enumerate(email_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# -- Pipeline ------------------------------------------------------------------
ws3  = wb["Pipeline"]
row3 = ws3.max_row + 1
pipeline_vals = [
    lead["num"], lead["company"], lead["full"],
    lead["email"], lead["industry"], lead["score"],
    "Email Pending", "2026-05-17", "Send cold email -- pitch AI automation (dual audit)",
    "",
]
for col, v in enumerate(pipeline_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

wb.save(WB_PATH)
print(f"OrganicLeads.xlsx updated -- Haitham Kalakesh / Reve Real Estate & Riversong added to all 3 sheets (row {row}).")
