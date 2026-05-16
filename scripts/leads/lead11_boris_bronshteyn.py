from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import copy as cp
import os

ROOT = Path(__file__).parent.parent.parent
WB_DIR = ROOT / "output" / "workbooks"

import glob
files = glob.glob(str(WB_DIR / "GenosAI_RealEstate_Leads*.xlsx"))
src = next((f for f in files if "UPDATED" in f), str(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx"))
print(f"Loading: {src}")
wb = load_workbook(src)

lead = {
    "num": 100,
    "first": "Boris",
    "last": "Bronshteyn",
    "full": "Boris Bronshteyn",
    "title": "CEO",
    "company": "Bronshteyn GmbH",
    "email": "boris@bronshteyn.com",
    "email_status": "Valid",
    "website": "https://www.bronshteyn.com/",
    "phone": "",
    "city": "Munich",
    "state": "Bavaria",
    "country": "DE",
    "employees": "11-50",
    "industry": "Real Estate - Family Office / Investment Advisory",
    "linkedin": "https://www.linkedin.com/company/bronshteyn/about/",
    "ppt": "Bronshteyn_Audit_GenosAI.pptx"
}

subject = "Capital Factory mentor, Munich family office deal flow, and a website that's still a contact card"

body = """\
Hi Boris! Rohan here - CEO of Genos AI. Came across Bronshteyn GmbH and really liked what you're building across Munich real estate and M&A advisory.

We help family offices and investment advisories with two things - AI automation (deal flow qualification, investor follow-up sequences) and website development (building sites that work as a credibility asset when HNW prospects research you before the first call). Put together a quick audit for you and thought it might be useful.

Also noticed you advise WhiteWill - great operation. The same stack scales well into something like that too if that's ever a conversation worth having.

Happy to walk through the audit whenever suits you, no pressure at all! - Rohan | genosai.tech\
"""

fu1 = """\
Hi Boris, just making sure my last message didn't get lost! The audit for Bronshteyn GmbH is ready whenever you'd like - happy to walk through it in 15 mins. - Rohan\
"""

fu2 = """\
Boris, last one from me! Totally understand if the timing isn't right - leaving the audit here whenever it's useful. Would love to connect down the road. - Rohan\
"""

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = cp.copy(src_cell.font)
        dst_cell.fill = cp.copy(src_cell.fill)
        dst_cell.border = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

# Leads
ws = wb["Leads"]
row = 101
vals = [lead["num"], lead["first"], lead["last"], lead["full"], lead["title"],
        lead["company"], lead["email"], lead["email_status"], lead["website"],
        lead["phone"], lead["city"], lead["state"], lead["country"],
        lead["employees"], lead["industry"], lead["linkedin"], lead["ppt"]]
for col, v in enumerate(vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# Website Audits
ws2 = wb["Website Audits"]
audit_vals = [
    lead["num"], lead["company"], lead["website"],
    "Unknown / Single-Page Landing Site", "No", "No", "No",
    "No", "None", "No", "No", "No", 3, "High",
    "bronshteyn.com is a single-page placeholder with only a name, subtitle (Family Office | Real estate), contact email, WhatsApp link, and Telegram handle; no navigation, no service descriptions, no deal track record, no team credentials, no testimonials or social proof, no case studies, no pricing or service scope, no blog or content hub, no SEO infrastructure, no lead capture form or qualifying inquiry flow, no trust indicators (certifications, awards, years in business, deal volume), no investor-grade presentation of Bronshteyn GmbH operations; the site functions as a contact card only and fails the credibility filter for HNW investors or M&A targets doing pre-call research",
    "Full site rebuild as family office authority: services scoped (investment mediation, M&A advisory, business acquisitions), deal track record where disclosable, team credentials for Boris/Anastasia/Ksenia, investor-grade design; AI chat for inbound inquiry qualification (deal size, asset class, geography screening before reaching Boris directly); investor follow-up automation for first-contact prospects who don't convert immediately; content layer with German real estate market commentary and Munich deal flow observations for search authority; structured inquiry routing by deal type (direct acquisition, co-investment, advisory mandate) replacing undifferentiated message channels"
]
row2 = 101
for col, v in enumerate(audit_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# Cold Emails
ws3 = wb["Cold Emails"]
email_vals = [lead["num"], lead["first"], lead["company"], lead["email"], subject, body, fu1, fu2]
row3 = 101
for col, v in enumerate(email_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

# Audit PPT Content
ws4 = wb["Audit PPT Content"]
ppt_vals = [
    lead["num"], lead["company"], lead["full"], lead["email"], lead["website"], 3, "High",
    "Bronshteyn GmbH - Website & AI Growth Audit | Prepared by Genos AI",
    "Boris Bronshteyn: CEO of Bronshteyn GmbH | Munich, Bavaria, Germany | Family office focused on investment real estate in Germany | Services: real estate mediation, family office management, M&A consulting, business consulting, marketing and PR, company acquisition advisory | Team: Anastasia Gerasimova (Head of Real Estate), Ksenia Kuligina (Key Account Manager) | 11 employees | Registered: Amtsgericht Munich #212179 / Leibnizstrasse 35, 80686 Munich | Capital Factory mentor (Austin, TX startup accelerator) indicating tech-world exposure alongside institutional real estate operations | Direct contact: boris@bronshteyn.com / Telegram @theboris7 / WhatsApp",
    "Munich family office model is relationship-driven and trust-based - exactly the context where digital presence acts as a credibility filter before the first call; Bronshteyn GmbH covers real estate mediation, M&A advisory, and business acquisitions under one operation - a breadth that most Munich advisories do not match; Capital Factory mentorship indicates Boris is comfortable with tech-enabled operations and AI tools - not a skeptic; 11-person team with dedicated Head of Real Estate and Key Account Manager means operational capacity exists to handle deal flow if inbound is improved; German real estate investment market is large, relationship-dense, and underdigitised - a family office that presents professionally online has a structural advantage over ones that do not",
    "bronshteyn.com is a single-page placeholder: name, subtitle, email, WhatsApp, Telegram - nothing else; no service descriptions, no scope, no track record, no credentials; no team presentation for Anastasia or Ksenia; no case studies or deal history where disclosable; no blog or content hub; no SEO infrastructure (single page cannot rank for any target keyword); no lead capture form or structured inquiry routing - inbound arrives undifferentiated via three different message channels; no trust indicators (years in business, deal volume, certifications); no investor-grade design or presentation matching HNW client expectations; the site fails the credibility filter a prospective investor or M&A target applies before deciding to call",
    "1. Full site rebuild as Munich family office authority: services scoped, track record presented, team credentials, investor-grade design | 2. AI chat for inbound inquiry qualification: deal size, asset class, geography screening before Boris's direct involvement | 3. Investor and seller follow-up automation: HNW relationships close on multiple touchpoints, not one | 4. Content layer: German real estate market commentary, Munich deal flow observations, M&A advisory insights for search authority | 5. Structured inquiry routing by deal type: direct acquisition, co-investment, advisory mandate | 6. SEO infrastructure: keyword targeting for family office, investment real estate Munich, M&A advisory Germany",
    "A family office managing HNW investor relationships and M&A transactions in a German market built on trust and track record. A prospective investor or acquisition target does their research before the first call. They land on bronshteyn.com and learn Boris's name, email address, and Telegram handle. No deal history. No service scope. No reason to proceed to the call that the business model depends on. The site fails the credibility filter before the relationship has a chance to start.",
    "Full site rebuild as family office authority: services clearly scoped, team credentials for Boris/Anastasia/Ksenia, deal track record where disclosable, investor-grade design | AI chat trained on service scope, investment focus areas, deal criteria - qualifies inbound before it reaches Boris | Follow-up automation: sequences for investors and sellers from first contact through to mandate, because HNW relationships close on the sixth touchpoint | Content layer: quarterly Munich/German real estate commentary and M&A insights for search authority and positioning | Structured inquiry form routing by deal type: acquisition, co-investment, advisory | SEO rebuild: keyword infrastructure for family office investment real estate Germany and Munich M&A advisory",
    "Month 1: Full site rebuild live on bronshteyn.com - services scoped, team presented, investor-grade design, contact flow restructured with deal-type routing; AI chat deployed and trained on Bronshteyn service scope and deal criteria | Month 2: Investor and seller follow-up sequences active; content layer launched with first German real estate market commentary pieces; SEO infrastructure in place with keyword targeting for family office and M&A advisory Munich | Month 3: Site ranking for target keywords in German real estate investment and Munich family office; inbound inquiry volume measurable and pre-qualified before reaching Boris; content cadence established for quarterly market commentary",
    "Reply to this email or book a 15-min call | Rohan Malik, CEO | hello@genosai.tech | www.genosai.tech"
]
row4 = 101
for col, v in enumerate(ppt_vals, 1):
    cell = ws4.cell(row4, col)
    cell.value = v
    copy_style(ws4.cell(row4 - 1, col), cell)

# Organic Leads
ws5 = wb["Organic Leads"]
thin = Side(style="thin", color="444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
row_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
row_font = Font(name="Calibri", size=10, color="1A1A2E")

organic_vals = [
    11, "Boris", "Bronshteyn", "Boris Bronshteyn", "CEO",
    "Bronshteyn GmbH", "boris@bronshteyn.com",
    "Valid", "https://www.bronshteyn.com/", "",
    "Munich", "Bavaria", "DE", "11-50", "Real Estate - Family Office / Investment Advisory",
    "https://www.linkedin.com/company/bronshteyn/about/", "Organic/Manual",
    "2026-05-10", "Bronshteyn_Audit_GenosAI.pptx", "Email Sent - Pending Reply"
]
for col, v in enumerate(organic_vals, 1):
    cell = ws5.cell(12, col)
    cell.value = v
    cell.font = row_font
    cell.fill = row_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border

for col in range(1, ws5.max_column + 1):
    if ws5.cell(1, col).value == "Cold Email Body":
        cell = ws5.cell(12, col)
        cell.value = body
        cell.font = row_font
        cell.fill = row_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        break

ws5.row_dimensions[12].height = 180

wb.save(WB_DIR / "GenosAI_RealEstate_Leads_UPDATED.xlsx")
print("Excel updated - Boris Bronshteyn / Bronshteyn GmbH added to all 5 sheets.")
