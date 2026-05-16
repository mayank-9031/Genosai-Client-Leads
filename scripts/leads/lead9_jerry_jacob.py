from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import copy as cp
import os

ROOT = Path(__file__).parent.parent.parent
WB_DIR = ROOT / "output" / "workbooks"

import glob
files = glob.glob(str(WB_DIR / "GenosAI_RealEstate_Leads*.xlsx"))
src = next((f for f in files if "UPDATED" in f), str(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx"))
print(f"Loading: {src}")
wb = load_workbook(src)

lead = {
    "num": 98,
    "first": "Jerry",
    "last": "Jacob",
    "full": "Jerry Jacob",
    "title": "Director / Partner",
    "company": "YPA Estate Agents - Tarneit / Truganina / Williams Landing",
    "email": "jerry.jacob@ypa.com.au",
    "email_status": "Verify",
    "website": "https://www.ypa.com.au/",
    "phone": "0404 727 772",
    "city": "Tarneit",
    "state": "VIC",
    "country": "AU",
    "employees": "11-50",
    "industry": "Real Estate - Residential / Property Mgmt",
    "linkedin": "https://www.linkedin.com/in/jerry-jacob-7613b068/",
    "ppt": "YPATarneit_Audit_GenosAI.pptx"
}

subject = "Top of 3029, Truganina sale $190K above median, and an office page that doesn't show it"

body = """\
Hi Jerry,

Top-performing team in the 3029 postcode. Recent Truganina sale at $190K above median after previous agents failed to find a buyer. One Tarneit campaign closed in a week against a market averaging 50 days. The track record you've built with Jalpa across Tarneit-Truganina is the kind that builds itself.

I'm Rohan Malik, CEO of Genos AI. We help offices like yours turn that sort of result into more inbound vendor conversations through AI tools and a website that does something at 11pm when a Tarneit homeowner is comparing agents.

Ran a free audit of ypa.com.au with focus on the Tarneit-Truganina office and attached it. A few things stood out.

A vendor in Tarneit thinking about listing in spring searches at 9pm and lands on a YPA office page that's mostly a roster. No instant valuation tool, no recent sold data by suburb, no story about the Truganina sale that beat the market by $190K. The website asks the vendor to fill out an appraisal request and wait. Most don't, especially when the next agency in their search results gives them a number on the spot.

The Elite Agent feature on Truganina is the kind of social proof that closes vendor decisions. Right now it lives off-site on a third-party publication and not on the office page where it would do the most work.

A few things worth a look:
- Instant property valuation on the office page with suburb-routed follow-up. Vendor enters an address, gets a comparable-based range from YPA's sold data, optionally requests a refined appraisal. Captures the lead instead of asking for it
- AI chat trained on YPA's appraisal process, current Tarneit-Truganina listings, and PM workflows. Handles vendor and buyer questions after hours, books appraisals, triages tenant maintenance
- Office page rebuilt as a 3029 postcode authority. Recent sold data, suburb market reports, the $190K-above-median Truganina sale as a permanent case study, agent matching for buyers
- Vendor nurture sequences with weekly Tarneit-Truganina market reports. Stays in front of vendors in the 6 to 12 month window before they list
- Past-client re-engagement at 1, 3, 5, 7-year anniversaries with current home value and suburb activity. Australians sell every 6-9 years, so the 2018 settled list is the 2026 vendor pipeline
- Tenant portal automation: online applications, maintenance triage, lease renewal sequences 90 days out, behavioural rent reminders

On the website side, we'd rebuild the Tarneit-Truganina page to reflect what your team actually does in 3029. Suburb authority, not just an agent roster. Awards and case studies surfaced where vendors decide.

Audit's attached, no strings. Happy to walk through it in 15 minutes.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi Jerry, just making sure the audit didn't get buried. The instant valuation tool and the office page rebuild are probably the most relevant for Tarneit-Truganina's spring vendor pipeline. Happy to walk through it in 15 minutes. - Rohan\
"""

fu2 = """\
Hi Jerry, last one from me. The Truganina sale at $190K above median is exactly the kind of story that should be doing trust work on the office page 24/7. The case study and past-client re-engagement automation we'd build is what makes that happen. Leaving the audit here in case timing improves. - Rohan\
"""

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = cp.copy(src_cell.font)
        dst_cell.fill = cp.copy(src_cell.fill)
        dst_cell.border = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

# Leads
ws = wb["Leads"]
row = 99
vals = [lead["num"], lead["first"], lead["last"], lead["full"], lead["title"],
        lead["company"], lead["email"], lead["email_status"], lead["website"],
        lead["phone"], lead["city"], lead["state"], lead["country"],
        lead["employees"], lead["industry"], lead["linkedin"], lead["ppt"]]
for col, v in enumerate(vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# Website Audits
ws2 = wb["Website Audits"]
audit_vals = [
    lead["num"], lead["company"], lead["website"],
    "Custom (likely Rex/Designly stack)", "Yes", "Unknown", "Unknown",
    "No", "None Detected", "Unknown", "Unknown", "Yes (suburb reports)", 5, "High",
    "No instant property valuation tool - only an appraisal REQUEST form (vendor must wait); no AI chat for vendor/buyer/tenant after-hours questions; office pages across 24-26 locations are mostly agent rosters, not suburb authorities; Tarneit-Truganina office page lacks recent sold data, suburb reports embedded, or the $190K-above-median Truganina case study; Elite Agent feature on Jerry's Truganina sale lives off-site, not on the YPA office page; no past-client re-engagement automation despite 6-9 year AU resale cycle; no tenant portal visible; no awards or REIWA/REIV-style social proof on homepage; only 10 testimonials surfaced",
    "Instant valuation tool with suburb-routed follow-up; AI chat trained on YPA appraisal process, current listings, and PM workflows for 24/7 vendor/buyer/tenant self-service; office page rebuild as 3029 postcode authority with embedded sold data, market reports, and Elite Agent case study; vendor nurture sequences with weekly Tarneit-Truganina market reports for 6-12 month listing window; past-client re-engagement at 1/3/5/7-year anniversaries; tenant portal automation (applications, maintenance, lease renewals); buyer alerts segmented by suburb and property type; awards/case study wall on every office page"
]
row2 = 99
for col, v in enumerate(audit_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# Cold Emails
ws3 = wb["Cold Emails"]
email_vals = [lead["num"], lead["first"], lead["company"], lead["email"], subject, body, fu1, fu2]
row3 = 99
for col, v in enumerate(email_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

# Audit PPT Content
ws4 = wb["Audit PPT Content"]
ppt_vals = [
    lead["num"], lead["company"], lead["full"], lead["email"], lead["website"], 5, "High",
    "YPA Estate Agents - Tarneit / Truganina / Williams Landing - Website & AI Growth Audit | Prepared by Genos AI",
    "Jerry Jacob: Director / Partner at YPA Tarneit-Truganina-Williams Landing | Co-leads office with Jalpa Patel | Top-performing team in 3029 postcode | Featured in Elite Agent for Truganina sale at $190K above median after previous agents failed | One Tarneit campaign closed in a week vs. 50-day market average | YPA Estate Agents: founded 2007, 24-26 offices across Victoria and QLD, 300+ team members, growth corridors including Tarneit, Truganina, Williams Landing, Point Cook, Sunbury, Werribee, Pakenham, Cranbourne, Mornington Peninsula | Direct contact: 0404 727 772 | Email pattern: jerry.jacob@ypa.com.au",
    "Top performer in Western Melbourne's largest growth corridor (3029 postcode) - the demographic and listing volume tailwind is real; Jerry-Jalpa partnership delivers concrete record results (Truganina $190K above median, week-long campaigns vs. 50-day market); Elite Agent media coverage gives third-party validation most agents can't access; YPA franchise scale (24-26 offices, 300+ staff) provides shared infrastructure and brand recognition; Tarneit-Truganina is one of the fastest-growing residential markets in Australia - rising vendor pool every quarter; PM book + sales operation gives recurring revenue base alongside transactional",
    "No instant valuation tool - YPA's appraisal form is a request that asks the vendor to wait, while comparable agencies offer instant ranges; no AI chat for after-hours vendor/buyer/tenant questions; Tarneit-Truganina office page is mostly an agent roster, not a 3029 suburb authority; the $190K-above-median Truganina sale story is on Elite Agent and not on the YPA office page where vendors decide; no past-client re-engagement automation despite Australian 6-9 year resale cycles; no tenant portal for PM clients; awards and case studies not surfaced on homepage or office pages",
    "1. Instant property valuation tool with suburb-routed follow-up | 2. AI chat for vendor/buyer/tenant 24/7 self-service | 3. Office page rebuild as 3029 postcode authority with embedded sold data + Elite Agent case study | 4. Vendor nurture sequences with weekly Tarneit-Truganina market reports | 5. Past-client re-engagement at 1/3/5/7-year anniversaries | 6. Tenant portal: applications, maintenance, lease renewals, rent reminders | 7. Awards / case study wall surfaced on every primary page",
    "A vendor in Tarneit searches at 9pm, lands on YPA's office page (a paginated agent roster), and either fills out an appraisal request and waits, or moves on to the agency with an instant range. The $190K-above-median Truganina story is exactly the proof point that closes that vendor's decision, and it's sitting on Elite Agent rather than the office page where it would matter. Past clients from settled 2018 sales aren't being re-engaged for the appraisal opportunity in 2026 they represent.",
    "Instant valuation tool on YPA homepage and Tarneit-Truganina office page with suburb-routed follow-up | AI chat trained on YPA's process, current listings, PM workflows | Office page rebuilt as 3029 postcode authority: recent sold data, suburb reports, $190K-above-median Truganina case study, agent matching | Vendor nurture: weekly Tarneit-Truganina market reports for 6-12 month listing window | Past-client re-engagement: 1/3/5/7-year anniversary sequences with current home value | Tenant portal: online applications, maintenance triage, lease renewal sequences 90 days out, behavioural rent reminders | Awards and case study wall on homepage and office page",
    "Month 1: Instant valuation tool live on YPA homepage and Tarneit-Truganina office page, AI chat deployed for vendor/buyer/tenant self-service, office page rebuild begins with sold data and case study integration | Month 2: Vendor nurture sequences active with weekly 3029 market reports, tenant portal live (applications, maintenance, renewals), past-client re-engagement sequences running, awards/case study wall on primary pages | Month 3: Office page acting as 3029 suburb authority with full sold data and matching, buyer alerts segmented by suburb, measurable inbound appraisal requests from valuation tool, vendor pipeline pre-listing measurable",
    "Reply to this email or book a 15-min call | Rohan Malik, CEO | hello@genosai.tech | www.genosai.tech"
]
row4 = 99
for col, v in enumerate(ppt_vals, 1):
    cell = ws4.cell(row4, col)
    cell.value = v
    copy_style(ws4.cell(row4 - 1, col), cell)

# Organic Leads
ws5 = wb["Organic Leads"]
thin = Side(style="thin", color="444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
row_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
row_font = Font(name="Calibri", size=10, color="1A1A2E")

organic_vals = [
    9, "Jerry", "Jacob", "Jerry Jacob", "Director / Partner",
    "YPA Estate Agents - Tarneit / Truganina / Williams Landing", "jerry.jacob@ypa.com.au",
    "Verify", "https://www.ypa.com.au/", "0404 727 772",
    "Tarneit", "VIC", "AU", "11-50", "Real Estate - Residential / Property Mgmt",
    "https://www.linkedin.com/in/jerry-jacob-7613b068/", "Organic/Manual",
    "2026-05-07", "YPATarneit_Audit_GenosAI.pptx", "Email Sent - Pending Reply"
]
for col, v in enumerate(organic_vals, 1):
    cell = ws5.cell(10, col)
    cell.value = v
    cell.font = row_font
    cell.fill = row_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border

for col in range(1, ws5.max_column + 1):
    if ws5.cell(1, col).value == "Cold Email Body":
        cell = ws5.cell(10, col)
        cell.value = body
        cell.font = row_font
        cell.fill = row_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        break

ws5.row_dimensions[10].height = 180

wb.save(WB_DIR / "GenosAI_RealEstate_Leads_UPDATED.xlsx")
print("Excel updated - Jerry Jacob / YPA Tarneit added to all 5 sheets.")
