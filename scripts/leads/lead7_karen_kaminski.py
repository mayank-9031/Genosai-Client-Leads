from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy as cp
import os

ROOT = Path(__file__).parent.parent.parent
WB_DIR = ROOT / "output" / "workbooks"

import glob
files = glob.glob(str(WB_DIR / "GenosAI_RealEstate_Leads*.xlsx"))
src = next((f for f in files if "UPDATED" in f), str(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx"))
print(f"Loading: {src}")
wb = load_workbook(src)

lead = {
    "num": 96,
    "first": "Karen",
    "last": "Kaminski",
    "full": "Karen A. Kaminski",
    "title": "Founder & CEO",
    "company": "Allure Salon Suite Consulting / Allure Salon Group",
    "email": "info@alluresalonsuiteconsulting.com",
    "email_status": "Verify",
    "website": "https://alluresalonsuiteconsulting.com/",
    "phone": "(570) 955-0671",
    "city": "Moosic",
    "state": "PA",
    "country": "US",
    "employees": "2-10",
    "industry": "Beauty / Salon Suite Consulting & Property Mgmt",
    "linkedin": "https://www.linkedin.com/in/karen-a-kaminski-86531228/",
    "ppt": "AllureSalon_Audit_GenosAI.pptx"
}

subject = "Modern Salon Top 10, 5 locations, 115+ studios, and a homepage that doesn't show it"

body = """\
Hi Karen,

Modern Salon nominated your 2024 piece for Top 10 of the year. Authority Magazine, IAMCEO, Salon Scoop, SF Commercial, Elite Property News - all in 2024 and 2025. Five owned salon suite locations, 115+ studios, roughly $150K per location. The track record is hard to match in this space.

I'm Rohan Malik, CEO of Genos AI. We help consultants like you turn that kind of credibility into more inbound investor conversations through AI tools and a website that does something at 11pm when an investor is comparing salon suite options.

Ran a free audit of alluresalonsuiteconsulting.com and attached it. A few things stood out.

The "Partner With Us" page is a phone number and an email address. No form, no calendar booking, no automated follow-up. An investor researching salon suites on a Sunday night reads the page, doesn't call, and moves on to the next consultant in their search results. Same goes for a beauty professional looking at Allure Salon Group for a suite. Two very different audiences, both ending at the same phone tag.

The press wins also don't show up. None of the Modern Salon, podcast, or industry feature content is on the homepage in a way that builds trust before the first conversation. Most consulting buyers do their research before they call.

A few things worth a look:
- Investor ROI calculator on the site so a prospect punching in square footage and metro area instantly sees projected annual revenue. Your $150K-per-location benchmark is exactly the kind of number to anchor it
- AI chat trained on Allure's process, TI negotiation playbook, and pricing approach so the after-hours visitor gets real answers instead of a contact form
- Allure University funnel: landing page, signup sequence, drip lessons, completion tracking. Right now there's no infrastructure for it
- Tenant retention automation for Allure Salon Group's 115+ studios. Lease renewal sequences starting 90 days out, early warning signals before a tenant goes quiet, employee-to-owner program nurture
- Press syndication so the next podcast or article spreads across your channels in minutes instead of hours

On the website side, we'd rebuild it to reflect what 5 locations and 115+ studios actually means. Investor track with case studies and financials, tenant track for Allure Salon Group, a press wall the brand has earned but the homepage doesn't show.

Audit's attached, no strings. Happy to walk through it in 15 minutes.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi Karen, just making sure the audit didn't get buried. The investor ROI calculator and the Allure University funnel are probably the most relevant to where the business is heading. Happy to walk through it in 15 minutes. - Rohan\
"""

fu2 = """\
Hi Karen, last one from me. You said "if your tenants are successful business owners, they stay long term" - the tenant retention automation we'd build is exactly that thesis applied across all 115+ studios at once. Leaving the audit here in case timing improves. - Rohan\
"""

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = cp.copy(src_cell.font)
        dst_cell.fill = cp.copy(src_cell.fill)
        dst_cell.border = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

# Leads
ws = wb["Leads"]
row = 97
vals = [lead["num"], lead["first"], lead["last"], lead["full"], lead["title"],
        lead["company"], lead["email"], lead["email_status"], lead["website"],
        lead["phone"], lead["city"], lead["state"], lead["country"],
        lead["employees"], lead["industry"], lead["linkedin"], lead["ppt"]]
for col, v in enumerate(vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# Website Audits
ws2 = wb["Website Audits"]
audit_vals = [
    lead["num"], lead["company"], lead["website"],
    "WordPress / Custom Theme", "Yes", "Unknown", "None Detected",
    "No", "None Detected", "Unknown", "Unknown", "Unknown", 5, "High",
    "Partner With Us page is a phone number and email - no lead capture form, no calendar booking, no automated follow-up; no AI chat for investor or tenant questions; no investor ROI calculator despite a clear $150K/location benchmark; press wins (Modern Salon Top 10 nominee, Authority Magazine, IAMCEO, SF Commercial, Elite Property News) not surfaced on homepage; no Allure University funnel despite plans to launch the platform; no tenant portal for Allure Salon Group's 115+ studios; two phone numbers listed without unified routing; no investor case studies or location-by-location financials; no clear separation between investor and tenant tracks",
    "Investor ROI calculator (sqft + metro -> projected annual revenue, anchored on $150K benchmark); AI chat trained on Allure's process, TI negotiation, and pricing approach for after-hours investor and tenant questions; Allure University funnel with landing page, signup sequence, drip lessons, completion tracking; tenant retention automation for 115+ studios with 90-day lease renewal sequences and early-warning churn signals; employee-to-owner program nurture sequences; press syndication automation for new media wins; cross-border (US/Canada) investor capture flows; site rebuild with separated investor and tenant tracks plus press wall"
]
row2 = 97
for col, v in enumerate(audit_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# Cold Emails
ws3 = wb["Cold Emails"]
email_vals = [lead["num"], lead["first"], lead["company"], lead["email"], subject, body, fu1, fu2]
row3 = 97
for col, v in enumerate(email_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

# Audit PPT Content
ws4 = wb["Audit PPT Content"]
ppt_vals = [
    lead["num"], lead["company"], lead["full"], lead["email"], lead["website"], 5, "High",
    "Allure Salon Suite Consulting & Allure Salon Group - Website & AI Growth Audit | Prepared by Genos AI",
    "Karen A. Kaminski: Founder & CEO Allure Salon Suite Consulting (2018) and Allure Salon Group | BA Advertising/Graphics/Fine Arts, Marywood University | 15+ years beauty industry (CosmoProf Beauty, East Coast Salon Services, Moroccan Oil) | Opened first salon suite 2016 | 5 owned locations housing 115+ studios | $150K avg annual per location | Saves clients 30-40% on startup/construction | Negotiates $100/sqft+ TI on 10-20 year leases | US + Canada operations | Modern Salon Top 10 Article of 2024 nominee | Featured: Authority Magazine, IAMCEO Podcast, Salon Scoop, SF Commercial Property Conversations, Elite Property News | HQ: 72 Glenmaura National Blvd, Moosic, PA",
    "Founded Allure Salon Suite Consulting in 2018 with rare hands-on credentials - she actually owns and operates 5 salon suite locations rather than just consulting on them; 115+ studios under management proves the model at scale; press footprint is exceptional for an industry niche - Modern Salon Top 10 nominee plus Authority Magazine and multiple podcast features; clear differentiation as 'built by beauty professionals for beauty professionals'; $150K-per-location benchmark and 30-40% startup savings are concrete numbers most consultants can't quote; Marywood University design background means brand sensibility is real; Employee-to-Owner Program and planned Allure University show product thinking beyond a typical consulting practice",
    "Partner With Us page is a phone number and email - no lead capture form, no calendar booking, no automated follow-up; no AI chat for investor or tenant questions; no investor ROI calculator despite the $150K/location benchmark she quotes publicly; press wins not surfaced on homepage where they would build trust pre-call; no Allure University funnel despite planned launch; no tenant portal for Allure Salon Group's 115+ studios; no investor case studies or location-by-location financials; no clear separation between investor and tenant audiences",
    "1. Investor ROI calculator (sqft + metro -> projected annual revenue, anchored on $150K) | 2. AI chat trained on Allure's process, TI negotiation, pricing approach - 24/7 investor and tenant self-service | 3. Allure University funnel - landing page, signup, drip lessons, completion tracking | 4. Tenant retention automation - 90-day lease renewal sequences and churn early-warning across 115+ studios | 5. Employee-to-Owner program drip nurture | 6. Press syndication automation for next Modern Salon or podcast feature | 7. Site rebuild with separated investor / tenant tracks and press wall",
    "An investor researching salon suite consultants on a Sunday night reads the Partner With Us page, doesn't call, and moves on. A beauty professional comparing Allure Salon Group vs. a franchise hits the same dead end. The press wins don't appear on the homepage where they would close the trust gap before the first call. Allure University has no infrastructure to capture the audience it would build. Tenant retention - the very thing Karen says drives long-term success - is run manually across 115+ studios.",
    "Investor lead capture: ROI calculator + qualified-lead form on Partner With Us, automated follow-up sequence | AI chat: trained on Allure's process, TI negotiation, pricing - handles investor and tenant questions 24/7 | Allure University: landing page, signup funnel, drip course delivery, completion tracking | Tenant retention automation: 90-day pre-renewal sequences, early-warning churn signals, employee-to-owner drip nurture | Press wall: homepage section + automated syndication when new features land | Site rebuild: investor track (case studies, financials, ROI) + tenant track (Allure Salon Group studios) + press wall + Allure University",
    "Month 1: AI chat live on alluresalonsuiteconsulting.com, investor ROI calculator deployed on Partner With Us page, lead capture form replacing phone-and-email-only contact, press wall added to homepage | Month 2: Allure University funnel live with first drip course, tenant retention sequences running for Allure Salon Group's 115+ studios, employee-to-owner nurture sequences active | Month 3: Site rebuild complete with investor and tenant tracks, press syndication automation in production, first cohort of inbound investor leads measurable, cross-border (US/Canada) capture flows live",
    "Reply to this email or book a 15-min call | Rohan Malik, CEO | hello@genosai.tech | www.genosai.tech"
]
row4 = 97
for col, v in enumerate(ppt_vals, 1):
    cell = ws4.cell(row4, col)
    cell.value = v
    copy_style(ws4.cell(row4 - 1, col), cell)

# Organic Leads
ws5 = wb["Organic Leads"]
thin = Side(style="thin", color="444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
row_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
row_font = Font(name="Calibri", size=10, color="1A1A2E")

organic_vals = [
    7, "Karen", "Kaminski", "Karen A. Kaminski", "Founder & CEO",
    "Allure Salon Suite Consulting / Allure Salon Group", "info@alluresalonsuiteconsulting.com",
    "Verify", "https://alluresalonsuiteconsulting.com/", "(570) 955-0671",
    "Moosic", "PA", "US", "2-10", "Beauty / Salon Suite Consulting & Property Mgmt",
    "https://www.linkedin.com/in/karen-a-kaminski-86531228/", "Organic/Manual",
    "2026-05-07", "AllureSalon_Audit_GenosAI.pptx", "Email Sent - Pending Reply"
]
for col, v in enumerate(organic_vals, 1):
    cell = ws5.cell(8, col)
    cell.value = v
    cell.font = row_font
    cell.fill = row_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border

for col in range(1, ws5.max_column + 1):
    if ws5.cell(1, col).value == "Cold Email Body":
        cell = ws5.cell(8, col)
        cell.value = body
        cell.font = row_font
        cell.fill = row_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        break

ws5.row_dimensions[8].height = 180

wb.save(WB_DIR / "GenosAI_RealEstate_Leads_UPDATED.xlsx")
print("Excel updated - Karen Kaminski added to all 5 sheets.")
