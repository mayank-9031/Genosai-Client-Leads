from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import copy as cp
import os

ROOT = Path(__file__).parent.parent.parent
WB_DIR = ROOT / "output" / "workbooks"

import glob
files = glob.glob(str(WB_DIR / "GenosAI_RealEstate_Leads*.xlsx"))
src = next((f for f in files if "UPDATED" in f), str(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx"))
print(f"Loading: {src}")
wb = load_workbook(src)

lead = {
    "num": 101,
    "first": "Kevin",
    "last": "Levine",
    "full": "Kevin Levine",
    "title": "Executive Vice President & Principal",
    "company": "Peak 1031 Exchange",
    "email": "kevinl@peakexchange.com",
    "email_status": "Valid",
    "website": "https://peakexchange.com/",
    "phone": "866-357-1031",
    "city": "Woodland Hills",
    "state": "CA",
    "country": "US",
    "employees": "11-50",
    "industry": "Real Estate - 1031 Exchange / Qualified Intermediary",
    "linkedin": "https://www.linkedin.com/company/peak-1031-exchange/",
    "ppt": "PeakExchange_Audit_GenosAI.pptx"
}

subject = "9 people, national 1031 volume, and a 45-day deadline that costs clients six figures if one reminder slips"

body = """\
Hi Kevin,

20 years building Peak 1031 into a nationally operating qualified intermediary inside the Peak Corporate Network. The FEA membership, the 4.9-star rating across 76 reviews, and the multi-state operation at 9 people is a lean, precise operation. The margin for error in your business is close to zero.

I'm Rohan Malik, CEO of Genos AI. We help firms like yours automate the deadline-critical and document-heavy workflows that grow faster than headcount can.

Ran a quick operational audit of Peak 1031's process and attached it. A few things stood out.

A national 1031 operation at 9 people means every team member carries significant per-transaction load. The 45-day identification window and 180-day exchange window are non-negotiable. A client whose identification letter arrives at Day 46 instead of Day 45 owes the IRS the full capital gains on their sale - potentially $80K, $200K, $500K. That liability doesn't fall on Peak. But the client relationship does.

The workflows running that deadline risk today are calendar entries, spreadsheet trackers, and email chains between your team, the investor, the title company, and the accountant. Every exchange has 20-plus documents. Every one of those documents has to be chased, classified, verified, and matched to the right file. That's not where your team's expertise is most valuable - and it's also where the errors happen.

A few things worth a look:
- Automated 45/180-day deadline tracking with tiered alerts to team and client at 15, 7, 3, and 1 day out. Escalates to senior staff if an action is still pending within 48 hours of a critical date. Zero calendar management required
- AI document collection and classification. Client submits documents, the system classifies by type, extracts key data by OCR, flags missing items, and alerts your team and the client with a specific list of what's outstanding. Document assembly time from 5-10 hours per exchange to under 1 hour
- Property identification verification bot. Client submits an address, the system cross-references parcel data, legal description, and county records, pre-fills the identification letter with IRS-compliant formatting, and flags any discrepancy before the letter goes out
- Client status portal with AI chat. Investors can check where their exchange stands without calling your office. The chatbot handles timeline questions, document status, and IRS rule FAQs. Your team gets involved only when a decision or exception actually requires them
- Compliance checklist automation by exchange type. Delayed, simultaneous, reverse, improvement - each has a different documentation and verification path. The system generates the right checklist for each exchange, tracks completion, and flags gaps before senior staff review

None of this replaces your team. It removes the 60 percent of each exchange that doesn't need their expertise, so the remaining 40 percent gets handled at higher quality and higher volume.

The operational audit is attached, no strings. Happy to walk through it in 15 minutes.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi Kevin, just making sure the audit didn't get buried. The deadline tracking automation and document classification are probably the fastest to deploy and the most direct on liability exposure. Happy to walk through it in 15 minutes. - Rohan\
"""

fu2 = """\
Hi Kevin, last one from me. The 9-to-3x transaction volume gap for a national 1031 operation closes when the manual overhead per exchange goes from 10 hours to 2. The automation stack we'd build is what makes that math work without proportional hiring. Leaving the audit here in case timing improves. - Rohan\
"""

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = cp.copy(src_cell.font)
        dst_cell.fill = cp.copy(src_cell.fill)
        dst_cell.border = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

# Leads
ws = wb["Leads"]
row = 102
vals = [lead["num"], lead["first"], lead["last"], lead["full"], lead["title"],
        lead["company"], lead["email"], lead["email_status"], lead["website"],
        lead["phone"], lead["city"], lead["state"], lead["country"],
        lead["employees"], lead["industry"], lead["linkedin"], lead["ppt"]]
for col, v in enumerate(vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# Website Audits
ws2 = wb["Website Audits"]
audit_vals = [
    lead["num"], lead["company"], lead["website"],
    "Custom / WordPress (peakexchange.com)", "Yes", "Unknown", "No",
    "No", "None Detected", "Yes (4.9 stars, 76 reviews)", "Yes (FEA member)", "Yes (blog)", 6, "High",
    "Website is professionally designed and functional; primary AI automation gaps are operational not digital: no automated 45/180-day deadline alert system (manual calendar/spreadsheet tracking creates liability risk); no AI document collection and classification (20+ docs per exchange manually chased and verified); no property identification verification bot (manual cross-reference of parcel data and legal descriptions); no client self-service portal or AI chat for exchange status and IRS rule FAQs; no compliance checklist automation by exchange type; no IRS Form 8824 auto-generation; no fund movement tracking automation; manual workflows cap transaction volume at current 9-person headcount ceiling",
    "Automated 45/180-day deadline tracking with tiered alerts (15/7/3/1 day) to team and client with senior staff escalation; AI document collection and classification with OCR data extraction and missing-document alerts; property identification verification bot cross-referencing parcel data and generating IRS-compliant identification letters; client status portal with AI chat for timeline questions, document status, and IRS FAQ deflection; compliance checklist automation by exchange type (delayed, simultaneous, reverse, improvement); IRS Form 8824 auto-generation from exchange file data; fund movement tracking and escrow reconciliation automation; investor communication orchestration at key exchange milestones"
]
row2 = 102
for col, v in enumerate(audit_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# Cold Emails
ws3 = wb["Cold Emails"]
email_vals = [lead["num"], lead["first"], lead["company"], lead["email"], subject, body, fu1, fu2]
row3 = 102
for col, v in enumerate(email_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

# Audit PPT Content
ws4 = wb["Audit PPT Content"]
ppt_vals = [
    lead["num"], lead["company"], lead["full"], lead["email"], lead["website"], 6, "High",
    "Peak 1031 Exchange - AI Automation Opportunity Audit | Prepared by Genos AI",
    "Kevin Levine: Executive Vice President & Principal / Co-Founder of Peak 1031 Exchange | Founded 2003 | Part of Peak Corporate Network (est. 1991) | Woodland Hills, California | National qualified intermediary (all 50 states) | 9+ team members | 4.9-star rating, 76+ reviews | FEA member | Exchange types: delayed, simultaneous, reverse, improvement, personal property | Team includes Steven Rosansky Esq. (Senior Director), Joshua Messian (Manager/Senior Exchange Officer), Jay Peet (VP Business Development), Eli Tene (Managing Director, Peak Entities) | Process: IRS Section 1031, 45-day identification window, 180-day exchange window, 20+ documents per exchange, escrow fund management, IRS Form 8824 filing",
    "20-year track record in national 1031 exchange operations - rare institutional depth; FEA membership signals compliance commitment and industry standing; 4.9-star rating across 76 reviews reflects genuine client satisfaction and outcome delivery; multi-exchange type capability (delayed through improvement) means larger addressable market than single-product QIs; Peak Corporate Network structure provides cross-referral from mortgage, realty, and property management operations; lean 9-person team delivering national volume demonstrates high per-person productivity that automation can multiply; deadline-critical business model creates clear and immediate ROI case for automation investment",
    "No automated 45/180-day deadline alert system - critical deadline tracking runs on manual calendars and spreadsheets, creating client liability exposure on every active exchange; no AI document collection or classification - 20+ documents per exchange chased and organised manually at 5-10 hours per file; no property identification verification bot - legal description cross-referencing done manually against county records and MLS; no client self-service portal or AI chat - investors call or email for exchange status updates and IRS timeline questions that a portal handles automatically; no compliance checklist automation by exchange type - checklists generated manually per transaction; no IRS Form 8824 auto-generation from exchange data; current manual workflows cap transaction volume at 9-person headcount ceiling",
    "1. Automated 45/180-day deadline tracking with tiered alerts at 15/7/3/1 day - zero missed deadline risk | 2. AI document collection and classification with OCR extraction and missing-document alerts | 3. Property identification verification bot: parcel lookup + IRS-compliant letter generation | 4. Client status portal + AI chat for exchange status and IRS FAQ deflection | 5. Compliance checklist automation by exchange type | 6. IRS Form 8824 auto-generation from exchange file | 7. Fund movement tracking and escrow reconciliation automation",
    "9 people managing national 1031 exchanges where the 45-day identification deadline is non-negotiable and one missed alert costs a client more than Peak's annual fee for that exchange. The deadline tracking is on spreadsheets. The 20-plus documents per exchange are chased manually by email. The compliance checklist is generated fresh each time. The client calling for an update gets whoever picks up the phone. None of that is where Kevin's team's expertise should be spending its time.",
    "Automated deadline tracking: tiered alerts to team and client at 15/7/3/1 day, senior staff escalation if action pending within 48 hours | AI document collection and classification: OCR extraction, missing-document alert, automatic file matching | Property identification verification: parcel data cross-reference, IRS-compliant letter pre-fill, discrepancy flagging before submission | Client portal + AI chat: exchange status self-service, document status, IRS FAQ handling - team involved only for exceptions | Compliance checklist automation by exchange type: delayed/simultaneous/reverse/improvement each generates the correct path | Form 8824 auto-generation from exchange file data | Fund movement and escrow reconciliation automation with bank feed integration",
    "Month 1: Deadline tracking automation live - tiered alerts to team and clients on all active and new exchanges; AI document classification deployed - incoming docs auto-sorted, OCR extracted, missing items flagged; client status portal live with exchange timeline and document checklist visible to investors | Month 2: Property identification verification bot live; compliance checklist automation by exchange type; IRS Form 8824 auto-generation from exchange file; investor communication sequences at key milestones | Month 3: Fund movement tracking and escrow reconciliation automation; full exchange file assembled automatically from close to completion; team handling exceptions only; measurable reduction in manual hours per exchange and measurable increase in transaction throughput",
    "Reply to this email or book a 15-min call | Rohan Malik, CEO | hello@genosai.tech | www.genosai.tech"
]
row4 = 102
for col, v in enumerate(ppt_vals, 1):
    cell = ws4.cell(row4, col)
    cell.value = v
    copy_style(ws4.cell(row4 - 1, col), cell)

# Organic Leads
ws5 = wb["Organic Leads"]
thin = Side(style="thin", color="444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
row_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
row_font = Font(name="Calibri", size=10, color="1A1A2E")

organic_vals = [
    12, "Kevin", "Levine", "Kevin Levine", "Executive Vice President & Principal",
    "Peak 1031 Exchange", "kevinl@peakexchange.com",
    "Valid", "https://peakexchange.com/", "866-357-1031",
    "Woodland Hills", "CA", "US", "11-50", "Real Estate - 1031 Exchange / Qualified Intermediary",
    "https://www.linkedin.com/company/peak-1031-exchange/", "Organic/Manual",
    "2026-05-10", "PeakExchange_Audit_GenosAI.pptx", "Email Sent - Pending Reply"
]
for col, v in enumerate(organic_vals, 1):
    cell = ws5.cell(13, col)
    cell.value = v
    cell.font = row_font
    cell.fill = row_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border

for col in range(1, ws5.max_column + 1):
    if ws5.cell(1, col).value == "Cold Email Body":
        cell = ws5.cell(13, col)
        cell.value = body
        cell.font = row_font
        cell.fill = row_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        break

ws5.row_dimensions[13].height = 180

wb.save(WB_DIR / "GenosAI_RealEstate_Leads_UPDATED.xlsx")
print("Excel updated - Kevin Levine / Peak 1031 Exchange added to all 5 sheets.")
