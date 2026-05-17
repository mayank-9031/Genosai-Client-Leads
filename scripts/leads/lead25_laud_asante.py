from pathlib import Path
import copy as cp
from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent.parent
WB_PATH = ROOT / "output" / "workbooks" / "GenosAI_OrganicLeads.xlsx"

print(f"Loading: {WB_PATH}")
wb = load_workbook(WB_PATH)

lead = {
    "num":   23,
    "first": "Laud",
    "last":  "Asante",
    "full":  "Laud Morgan Asante",
    "title": "CEO",
    "company": "Regnum Properties",
    "email": "info@theregnumproperties.com",
    "email_status": "General contact -- personal email not publicly listed",
    "website":  "https://www.theregnumproperties.com",
    "phone":    "+233 50 863 5479",
    "city":     "Accra",
    "state":    "Greater Accra, Ghana",
    "industry": "Real Estate / Property Development",
    "linkedin": "https://www.linkedin.com/company/regnum-properties/",
    "ppt":   "RegnumProperties_Pitch_GenosAI.pptx",
    "score": 6.5,
    "automation": (
        "Lead qualification AI via WhatsApp (capture budget, property type, timeline, "
        "location preference before agent involvement -- critical for diaspora/international buyers "
        "enquiring across time zones); "
        "investor nurture sequences (automated email/WhatsApp drip for overseas buyers: "
        "construction progress updates, payment plan details, market context, viewing booking); "
        "social media and content AI (LinkedIn and Instagram content generation -- "
        "market insights, project updates, buyer guides -- to grow from 2 followers organically); "
        "AI property inquiry chatbot on website with 24/7 availability for diaspora buyers; "
        "CRM pipeline automation (automated deal stage tracking, follow-up reminders, "
        "buyer communication history -- replacing manual WhatsApp/email tracking)"
    ),
    "notes": (
        "Founded 2023. East Legon, Accra, Ghana. CEO: Laud Morgan Asante "
        "(BA Political Science/Sociology, U of Ghana; MA International Relations, "
        "Webster University St. Louis). Finance Director: Maxwell Addo (48 yrs experience). "
        "Active projects: The Alexandria (4.5BR smart townhouses, Trasacco Phase 1 area) and "
        "The Arlington (3.5BR, 5-unit gated community). Explicitly targets diaspora/international buyers. "
        "Smart home features in all units. No CRM, no WhatsApp button, no virtual tour, "
        "no booking system. Website on SitePad/Estatik (placeholder content still live on About page). "
        "LinkedIn: 2 followers. Phone: +233 50 863 5479 / +233 (0)302 547 951. "
        "Ice-breaker: CEO's MA International Relations background shaped the diaspora buyer thesis -- "
        "bridging Ghanaian diaspora investors to luxury Accra developments."
    ),
}

subject = "The Arlington, 5 units, diaspora buyers -- the automation layer that converts overseas interest before it goes cold"

body = """\
Hi Laud,

Regnum Properties is built around a specific thesis: luxury residential developments in East Legon sold to international buyers and the Ghanaian diaspora. That thesis is commercially strong -- diaspora demand for Accra real estate is real and growing. The challenge it creates is operational: a buyer in London or Atlanta sees The Arlington at 10 PM their time, fills in a contact form, and by the time someone picks it up Monday morning, the moment is gone.

The gap between international interest and signed deals is almost always a follow-up problem. Not a product problem.

What Genos AI builds for real estate developers at Regnum's stage:

WhatsApp lead qualification AI -- a WhatsApp number displayed on the Regnum website and social profiles handles inbound enquiries automatically: budget, property type preference, timeline, and location interest captured before any agent is involved. A diaspora buyer in Toronto can qualify themselves at midnight and wake up to a confirmed viewing booking.

Investor nurture sequences -- buyers who enquire but do not convert immediately enter an automated sequence: construction progress photos, payment plan breakdowns, Accra market context, testimonials. The sequence runs for 12 weeks without manual follow-up. Most luxury property deals in Ghana close on the third or fourth contact -- automation ensures those contacts happen.

Social media content AI -- Regnum is actively posting on LinkedIn and Instagram, which is exactly right. The gap is consistency and volume. An AI layer generates market insight posts, project update content, and diaspora buyer guides on a weekly schedule -- growing an audience without consuming founder time.

AI property chatbot -- a 24/7 chat widget on theregnumproperties.com that answers property questions, shares floor plans, and books viewings. For a developer selling smart homes, operating on a basic contact form is a credibility gap.

CRM pipeline automation -- as The Arlington closes and the next project launches, managing buyer conversations across WhatsApp threads and emails stops scaling. Automated pipeline tracking, follow-up reminders, and communication history in one view means Laud and Maxwell always know exactly where every buyer stands.

I'm Rohan Malik, CEO of Genos AI. We build AI automation for real estate developers. Happy to show you how it maps to Regnum's sales model in 20 minutes.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi Laud, following up on my last message. The fastest win for Regnum -- a WhatsApp AI that qualifies diaspora buyers the moment they enquire, regardless of time zone -- is typically live in under 2 weeks and directly captures the leads that currently go cold over the weekend. Happy to walk through it in 20 minutes. -- Rohan | Genos AI\
"""

fu2 = """\
Laud, last one from me. A developer targeting international buyers with no WhatsApp qualification, no nurture sequences, and no 24/7 inquiry capture is leaving deals on the table every week. That is exactly what we fix. Leaving this here if the timing is not right. -- Rohan | Genos AI\
"""


def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font      = cp.copy(src_cell.font)
        dst_cell.fill      = cp.copy(src_cell.fill)
        dst_cell.border    = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


# -- Organic Leads -------------------------------------------------------------
ws  = wb["Organic Leads"]
row = ws.max_row + 1
organic_vals = [
    lead["num"], lead["company"], lead["full"], lead["title"],
    lead["email"], lead["phone"], lead["website"], lead["linkedin"],
    lead["city"], lead["state"], lead["industry"],
    "Organic/Manual", "Pending Send", lead["score"],
    lead["ppt"], lead["automation"], lead["notes"],
]
for col, v in enumerate(organic_vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# -- Cold Emails ---------------------------------------------------------------
ws2  = wb["Cold Emails"]
row2 = ws2.max_row + 1
email_vals = [
    lead["num"], lead["company"], lead["full"],
    lead["email"], subject, body, fu1, fu2,
]
for col, v in enumerate(email_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# -- Pipeline ------------------------------------------------------------------
ws3  = wb["Pipeline"]
row3 = ws3.max_row + 1
pipeline_vals = [
    lead["num"], lead["company"], lead["full"],
    lead["email"], lead["industry"], lead["score"],
    "Email Pending", "2026-05-17", "Send cold email -- pitch AI automation for diaspora real estate sales",
    "",
]
for col, v in enumerate(pipeline_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

wb.save(WB_PATH)
print(f"OrganicLeads.xlsx updated -- Laud Morgan Asante / Regnum Properties added to all 3 sheets (row {row}).")
