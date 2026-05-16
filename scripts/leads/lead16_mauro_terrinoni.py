from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import copy as cp
import os

ROOT = Path(__file__).parent.parent.parent
WB_DIR = ROOT / "output" / "workbooks"

import glob
files = glob.glob(str(WB_DIR / "GenosAI_RealEstate_Leads*.xlsx"))
src = next((f for f in files if "UPDATED" in f), str(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx"))
print(f"Loading: {src}")
wb = load_workbook(src)

lead = {
    "num": 105,
    "first": "Mauro",
    "last": "Terrinoni",
    "full": "Mauro Terrinoni",
    "title": "CEO & Co-Founder",
    "company": "Enerinvest / Crossroad",
    "email": "mterrinoni@enerinvest.com",
    "email_status": "Verified (listed on About page)",
    "website": "https://crossroad.app/",
    "phone": "",
    "city": "London",
    "state": "",
    "country": "UK",
    "employees": "2-10",
    "industry": "PropTech / Property Management SaaS",
    "linkedin": "https://uk.linkedin.com/in/mauroterrinoni",
    "ppt": "Crossroad_Audit_GenosAI.pptx"
}

subject = "Crossroad sells digital transformation to property managers - demo request is still an email"

body = """\
Hi Mauro,

Read your piece "Property Management enters the digital era" and the follow-up on what PropTech can do for property owners. The argument is right - fragmented spreadsheets, disconnected service providers, and manual reporting are exactly where digitisation pays.

One thing caught my eye on crossroad.app. The headline CTA is "Go Digital with Crossroad Today." The mechanism to act on that is: send an email to info@enerinvest.com.

A prospect who wants to see a demo lands on a well-designed site, reads about centralised asset data and automated workflows, and then gets an email inbox as the response mechanism. No chatbot to answer the qualifying questions they have right now. No automated demo booking. No free trial. No nurture sequence if they don't follow through immediately.

For a platform built to eliminate manual workstreams in property management, the sales process into Crossroad is still entirely manual.

The irony is the pitch. If Crossroad automates what property managers do manually, an AI-powered sales funnel for Crossroad itself demonstrates the product philosophy before the demo even starts.

I'm Rohan Malik, CEO of Genos AI. We build AI agents and sales automation for B2B SaaS and service businesses - specifically the demo booking flows, qualification chatbots, and post-visit nurture sequences that convert site traffic without adding headcount. Did a quick audit of crossroad.app and found a few specific things worth 15 minutes.

Happy to send it across or talk through it whenever works.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi Mauro, following up on my last note. The specific point - a PropTech platform with a manual-email demo CTA - is the clearest version of the gap. An AI chatbot that qualifies prospects and books demos automatically is a 2-week build. Happy to walk through the audit in 15 minutes. - Rohan | Genos AI\
"""

fu2 = """\
Mauro, last one from me. The Crossroad audit is ready whenever the timing is right. The case for automating your own sales funnel using the same logic you apply to property management is in the deck. Leaving it here if it's useful. - Rohan | Genos AI\
"""

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = cp.copy(src_cell.font)
        dst_cell.fill = cp.copy(src_cell.fill)
        dst_cell.border = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

# Leads
ws = wb["Leads"]
row = 106
vals = [lead["num"], lead["first"], lead["last"], lead["full"], lead["title"],
        lead["company"], lead["email"], lead["email_status"], lead["website"],
        lead["phone"], lead["city"], lead["state"], lead["country"],
        lead["employees"], lead["industry"], lead["linkedin"], lead["ppt"]]
for col, v in enumerate(vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# Website Audits
ws2 = wb["Website Audits"]
audit_vals = [
    lead["num"], lead["company"], lead["website"],
    "Custom / unknown", "Yes", "No (email link only)", "No",
    "No", "No", "No (no case studies, no client logos)", "No", "No", 5, "High",
    "Crossroad sells digital transformation but its own customer acquisition is entirely manual: primary CTA 'Go Digital with Crossroad Today' resolves to an email to info@enerinvest.com - no demo booking automation, no chatbot; no qualifying intake form - prospect interest captured with no context (portfolio size, property type, geography, pain point); no free trial or freemium tier to reduce friction for inbound evaluators; no post-visit nurture sequence - prospect who visits and does not email immediately disappears; no pricing page - B2B SaaS without pricing transparency lengthens sales cycles; no case studies or client logos - 8 years of proptech work with no social proof on site; no blog or SEO content - 'What PropTech can do for you' published on LinkedIn, not indexed as site content; no ROI calculator for property managers to self-qualify",
    "Deploy AI chatbot trained on Crossroad features (property management, sales management, CRM, document management) - qualifies visitors by portfolio size, property type, current pain, and books demo automatically; replace email CTA with automated demo booking flow (Calendly-equivalent with pre-qualification); build post-visit nurture sequence (Day 1/3/7) re-engaging visitors who don't immediately book; add pricing page or transparent tiering (even 'contact for enterprise' with a clear starter tier reduces friction); import Mauro's LinkedIn articles as SEO-indexed blog content - 'Property Management enters the digital era' and 'What PropTech can do for you' are strong top-of-funnel pieces; build ROI calculator (hours saved per property manager per month) as lead capture tool; add case studies / client testimonials with portfolio size and outcome metrics"
]
row2 = 106
for col, v in enumerate(audit_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# Cold Emails
ws3 = wb["Cold Emails"]
email_vals = [lead["num"], lead["first"], lead["company"], lead["email"], subject, body, fu1, fu2]
row3 = 106
for col, v in enumerate(email_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

# Audit PPT Content
ws4 = wb["Audit PPT Content"]
ppt_vals = [
    lead["num"], lead["company"], lead["full"], lead["email"], lead["website"], 5, "High",
    "Crossroad / Enerinvest - Website & Sales Automation Audit | Prepared by Genos AI",
    "Mauro Terrinoni: CEO & Co-Founder of Enerinvest PropTech / Crossroad | Location: London, UK | Email: mterrinoni@enerinvest.com | Credentials: CFA charter holder (October 2004), MBA Cornell Johnson Graduate School | Career: JPMorgan, Citi, Fortress Investment Group, ACP Capital - investment banking, private equity, distressed credit portfolios for European commercial banks and credit funds | 5K LinkedIn followers | Published: 'What PropTech can do for you' (Jan 2021), 'Property Management enters the digital era' (Nov 2020) | Languages: English, Italian (native), Spanish | Co-founder: Alex Coutlis (acoutlis@enerinvest.com) | Enerinvest PropTech: founded 2020, London, 2-10 employees, 8 years proptech innovation | Crossroad: B2B property management SaaS - asset management, project management, document management, CRM (buyers/tenants/borrowers), branded marketing website, collaborative workspace for internal/external partners | Address: 36-37 Albert Embankment, London SE1 7TL | Target clients: organisations managing large property portfolios",
    "Mauro has already done the intellectual work on why property management needs to go digital - two published pieces and 8 years of product development prove conviction; CFA + Cornell MBA + JPMorgan/Citi/Fortress background means he evaluates ROI arguments rigorously - the pitch must be specific and data-led; 50 years of combined founder experience executing trades valued in hundreds of millions across European markets gives institutional credibility; the irony angle is strong and non-generic: a PropTech company selling digital transformation via a manual email CTA is a pointed, specific observation that only someone who audited the site would make; Crossroad solves a real problem (fragmented property data, disconnected teams, manual reporting) - the product logic is sound, the GTM automation is the gap",
    "CTA 'Go Digital with Crossroad Today' resolves to email to info@enerinvest.com - no demo booking automation; no AI chatbot or qualifying intake - prospect interest captured with zero context; no free trial or self-serve option for inbound evaluators; no post-visit nurture sequence - non-converting visitors disappear permanently; no pricing page - increases sales cycle length for B2B SaaS; no case studies or client logos despite 8 years in market; Mauro's two LinkedIn articles not indexed as site blog content - SEO value lost; no ROI calculator for self-qualification",
    "1. AI chatbot qualifying by portfolio size, property type, pain point - books demo automatically | 2. Automated demo booking flow replacing email CTA | 3. Post-visit Day 1/3/7 nurture sequence for non-converting visitors | 4. Pricing page or transparent starter tier to reduce B2B sales friction | 5. Import LinkedIn articles as SEO-indexed blog (existing content, zero additional writing) | 6. ROI calculator as lead magnet with email gate",
    "Mauro published 'Property Management enters the digital era' in 2020. Six years later, the process for entering Crossroad's sales funnel is: send an email and wait. The same logic he applies to property managers - eliminate manual workstreams, centralise data, automate workflows - applies to how Crossroad itself acquires customers. Genos AI builds the automated sales layer that makes the product philosophy visible before the first demo.",
    "Website: replace email CTA with automated demo booking + AI chatbot; add pricing page; import LinkedIn articles as blog | AI automation: qualifying chatbot, post-visit nurture sequence, ROI calculator lead magnet | SEO: Mauro's existing articles indexed; property management guides targeting UK and European portfolio manager keywords | Implementation: chatbot + demo automation first (fastest pipeline impact), nurture sequence second, SEO and ROI calculator third",
    "Week 1: AI chatbot deployed trained on Crossroad features and property management use cases; automated demo booking replaces info@ email CTA; post-visit Day 1/3/7 nurture sequence live | Week 2: Mauro's two LinkedIn articles imported and published as indexed blog posts; pricing page (or transparent starter tier) live; ROI calculator with email gate published | Month 2: first 10 property management SEO articles published targeting UK, Italian, Spanish market keywords; case study template built and first 2 client stories published with portfolio size and outcome metrics | Month 3: chatbot qualifying data informing ICP refinement; nurture sequence A/B tested; full content calendar for European PropTech market planned",
    "Reply to this email or book a 15-min call | Rohan Malik, CEO | hello@genosai.tech | www.genosai.tech"
]
row4 = 106
for col, v in enumerate(ppt_vals, 1):
    cell = ws4.cell(row4, col)
    cell.value = v
    copy_style(ws4.cell(row4 - 1, col), cell)

# Organic Leads
ws5 = wb["Organic Leads"]
thin = Side(style="thin", color="444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
row_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
row_font = Font(name="Calibri", size=10, color="1A1A2E")

organic_vals = [
    16, "Mauro", "Terrinoni", "Mauro Terrinoni", "CEO & Co-Founder",
    "Enerinvest / Crossroad", "mterrinoni@enerinvest.com",
    "Verified (About page)", "https://crossroad.app/",
    "",
    "London", "", "UK", "2-10", "PropTech / Property Management SaaS",
    "https://uk.linkedin.com/in/mauroterrinoni", "Organic/Manual",
    "2026-05-11", "Crossroad_Audit_GenosAI.pptx", "Email Sent - Pending Reply"
]
for col, v in enumerate(organic_vals, 1):
    cell = ws5.cell(17, col)
    cell.value = v
    cell.font = row_font
    cell.fill = row_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border

for col in range(1, ws5.max_column + 1):
    if ws5.cell(1, col).value == "Cold Email Body":
        cell = ws5.cell(17, col)
        cell.value = body
        cell.font = row_font
        cell.fill = row_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        break

ws5.row_dimensions[17].height = 180

wb.save(WB_DIR / "GenosAI_RealEstate_Leads_UPDATED.xlsx")
print("Excel updated - Mauro Terrinoni / Crossroad added to all 5 sheets.")
