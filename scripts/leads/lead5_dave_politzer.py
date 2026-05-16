from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy as cp
import os

ROOT = Path(__file__).parent.parent.parent
WB_DIR = ROOT / "output" / "workbooks"

import glob
files = glob.glob(str(WB_DIR / "GenosAI_RealEstate_Leads*.xlsx"))
src = next((f for f in files if "UPDATED" in f), str(WB_DIR / "GenosAI_RealEstate_Leads (2).xlsx"))
print(f"Loading: {src}")
wb = load_workbook(src)

lead = {
    "num": 94,
    "first": "Dave",
    "last": "Politzer",
    "full": "Dave Politzer",
    "title": "REALTOR / Founder",
    "company": "The First Responders (Compass)",
    "email": "dave@davepolitzer.com",
    "email_status": "Valid",
    "website": "https://www.thefirstresponders.org/",
    "phone": "(925) 890-9188",
    "city": "Danville",
    "state": "CA",
    "country": "US",
    "employees": "1",
    "industry": "Real Estate - Residential / First Responders",
    "linkedin": "https://www.linkedin.com/in/dave-politzer-376861a/",
    "ppt": "FirstResponders_Audit_GenosAI.pptx"
}

subject = "You built Interloan.com in 1999 - free First Responders audit attached"

body = """\
Hi Dave,

You were one of the first agents to take mortgage lending online in 1999. Most of your peers thought the internet was a fad. Twenty-six years later, the early-mover instinct is still the right one.

I'm Rohan Malik, CEO of Genos AI. We help real estate agents with a specific niche use AI to actually serve that niche instead of just marketing to it.

Ran a free audit of thefirstresponders.org and attached it. A few things stood out.

First responders work shifts. The buyer who's thinking about a home is browsing at 2am after a night shift, not at 10am when most agents are available. There's no AI chat on the site to catch that conversation, no automated follow-up when someone fills out a contact form at midnight, and no way to stay top of mind with the 100+ first responders who probably found the site and never heard back.

The fact that you spent 21 years in Oakland PD is the thing that earns trust with a firefighter or paramedic that no other agent can fake. The digital side should match that credential.

A few things worth a look:
- AI chat trained on first responder home buying programs - captures the 2am browser who can't call during business hours
- Lead follow-up sequences for first responder buyers - automated, personal, built around their schedule not a 9-5
- Automated market reports to your first responder database, keeping you top of mind before they're ready to buy
- Post-closing review automation - first responders trust their colleagues. One review from a fellow officer carries more weight than five from strangers
- Program landing pages that explain the first responder benefits clearly and convert visitors to appointments

Audit's attached, no strings. Happy to walk through it in 15 minutes.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi Dave, just checking the audit didn't get buried. The AI chat and lead follow-up sections are probably most relevant given the after-hours nature of your buyer pool. Happy to walk through it in 15 minutes. - Rohan\
"""

fu2 = """\
Hi Dave, last one from me. 21 years in Oakland PD and 40+ years in real estate is a serious combination of credentials - the digital side should match. Leaving the audit here in case timing is better later. - Rohan\
"""

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = cp.copy(src_cell.font)
        dst_cell.fill = cp.copy(src_cell.fill)
        dst_cell.border = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

# ── 1. Leads ──
ws = wb["Leads"]
row = 95
vals = [lead["num"], lead["first"], lead["last"], lead["full"], lead["title"],
        lead["company"], lead["email"], lead["email_status"], lead["website"],
        lead["phone"], lead["city"], lead["state"], lead["country"],
        lead["employees"], lead["industry"], lead["linkedin"], lead["ppt"]]
for col, v in enumerate(vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# ── 2. Website Audits ──
ws2 = wb["Website Audits"]
audit_vals = [
    lead["num"], lead["company"], lead["website"],
    "Custom / Compass-integrated", "Yes", "Unknown", "Unknown",
    "Yes", "None Detected", "Unknown", "Unknown", "Unknown", 5, "High",
    "No AI chat for after-hours first responder buyers (shift workers browse at 2am post-shift); no automated follow-up for contact form submissions - first responders on 48-hour shifts miss the follow-up window; text testimonials exist (Pleasanton PD, Oakland PD, Sr DA) but no video testimonials or automated review collection system; no automated market report sequences to first responder database; no department-specific landing pages for Police/Fire/EMS/Dispatch with profession-targeted program calculators; no post-closing review automation to compound profession-identified social proof",
    "AI chat trained on first responder home buying programs (24/7 for shift workers); lead follow-up automation built around shift schedules not 9-5; automated market reports to police/fire/EMS database segmented by East Bay location; post-closing review automation prompting profession-identified testimonials; department-specific landing pages (Police/Fire/EMS/Dispatch) with program benefit calculators; referral network automation through department channels"
]
row2 = 95
for col, v in enumerate(audit_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# ── 3. Cold Emails ──
ws3 = wb["Cold Emails"]
email_vals = [lead["num"], lead["first"], lead["company"], lead["email"], subject, body, fu1, fu2]
row3 = 95
for col, v in enumerate(email_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

# ── 4. Audit PPT Content ──
ws4 = wb["Audit PPT Content"]
ppt_vals = [
    lead["num"], lead["company"], lead["full"], lead["email"], lead["website"], 5, "High",
    "The First Responders / Dave Politzer - Website & AI Growth Audit | Prepared by Genos AI",
    "Dave Politzer: 21 years Oakland PD, retired Sergeant & Detective | Started in real estate 1984 while still a cop | Founded Interloan.com 1999 (one of first online multi-lender mortgage companies, sold to public company 2000) | 40+ years combined real estate and mortgage | Compass agent, Danville CA | DRE# 00866434 | 5.0 stars, 14 reviews | Specialties: buyer, listing, relocation, new construction | thefirstresponders.org program serving police/fire/EMS buyers",
    "Authentic first responder credibility that no competitor can replicate (21 years OPD); early tech adopter - built Interloan.com in 1999 before anyone else; 40+ year track record across mortgage and real estate; 5.0 star rating; Compass platform access and tools; trusted by first responder community in East Bay; niche positioning = high referral potential within tight-knit first responder network",
    "No AI chat for after-hours buyers (first responders browse at 2am post-shift); no automated follow-up for contact form submissions; text testimonials exist (Pleasanton PD, Oakland PD) but no video testimonials or automated review collection; no department-specific landing pages by profession (Police/Fire/EMS/Dispatch); no automated market reports to first responder database; no post-closing review automation for profession-identified social proof",
    "1. AI chat trained on first responder home buying programs - 24/7 for shift workers | 2. Lead follow-up automation built around shift schedules not 9-5 | 3. Automated market reports to police/fire/EMS database segmented by East Bay location | 4. Post-closing review automation prompting profession-identified testimonials | 5. Department-specific landing pages (Police/Fire/EMS/Dispatch) with program benefit calculators | 6. Referral network automation through department channels",
    "First responders work shifts. The buyer who's ready is browsing at 2am, not 10am. No AI chat = missed conversations from the exact buyers Dave is positioned to serve. No automated follow-up = leads going cold while the first responder is on a 48-hour shift. Site has good content and real testimonials - the gap is the automation layer underneath.",
    "thefirstresponders.org enhancement: department-specific pages with program calculators (Police/Fire/EMS/Dispatch), video testimonials section | AI chat trained on first responder programs for 24/7 inquiry capture | Lead follow-up sequences built around shift schedules | Automated market report pipeline to Dave's first responder database | Post-closing review automation prompting profession-identified testimonials | Department referral pages for OPD, SFFD, Contra Costa EMS",
    "Month 1: AI chat live on thefirstresponders.org for 24/7 first responder buyer capture, lead follow-up sequences active | Month 2: Department-specific landing pages live (Police/Fire/EMS/Dispatch), market report pipeline automated, review automation collecting profession-identified testimonials | Month 3: Referral sequences running through departments, review velocity and video testimonials compounding, qualified first responder leads measurably up",
    "Reply to this email or book a 15-min call | Rohan Malik, CEO | hello@genosai.tech | www.genosai.tech"
]
row4 = 95
for col, v in enumerate(ppt_vals, 1):
    cell = ws4.cell(row4, col)
    cell.value = v
    copy_style(ws4.cell(row4 - 1, col), cell)

# ── 5. Organic Leads ──
ws5 = wb["Organic Leads"]
thin = Side(style="thin", color="444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
row_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
row_font = Font(name="Calibri", size=10, color="1A1A2E")

organic_vals = [
    5, "Dave", "Politzer", "Dave Politzer", "REALTOR / Founder",
    "The First Responders (Compass)", "dave@davepolitzer.com",
    "Valid", "https://www.thefirstresponders.org/", "(925) 890-9188",
    "Danville", "CA", "US", "1", "Real Estate - Residential / First Responders",
    "https://www.linkedin.com/in/dave-politzer-376861a/", "Organic/Manual",
    "2026-05-06", "FirstResponders_Audit_GenosAI.pptx", "Email Sent - Pending Reply"
]
for col, v in enumerate(organic_vals, 1):
    cell = ws5.cell(6, col)
    cell.value = v
    cell.font = row_font
    cell.fill = row_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = border

for col in range(1, ws5.max_column + 1):
    if ws5.cell(1, col).value == "Cold Email Body":
        cell = ws5.cell(6, col)
        cell.value = body
        cell.font = row_font
        cell.fill = row_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        break

ws5.row_dimensions[6].height = 180

wb.save(WB_DIR / "GenosAI_RealEstate_Leads_UPDATED.xlsx")
print("Excel updated - Dave Politzer added to all 5 sheets.")
