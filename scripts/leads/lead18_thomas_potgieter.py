from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import copy as cp

ROOT = Path(__file__).parent.parent.parent
WB_PATH = ROOT / "output" / "workbooks" / "GenosAI_OrganicLeads.xlsx"

print(f"Loading: {WB_PATH}")
wb = load_workbook(WB_PATH)

lead = {
    "num": 15,
    "first": "Thomas",
    "last": "Potgieter",
    "full": "Thomas Potgieter",
    "title": "Co-Founder & Managing Director",
    "company": "Home Seekers",
    "email": "info@homeseeker.co.za",
    "email_status": "Best guess (no email publicly listed on site)",
    "website": "https://www.homeseeker.co.za/",
    "phone": "+27 012 880 3127",
    "city": "Pretoria",
    "state": "Gauteng, South Africa",
    "industry": "Real Estate",
    "linkedin": "https://www.linkedin.com/in/thomas-potgieter-a4972b96/",
    "ppt": "HomeSeeker_Audit_GenosAI.pptx",
    "score": 5.5,
    "automation": (
        "WhatsApp AI for buyer inquiries (24/7 response + viewing booking); "
        "lead routing from Property24/Private Property to agents; "
        "FICA compliance document collection workflow; "
        "agent recruitment intake automation (agents.homeseeker.co.za); "
        "buyer nurture drip for non-converting inquiries; "
        "post-sale review request sequence"
    ),
    "notes": (
        "Fixed-fee platform — agents keep 100% commission. Founded 2025. "
        "416+ active listings. Ice-breaker: Rode Potgieter 8 deals / R11.44M Oct 2025 (LinkedIn). "
        "Co-LinkedIn: linkedin.com/company/home-seekers-sa"
    ),
}

subject = "416 listings and buyer inquiries waiting while agents are at viewings"

body = """\
Hi Thomas,

Saw the LinkedIn post on Rode's October run — 8 deals for R11.44M on a 100%-commission model. That number makes the platform's value proposition more credible than any recruitment copy.

The question it raised for me: when a buyer submits an inquiry on one of your 416 listings at 9pm on a Sunday, what happens to that lead before Monday morning? In South Africa's market, that buyer has already sent the same inquiry to three other agents on Property24. The one who responds on WhatsApp first — even if it's an AI — controls the conversation.

Genos AI builds WhatsApp automation specifically for real estate platforms. For HomeSeeker's model:

WhatsApp AI for buyer inquiries — responds 24/7, answers property questions, qualifies budget and timeline, books the viewing without the agent being online. Lead goes from Property24 to confirmed appointment automatically.

Lead routing by area and type — inquiry captured via WhatsApp, AI qualifies the buyer, assigns to the right agent based on property type and location, logged in your system.

FICA and compliance document collection — for your back-office platform, an AI workflow that requests, follows up, and confirms receipt of required documents from buyers and sellers. Agents stop chasing paper manually.

Agent recruitment intake — agents.homeseeker.co.za gets an AI that screens applicants, asks qualifying questions, and schedules intro calls. You're not reading every enquiry manually.

I'm Rohan Malik, CEO of Genos AI. We build AI automation for real estate operations — WhatsApp agents, lead qualification flows, compliance automations. Happy to show you how it works in 20 minutes.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi Thomas, following up on my last message. The core gap — buyer inquiries on your 416 listings sitting unanswered while agents are at viewings — is the fastest to close. A WhatsApp AI agent for Home Seekers is typically live in under 2 weeks. Happy to walk through it in 20 minutes. — Rohan | Genos AI\
"""

fu2 = """\
Thomas, last one from me. On a 100%-commission platform, an agent's income depends entirely on their lead conversion speed. An AI that ensures no buyer inquiry goes unanswered — even at 9pm on a Sunday — is the infrastructure that makes your agents more productive than any franchise competitor. Leaving this here if the timing isn't right. — Rohan | Genos AI\
"""


def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = cp.copy(src_cell.font)
        dst_cell.fill = cp.copy(src_cell.fill)
        dst_cell.border = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


# ── Organic Leads ──────────────────────────────────────────────────────────────
ws = wb["Organic Leads"]
row = 16
# Columns: #, Company, Contact Person, Title, Email, Phone, Website, LinkedIn,
#          City, State/Country, Industry, Source, Status, Audit Score, Deck Path,
#          Automation Opportunities, Notes
organic_vals = [
    lead["num"], lead["company"], lead["full"], lead["title"],
    lead["email"], lead["phone"], lead["website"], lead["linkedin"],
    lead["city"], lead["state"], lead["industry"],
    "Organic/Manual", "Pending Send", lead["score"],
    lead["ppt"], lead["automation"], lead["notes"],
]
for col, v in enumerate(organic_vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# ── Cold Emails ────────────────────────────────────────────────────────────────
ws2 = wb["Cold Emails"]
row2 = 10
# Columns: #, Company, Contact, Email, Subject, Body, Follow-Up 1, Follow-Up 2
email_vals = [
    lead["num"], lead["company"], lead["full"],
    lead["email"], subject, body, fu1, fu2,
]
for col, v in enumerate(email_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# ── Pipeline ───────────────────────────────────────────────────────────────────
ws3 = wb["Pipeline"]
row3 = 10
# Columns: #, Company, Contact, Email, Industry, Audit Score, Status,
#          Last Touchpoint, Next Action, Est. Deal Value
pipeline_vals = [
    lead["num"], lead["company"], lead["full"],
    lead["email"], lead["industry"], lead["score"],
    "Email Pending", "2026-05-17", "Send cold email — pitch AI automation",
    "",
]
for col, v in enumerate(pipeline_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

wb.save(WB_PATH)
print("OrganicLeads.xlsx updated — Thomas Potgieter / Home Seekers added to all 3 sheets.")
