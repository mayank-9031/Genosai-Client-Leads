from pathlib import Path
import copy as cp
from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent.parent
WB_PATH = ROOT / "output" / "workbooks" / "GenosAI_OrganicLeads.xlsx"

print(f"Loading: {WB_PATH}")
wb = load_workbook(WB_PATH)

lead = {
    "num":   18,
    "first": "John",
    "last":  "Romeo",
    "full":  "John Romeo",
    "title": "Managing Partner & CEO, Oliver Wyman Forum",
    "company": "Oliver Wyman Forum",
    "email": "owforum@oliverwyman.com",
    "email_status": "Official Forum contact email — personal email not publicly listed",
    "website":  "https://www.oliverwymanforum.com/",
    "phone":    "N/A",
    "city":     "New York",
    "state":    "New York, USA",
    "industry": "Management Consulting / Think Tank",
    "linkedin": "https://www.linkedin.com/company/oliverwymanforum/about/",
    "ppt":   "OliverWymanForum_Pitch_GenosAI.pptx",
    "score": 7.5,
    "automation": (
        "AI-powered executive community management (invitation, onboarding, engagement sequences for "
        "CEO/CFO Leadership Reimagined community); research distribution automation with audience "
        "segmentation for reports and Uncharted weekly insights; event and convening coordination AI "
        "(registration, reminders, post-event follow-ups, speaker management); member qualification "
        "flow for Leadership Reimagined applicants; CRM automation for C-suite relationship touchpoints; "
        "content distribution AI for research publications across LinkedIn and email"
    ),
    "notes": (
        "Think tank arm of Oliver Wyman (part of Marsh McLennan). CEO Agenda 2026 covers 10% of global "
        "market cap. CFO Agenda 2026 covers 12% of global market cap. Leadership Reimagined is exclusive "
        "CEO community. 'Uncharted' is their weekly data-driven insights series. "
        "Global Consumer Sentiment Survey reaches 300K+ respondents. "
        "Ice-breaker: CEO Agenda 2026 scale — managing relationships at 10% of global market cap is "
        "a community-management infrastructure problem. Personal LinkedIn not found — used Forum activity."
    ),
}

subject = "Leadership Reimagined runs on relationships — the follow-through infrastructure could be automated"

body = """\
Hi John,

The CEO Agenda 2026 covering leaders at 10% of global market capitalization is an extraordinary convening mandate. The operational question it raises: maintaining meaningful, consistent touchpoints with that many C-suite relationships — invitations, pre-event briefings, post-event follow-up, research delivery, renewal — is a community-management infrastructure problem at scale. Right now that follow-through depends on who is at their desk.

Genos AI builds the automation layer for organizations that run on executive relationships. For the Oliver Wyman Forum specifically:

Executive community management automation — personalized outreach sequences for CEO and CFO community members. Event invitations, pre-reading delivery, post-convening follow-up, and renewal outreach — all triggered by calendar and behavior, not by someone tracking a spreadsheet.

Research distribution and segmentation — automated delivery of Uncharted insights and Forum reports to audience segments based on stated interests and engagement history. Reports that executives subscribed to but never received a coherent follow-up on.

Event and convening coordination — automated registration flows, speaker briefing distribution, attendee reminder sequences, and post-event summary delivery. Removes the admin overhead from every convening without reducing quality.

Member qualification for Leadership Reimagined — when executives express interest in the community, an automated flow qualifies their profile, routes to the right engagement track, and ensures no application sits cold.

I'm Rohan Malik, CEO of Genos AI. We build AI automation for research-led organizations and executive communities. Happy to show you how it maps to the Forum's workflow in a 15-minute call.

Rohan Malik
CEO, Genos AI
+91 638-714-2699  |  hello@genosai.tech  |  www.genosai.tech\
"""

fu1 = """\
Hi John, following up on my last message. The most immediate value — automating the follow-up and engagement sequences around your executive convenings — is typically live within 2 weeks. Happy to walk through it in 15 minutes. — Rohan | Genos AI\
"""

fu2 = """\
John, last one from me. An organization running CEO and CFO agenda programs at the scale of global market capitalization needs the operational infrastructure to match. AI automation for community management, research distribution, and event follow-through is exactly that infrastructure. Leaving this here if the timing isn't right. — Rohan | Genos AI\
"""


def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font      = cp.copy(src_cell.font)
        dst_cell.fill      = cp.copy(src_cell.fill)
        dst_cell.border    = cp.copy(src_cell.border)
        dst_cell.alignment = cp.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


# ── Organic Leads ──────────────────────────────────────────────────────────────
ws  = wb["Organic Leads"]
row = ws.max_row + 1
organic_vals = [
    lead["num"], lead["company"], lead["full"], lead["title"],
    lead["email"], lead["phone"], lead["website"], lead["linkedin"],
    lead["city"], lead["state"], lead["industry"],
    "Organic/Manual", "Pending Send", lead["score"],
    lead["ppt"], lead["automation"], lead["notes"],
]
for col, v in enumerate(organic_vals, 1):
    cell = ws.cell(row, col)
    cell.value = v
    copy_style(ws.cell(row - 1, col), cell)

# ── Cold Emails ────────────────────────────────────────────────────────────────
ws2  = wb["Cold Emails"]
row2 = ws2.max_row + 1
email_vals = [
    lead["num"], lead["company"], lead["full"],
    lead["email"], subject, body, fu1, fu2,
]
for col, v in enumerate(email_vals, 1):
    cell = ws2.cell(row2, col)
    cell.value = v
    copy_style(ws2.cell(row2 - 1, col), cell)

# ── Pipeline ───────────────────────────────────────────────────────────────────
ws3  = wb["Pipeline"]
row3 = ws3.max_row + 1
pipeline_vals = [
    lead["num"], lead["company"], lead["full"],
    lead["email"], lead["industry"], lead["score"],
    "Email Pending", "2026-05-17", "Send cold email — pitch AI automation",
    "",
]
for col, v in enumerate(pipeline_vals, 1):
    cell = ws3.cell(row3, col)
    cell.value = v
    copy_style(ws3.cell(row3 - 1, col), cell)

wb.save(WB_PATH)
print(f"OrganicLeads.xlsx updated — John Romeo / Oliver Wyman Forum added to all 3 sheets (row {row}).")
